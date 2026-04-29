import openpyxl

wb = openpyxl.load_workbook('espelho_final_opcoes.xlsx')
ws = wb.active

print('✅ OPÇÕES DO EXCEL CONFIGURADAS')
print()
print('Configurações Aplicadas:')
print(f'  Date 1904: {wb.date1904}')
print(f'  Title: {wb.title}')

print()
print('Estrutura do Arquivo:')
print(f'  Worksheets: {len(wb.sheetnames)}')
print(f'  Sheet name: {ws.title}')

print()
print('Verificação de Fórmulas:')
formula_count = 0
formulas = []
for row in ws.iter_rows():
    for cell in row:
        if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
            formula_count += 1
            if len(formulas) < 3:
                formulas.append(f'{cell.coordinate}: {cell.value}')

print(f'  Total de fórmulas: {formula_count}')
for f in formulas:
    print(f'  - {f}')

print()
print('Dados Sample:')
print(f'  A1: {ws["A1"].value}')
print(f'  A10 (Data): {ws["A10"].value}')
print(f'  L10 (H.N.): {ws["L10"].value}')
print(f'  M10 (H.E.): {ws["M10"].value}')

print()
print('✅ ARQUIVO PRONTO COM OPÇÕES AVANÇADAS!')
