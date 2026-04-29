import openpyxl

wb = openpyxl.load_workbook('espelho_formulas.xlsx')
ws = wb.active

print('=== DIA 2 (COM DADOS) - Linha 10 ===')
for col in range(1, 16):
    cell = ws.cell(10, col)
    val = cell.value
    col_letter = chr(64 + col) if col <= 26 else chr(64 + col // 26) + chr(64 + col % 26)
    if isinstance(val, str) and val.startswith('='):
        print(f'Col {col:2d} ({col_letter}): FÓRMULA: {val}')
    else:
        print(f'Col {col:2d} ({col_letter}): {val}')

print()
print('=== DIA 5 (ABONO) - Linha 13 ===')
for col in range(1, 16):
    cell = ws.cell(13, col)
    val = cell.value
    col_letter = chr(64 + col) if col <= 26 else chr(64 + col // 26) + chr(64 + col % 26)
    if isinstance(val, str) and val.startswith('='):
        print(f'Col {col:2d} ({col_letter}): FÓRMULA: {val}')
    else:
        print(f'Col {col:2d} ({col_letter}): {val}')
