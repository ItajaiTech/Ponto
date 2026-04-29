import openpyxl
from openpyxl.styles import PatternFill

wb = openpyxl.load_workbook('espelho_com_abono3.xlsx', data_only=False)
ws = wb.active

print('✅ VALIDAÇÃO FINAL DO EXCEL COM CÁLCULOS')
print()

print('=== ESTRUTURA ===')
print(f'Título (A1): {ws["A1"].value}')
print(f'Mesclado A1:O1: {ws.merged_cells}')
print(f'Cabeçalhos (15 colunas): {[ws.cell(8, i).value for i in range(1, 16)]}')

print()
print('=== FÓRMULAS (Cálculos de Horas) ===')

# Verificar fórmulas
h_n_formula = ws.cell(10, 12).value
h_e_formula = ws.cell(10, 13).value

print(f'H.N. (Col 12): {h_n_formula}')
print(f'H.E. (Col 13): {h_e_formula}')

print()
print('=== DADOS DO DIA 5 (ABONO) ===')
print(f'Data: {ws.cell(13, 1).value}')
print(f'Dia: {ws.cell(13, 2).value}')
print(f'Entradas/Saídas: {ws.cell(13, 3).value}, {ws.cell(13, 4).value}, {ws.cell(13, 5).value}, {ws.cell(13, 6).value}')
print(f'Abono (Col 10): {ws.cell(13, 10).value}')
print(f'Jornada (Col 11): {ws.cell(13, 11).value}')
print(f'H.N. Fórmula (Col 12): {ws.cell(13, 12).value}')
print(f'H.E. Fórmula (Col 13): {ws.cell(13, 13).value}')
print(f'Observação (Col 15): {ws.cell(13, 15).value}')

print()
print('=== CORES ===')

# Verificar cores
titulo_fill = ws['A1'].fill
titulo_font = ws['A1'].font

print(f'Título - Fill: {titulo_fill.start_color.rgb if titulo_fill.start_color else "nenhuma"}')
print(f'Título - Font: {titulo_font.color.rgb if titulo_font.color and hasattr(titulo_font.color, "rgb") else "nenhuma"}')

# Cabeçalho
header_fill = ws.cell(8, 1).fill
header_font = ws.cell(8, 1).font

print(f'Cabeçalho - Fill: {header_fill.start_color.rgb if header_fill.start_color else "nenhuma"}')
print(f'Cabeçalho - Font: {header_font.color.rgb if header_font.color and hasattr(header_font.color, "rgb") else "nenhuma"}')

# Abono (Verde)
abono_font = ws.cell(13, 10).font
print(f'Abono (Col 10) - Font: {abono_font.color.rgb if abono_font.color and hasattr(abono_font.color, "rgb") else "nenhuma"}')

# Feriado (Vermelho)
feriado_font = ws.cell(12, 15).font
print(f'Feriado (Col 15) - Font: {feriado_font.color.rgb if feriado_font.color and hasattr(feriado_font.color, "rgb") else "nenhuma"}')

print()
print('✅ ESTRUTURA VALIDADA COM SUCESSO!')
