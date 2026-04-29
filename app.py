from flask import Flask, request, session, jsonify, render_template_string, redirect
import secrets
app = Flask(__name__)
app.secret_key = secrets.token_hex(32)

# ...existing code...

# Rota para redirecionar o menu 'Inserir Batidas Manuais' para a página correta
@app.route('/admin/inserir_batidas')
def admin_inserir_batidas():
    # Redireciona para a página principal onde está o formulário de inserir batidas manuais
    return redirect('/admin')
from flask import send_from_directory
import os
import logging
# Função utilitária para calcular horas do dia
def calcular_horas_do_dia(registros):
    def _parse_data_hora(data_hora):
        try:
            return datetime.strptime(data_hora, '%Y-%m-%d %H:%M:%S')
        except ValueError:
            try:
                return datetime.strptime(data_hora, '%Y-%m-%dT%H:%M')
            except ValueError:
                return None
    entradas = [_parse_data_hora(r['data_hora']) for r in registros if r['tipo'] == 'entrada']
    saidas = [_parse_data_hora(r['data_hora']) for r in registros if r['tipo'] == 'saida']
    entradas = [e for e in entradas if e]
    saidas = [s for s in saidas if s]
    if entradas and saidas:
        return (max(saidas) - min(entradas)).total_seconds() / 3600
    return 0.0
from flask import Flask, request, session, jsonify, render_template_string, redirect
import logging
from werkzeug.security import generate_password_hash, check_password_hash
from datetime import datetime, timedelta
from feriados import is_feriado
import calendar
import json
import sqlite3
import os
import random
import subprocess
import csv
import io
import urllib.parse
import threading
import secrets

# Configuração do logging para registrar erros em debug.log
logging.basicConfig(
    filename='debug.log',
    level=logging.ERROR,
    format='%(asctime)s %(levelname)s %(name)s %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)

app = Flask(__name__)
app.secret_key = secrets.token_hex(32)
from flask import Flask, request, session, jsonify, render_template_string, redirect
import logging

# Middleware para redirecionar HTTP → HTTPS
@app.before_request
def https_redirect():
    """Redireciona requisições HTTP para HTTPS"""
    if request.scheme == 'http':
        url = request.url.replace('http://', 'https://', 1)
        return redirect(url, code=301)

@app.route('/admin/gerar', methods=['GET', 'POST'])
def admin_gerar():
    if session.get('tipo') != 'admin':
        return redirect('/')

# --- MOVIDO PARA O INÍCIO PARA EVITAR NameError ---
from flask import Flask, request, session, jsonify, render_template_string, redirect
import secrets
app = Flask(__name__)
app.secret_key = secrets.token_hex(32)
# -------------------------------------------------

# Rota para redirecionar o menu 'Inserir Batidas Manuais' para a página correta
@app.route('/admin/inserir_batidas')
def admin_inserir_batidas():
    # Redireciona para a página principal onde está o formulário de inserir batidas manuais
    return redirect('/admin')
from flask import send_from_directory
import os
import logging
# Função utilitária para calcular horas do dia
def calcular_horas_do_dia(registros):
    def _parse_data_hora(data_hora):
        try:
            return datetime.strptime(data_hora, '%Y-%m-%d %H:%M:%S')
        except ValueError:
            try:
                return datetime.strptime(data_hora, '%Y-%m-%dT%H:%M')
            except ValueError:
                return None
    entradas = [_parse_data_hora(r['data_hora']) for r in registros if r['tipo'] == 'entrada']
    saidas = [_parse_data_hora(r['data_hora']) for r in registros if r['tipo'] == 'saida']
    entradas = [e for e in entradas if e]
    saidas = [s for s in saidas if s]
    if entradas and saidas:
        return (max(saidas) - min(entradas)).total_seconds() / 3600
    return 0.0
# ...existing code...

    conn = get_db()
    c = conn.cursor()
    mensagem = None

    # Gerar batidas aleatórias
    if request.method == 'POST' and request.form.get('form_tipo') == 'gerar_batidas':
        mensagem = processar_geracao_batidas(request, conn)

    funcionarios_raw = c.execute(
        'SELECT id, nome, cpf, jornada_diaria FROM usuarios WHERE tipo="funcionario"'
    ).fetchall()
    funcionarios = []
    for f in funcionarios_raw:
        funcionarios.append({
            'id': f['id'] if f['id'] is not None else '',
            'nome': f['nome'] if f['nome'] is not None else '',
            'cpf': f['cpf'] if f['cpf'] is not None else '',
            'jornada_diaria': f['jornada_diaria'] if f['jornada_diaria'] is not None else 8.0
        })
    conn.close()
    # Dados mínimos para o template funcionar sem erro
    labels = []
    horas = []
    filtro_atual = 'mes'
    usuario_id = 'all'
    return render_with_theme(
        ADMIN_GERAR_HTML,
        funcionarios=funcionarios,
        mensagem=mensagem,
        labels=labels,
        horas=horas,
        filtro_atual=filtro_atual,
        usuario_id=usuario_id
    )

# Rota para favicon.ico (deve ser registrada após o app ser criado)
@app.route('/favicon.ico', methods=['GET'])
def favicon():
    caminho = os.path.join('static', 'favicon-clock.svg')
    logging.warning(f"Tentando servir favicon em: {os.path.abspath(caminho)}")
    if not os.path.exists(caminho):
        logging.error("favicon-clock.svg não encontrado!")
        return '', 404
    return send_from_directory('static', 'favicon-clock.svg', mimetype='image/svg+xml')

# Handler para erro 404
@app.errorhandler(404)
def page_not_found(e):
    return render_template_string('<h1>Página não encontrada</h1><p>A URL requisitada não existe.</p>'), 404
@app.errorhandler(Exception)
def handle_exception(e):
    logging.error("Erro não tratado", exc_info=e)
    # Opcional: retornar uma página de erro amigável
    return "Ocorreu um erro interno no servidor. Consulte o arquivo debug.log para mais detalhes.", 500

 
DATABASE = 'ponto.db'

# Criação automática da tabela 'usuarios' se não existir
def criar_tabelas_se_necessario():
    conn = sqlite3.connect(DATABASE)
    c = conn.cursor()
    c.execute('''
        CREATE TABLE IF NOT EXISTS usuarios (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            nome TEXT NOT NULL,
            cpf TEXT NOT NULL UNIQUE,
            senha TEXT NOT NULL,
            tipo TEXT NOT NULL,
            jornada_diaria REAL DEFAULT 8.0,
            matricula TEXT,
            cargo TEXT,
            departamento TEXT,
            pis TEXT
        )
    ''')
    c.execute('''
        CREATE TABLE IF NOT EXISTS registros (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            usuario_id INTEGER NOT NULL,
            tipo TEXT NOT NULL,
            data_hora TEXT NOT NULL,
            justificativa TEXT,
            created_at TEXT DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY(usuario_id) REFERENCES usuarios(id)
        )
    ''')
    c.execute('''
        CREATE INDEX IF NOT EXISTS idx_registros_usuario_data
        ON registros(usuario_id, data_hora)
    ''')
    conn.commit()
    conn.close()

criar_tabelas_se_necessario()

def get_db():
    conn = sqlite3.connect(DATABASE)
    conn.row_factory = sqlite3.Row
    return conn


def formatar_horas(valor):
    """Formata um número decimalem formato HH:MM"""
    horas = int(valor)
    minutos = int(abs(valor - horas) * 60)
    return f"{horas:02d}:{minutos:02d}"


def calcular_total_dia(registros):
    """Calcula o total de horas trabalhadas em um dia"""
    if len(registros) < 2:
        return 0

    registros.sort()
    entrada = datetime.fromisoformat(registros[0])
    saida = datetime.fromisoformat(registros[-1])
    
    total = (saida - entrada).total_seconds() / 3600
    return round(total, 2)


def calcular_banco(total, jornada):
    """Calcula o saldo do banco de horas"""
    return round(total - jornada, 2)


def normalizar_cpf(cpf):
    return "".join(ch for ch in cpf if ch.isdigit())


def append_debug_log(message):
    try:
        timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        with open('debug.log', 'a', encoding='utf-8') as f:
            f.write(f"{timestamp} DEBUG login {message}\n")
    except Exception:
        pass


def cpf_valido(cpf):
    cpf = normalizar_cpf(cpf)
    if len(cpf) != 11 or cpf == cpf[0] * 11:
        return False

    for i in (9, 10):
        soma = sum(int(cpf[num]) * ((i + 1) - num) for num in range(i))
        digito = (soma * 10) % 11
        if digito == 10:
            digito = 0
        if digito != int(cpf[i]):
            return False

    return True


def validar_cpfs_banco():
    conn = get_db()
    rows_raw = conn.execute("SELECT id, nome, cpf FROM usuarios").fetchall()
    rows = [dict(r) for r in rows_raw]
    conn.close()

    invalidos = []
    for r in rows:
        if not cpf_valido(r["cpf"] or ""):
            invalidos.append(r)

    if not invalidos:
        return

    log_path = os.path.join(os.path.dirname(__file__), "cpf_invalidos.log")
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    with open(log_path, "a", encoding="utf-8") as f:
        f.write(f"[{now}] CPFs invalidos encontrados: {len(invalidos)}\n")
        for r in invalidos:
            f.write(f"- id={r['id']} nome={r['nome']} cpf={r['cpf']}\n")


HEAD_TAG = "<head>"

def render_with_theme(template, **context):
    theme_script = (
        '<link rel="icon" type="image/svg+xml" href="/favicon.ico">'
        '<script>(function(){var m=document.cookie.match(/(?:^|; )theme=([^;]+)/);'
        'var t=m?m[1]:"green";document.documentElement.setAttribute("data-theme",t);})();</script>'
    )
    if HEAD_TAG in template:
        template = template.replace(HEAD_TAG, HEAD_TAG + theme_script, 1)
    return render_template_string(template, **context)


 

validar_cpfs_banco()

CADASTRO_HTML = (
    '<!DOCTYPE html><html><head><title>Cadastro</title>'
    '<meta name="viewport" content="width=device-width, initial-scale=1.0">'
    '<link rel="stylesheet" href="https://cdn.jsdelivr.net/npm/bootstrap@5.3.2/dist/css/bootstrap.min.css">'
    '<style>body{background:linear-gradient(135deg,#0f2027,#203a43,#2c5364);display:flex;justify-content:center;align-items:center;min-height:100vh;font-family:Arial;margin:0;padding:20px}'
    '.card{background:white;padding:40px;border-radius:20px;width:400px;box-shadow:0 20px 50px rgba(0,0,0,0.4);text-align:center}'
    'h2{color:#2c5364;margin-bottom:30px}'
    'input,select{width:100%;padding:12px;margin:10px 0;border-radius:10px;border:1px solid #ddd;box-sizing:border-box;font-size:14px}'
    'button{width:100%;padding:12px;margin:15px 0 10px 0;border-radius:10px;background:#2c5364;color:white;border:none;font-weight:bold;cursor:pointer;transition:0.3s;font-size:16px}'
    'button:hover{background:#1a3a4a;transform:scale(1.02)}'
    '.btn-voltar{background:#95a5a6;margin-top:20px}.btn-voltar:hover{background:#7f8c8d}'
    '.mensagem-sucesso{background:#27ae60;color:white;padding:12px;border-radius:10px;margin-bottom:20px}'
    '.mensagem-erro{background:#e74c3c;color:white;padding:12px;border-radius:10px;margin-bottom:20px}'
    '</style></head><body><div class="card"><h2>Criar Conta</h2>'
    '{% if mensagem %}<div class="{% if tipo_mensagem == "sucesso" %}mensagem-sucesso{% else %}mensagem-erro{% endif %}">{{mensagem}}</div>{% endif %}'
    '<form method="POST" onsubmit="return validarFormulario();">'
    '<label>Nome:</label><input id="nome" name="nome" required>'
    '<label>CPF:</label><input id="cpf" name="cpf" maxlength="14" oninput="formatarCpf(this)" required>'
    '<label>Senha:</label><input id="senha" name="senha" type="password" required>'
    '<label>Confirmar Senha:</label><input id="confirmar_senha" name="confirmar_senha" type="password" required>'
    '<label>Jornada (horas):</label><input name="jornada" type="number" value="8" min="1" max="12" required>'
    '<button type="submit">Criar Conta</button></form>'
    '<a href="/"><button type="button" class="btn-voltar">Voltar</button></a></div>'
    '<script>'
    'function validarFormulario(){var cpf=document.getElementById("cpf").value;var senha=document.getElementById("senha").value;var confirmar=document.getElementById("confirmar_senha").value;'
    'var cpfDigits=cpf.replace(/\\D/g,"");'
    'if(cpfDigits.length!=11){alert("CPF deve ter 11 dígitos!");return false}'
    'if(senha.length<6){alert("Senha deve ter 6+ caracteres!");return false}'
    'if(senha!=confirmar){alert("Senhas não conferem!");return false}return true}'
    '{% if tipo_mensagem == "sucesso" %}setTimeout(function(){window.location.href="/"},2000);{% endif %}'
    '</script><script src="/static/cpf.js"></script></body></html>'
)
LOGIN_HTML = '''<!DOCTYPE html>
<html>
<head>
    <title>Login do Sistema</title>
    <style>
        :root {
            --bg-gradient-1: #0f2027;
            --bg-gradient-2: #203a43;
            --bg-gradient-3: #2c5364;
            --btn-color: #2c5364;
            --btn-hover: #1a3a4a;
        }
        
        :root[data-theme="green"] {
            --bg-gradient-1: #0f2027;
            --bg-gradient-2: #203a43;
            --bg-gradient-3: #2c5364;
            --btn-color: #2c5364;
            --btn-hover: #1a3a4a;
        }
        
        :root[data-theme="red"] {
            --bg-gradient-1: #1a0a0f;
            --bg-gradient-2: #3d1a25;
            --bg-gradient-3: #5a2a38;
            --btn-color: #ff6b6b;
            --btn-hover: #ff3838;
        }
        
        :root[data-theme="blue"] {
            --bg-gradient-1: #0a0f27;
            --bg-gradient-2: #1a2a5e;
            --bg-gradient-3: #2a3f7a;
            --btn-color: #4d9fff;
            --btn-hover: #1e5eff;
        }
        
        body {
            background: linear-gradient(135deg, var(--bg-gradient-1), var(--bg-gradient-2), var(--bg-gradient-3));
            display: flex;
            justify-content: center;
            align-items: center;
            min-height: 100vh;
            font-family: Arial;
            margin: 0;
            transition: background 0.5s ease;
        }
        
        .card {
            background: white;
            padding: 40px;
            border-radius: 20px;
            width: 350px;
            box-shadow: 0 20px 50px rgba(0,0,0,0.4);
            text-align: center;
        }
        
        input, button {
            width: 100%;
            padding: 12px;
            margin: 10px 0;
            border-radius: 10px;
            box-sizing: border-box;
        }
        
        button {
            background: var(--btn-color);
            color: white;
            border: none;
            font-weight: bold;
            cursor: pointer;
            transition: 0.3s;
        }
        
        button:hover {
            background: var(--btn-hover);
            transform: scale(1.02);
        }
        
        .btn-cadastro {
            background: #27ae60;
            margin-top: 15px;
        }
        
        .btn-cadastro:hover {
            background: #1e8449;
        }
        
        .separador {
            color: #999;
            margin: 20px 0;
            font-size: 12px;
        }
        
        .theme-selector {
            margin: 20px 0;
            padding: 15px;
            background: #f8f9fa;
            border-radius: 10px;
        }
        
        .theme-selector h4 {
            margin: 0 0 10px 0;
            font-size: 14px;
            color: #666;
        }
        
        .theme-buttons {
            display: flex;
            gap: 10px;
            justify-content: center;
        }
        
        .theme-btn {
            width: 50px;
            height: 50px;
            border: 3px solid #ddd;
            border-radius: 50%;
            cursor: pointer;
            transition: all 0.3s;
            padding: 0;
            margin: 0;
        }
        
        .theme-btn:hover {
            transform: scale(1.1);
            box-shadow: 0 4px 15px rgba(0,0,0,0.3);
        }
        
        .theme-btn.active {
            border-color: #333;
            box-shadow: 0 0 0 3px rgba(0,0,0,0.1);
        }
        
        .theme-btn.green {
            background: linear-gradient(135deg, #0f2027, #203a43, #2c5364);
        }
        
        .theme-btn.red {
            background: linear-gradient(135deg, #1a0a0f, #3d1a25, #5a2a38);
        }
        
        .theme-btn.blue {
            background: linear-gradient(135deg, #0a0f27, #1a2a5e, #2a3f7a);
        }
    </style>
</head>
<body>
    <div class="card">
        <h2>🕐 Sistema de Ponto</h2>
        
        <div class="theme-selector">
            <h4>🎨 Escolha seu tema:</h4>
            <div class="theme-buttons">
                <button type="button" class="theme-btn green" onclick="setTheme('green')" title="Green"></button>
                <button type="button" class="theme-btn red" onclick="setTheme('red')" title="Red"></button>
                <button type="button" class="theme-btn blue" onclick="setTheme('blue')" title="Blue"></button>
            </div>
        </div>
        
        {% if erro %}
        <div style="background:#e74c3c;color:#fff;padding:10px;border-radius:8px;margin-bottom:10px;font-size:13px;">{{ erro }}</div>
        {% endif %}

        <form method="POST">
            <input name="cpf" placeholder="CPF ou usuário" required>
            <input name="senha" type="password" placeholder="Senha" required>
            <button>Entrar</button>
        </form>
        
        <div class="separador">Não tem conta?</div>
        <a href="/cadastro" style="text-decoration:none;">
            <button type="button" class="btn-cadastro">Criar Conta</button>
        </a>
        <script>
        function loadTheme() {
            const cookies = document.cookie.split(';');
            let theme = 'green'; // padrão
            let encontrou = false;
            for (let cookie of cookies) {
                const [name, value] = cookie.trim().split('=');
                if (name === 'theme') {
                    theme = value;
                    encontrou = true;
                    break;
                }
            }
            // Se não houver tema salvo, salva green como padrão
            if (!encontrou) {
                document.cookie = "theme=green; path=/; max-age=31536000";
            }
            document.documentElement.setAttribute('data-theme', theme);
            document.querySelector('.theme-btn.' + theme)?.classList.add('active');
        }
        function setTheme(theme) {
            document.cookie = `theme=${theme}; path=/; max-age=31536000`;
            document.documentElement.setAttribute('data-theme', theme);
            document.querySelectorAll('.theme-btn').forEach(btn => btn.classList.remove('active'));
            document.querySelector('.theme-btn.' + theme)?.classList.add('active');
        }
        loadTheme();
        </script>
    </body>
    </html>'''

ADMIN_HTML = '<!DOCTYPE html><html><head><meta name="viewport" content="width=device-width, initial-scale=1.0"><style>@import url("https://fonts.googleapis.com/css2?family=Inter:wght@300;400;600;700&display=swap");:root{--bg-1:#0f2027;--bg-2:#203a43;--bg-3:#2c5364;--accent-1:#00c6ff;--accent-2:#0072ff}:root[data-theme="green"]{--bg-1:#0f2027;--bg-2:#203a43;--bg-3:#2c5364;--accent-1:#00c6ff;--accent-2:#0072ff}:root[data-theme="red"]{--bg-1:#1a0a0f;--bg-2:#3d1a25;--bg-3:#5a2a38;--accent-1:#ff6b6b;--accent-2:#ff3838}:root[data-theme="blue"]{--bg-1:#0a0f27;--bg-2:#1a2a5e;--bg-3:#2a3f7a;--accent-1:#4d9fff;--accent-2:#1e5eff}body{margin:0;font-family:Inter, sans-serif;background:linear-gradient(135deg,var(--bg-1),var(--bg-2),var(--bg-3));color:#fff;transition:background 0.5s ease}.layout{display:flex;gap:20px;padding:24px}.sidebar{width:220px;background:rgba(255,255,255,0.04);padding:18px;border-radius:12px;height:calc(100vh - 48px);position:sticky;top:24px}.content{flex:1}.card{background:rgba(255,255,255,0.06);padding:16px;border-radius:12px;margin-bottom:16px}.section-head{display:flex;justify-content:space-between;align-items:center;margin-bottom:8px}.btn{padding:10px 14px;border-radius:8px;border:none;cursor:pointer}.btn-primary{background:var(--accent-1);color:#001}.btn-danger{background:#c0392b;color:#fff}.form-row{display:flex;gap:8px}input,select{padding:8px;border-radius:8px;border:none;width:100%}.table{width:100%;border-collapse:collapse}.table td{padding:8px;border-bottom:1px solid rgba(255,255,255,0.03)}.table-wrap{overflow-x:auto}.muted{color:rgba(255,255,255,0.7)}.alert{padding:12px;border-radius:8px;margin-bottom:16px}.alert-error{background:rgba(192,57,43,0.2);border:1px solid rgba(192,57,43,0.5);color:#ff6b6b}.alert-success{background:rgba(46,204,113,0.2);border:1px solid rgba(46,204,113,0.5);color:#2ecc71}@media(max-width:980px){.layout{flex-direction:column}.sidebar{width:100%;height:auto;position:relative}.section-head{flex-direction:column;align-items:flex-start;gap:4px}}@media(max-width:720px){.card{padding:14px}.table{min-width:0}.table tr{display:block;background:rgba(255,255,255,0.04);border:1px solid rgba(255,255,255,0.08);border-radius:12px;margin-bottom:10px;padding:8px}.table tr:first-child{display:none}.table td{display:flex;justify-content:space-between;gap:12px;border:none;padding:6px 4px}.table td::before{content:attr(data-label);color:rgba(255,255,255,0.7);font-weight:600}.table td:last-child{flex-direction:column;align-items:stretch}.form-stack{display:flex;flex-direction:column;gap:8px}.action-stack{display:flex;flex-direction:column;align-items:stretch;gap:8px}}</style></head><body><div class="layout"><aside class="sidebar"><h2 style="margin:0 0 12px 0;background:linear-gradient(135deg,var(--accent-1),var(--accent-2));-webkit-background-clip:text;-webkit-text-fill-color:transparent">⏱ PontoPro</h2><nav><a href="/admin" style="color:#fff;display:block;margin:8px 0">Visão Geral</a><a href="/dashboard" style="color:#fff;display:block;margin:8px 0">📊 Dashboard</a><a href="/admin/empresa" style="color:#fff;display:block;margin:8px 0">🏢 Empresa</a><a href="/admin/funcionarios" style="color:#fff;display:block;margin:10px 0;padding:10px 12px;border-radius:10px;background:linear-gradient(135deg,rgba(0,198,255,0.25),rgba(0,114,255,0.05));border:1px solid rgba(0,198,255,0.35);box-shadow:0 8px 18px rgba(0,198,255,0.18);font-weight:600;letter-spacing:0.2px">👥 Funcionários</a><a href="/admin/gerar" style="color:#fff;display:block;margin:8px 0">⚙️ Gerar Batidas</a><a href="/admin/inserir_batidas" style="color:#fff;display:block;margin:8px 0">📝 Inserir Batidas Manuais</a><a href="/logout" style="color:#fff;display:block;margin:8px 0">Sair</a></nav></aside><main class="content">{% if request.args.get("erro") %}<div class="alert alert-error">❌ {{request.args.get("erro")}}</div>{% endif %}{% if request.args.get("sucesso") %}<div class="alert alert-success">✅ {{request.args.get("sucesso")}}</div>{% endif %}<div class="card"><h2 style="margin:0 0 16px 0;color:var(--accent-1)">Bem-vindo ao PontoPro</h2><p style="color:rgba(255,255,255,0.7)">Selecione uma opção no menu lateral para gerenciar o sistema.</p></div><section class="card"><h4 style="color:var(--accent-1)">Inserir Batida Manual</h4><form method="POST" action="/admin/editar" style="display:flex;gap:8px;flex-wrap:wrap;align-items:center"><select name="usuario_id">{% for f in funcionarios %}<option value="{{f.id}}">{{f.nome}}</option>{% endfor %}</select><input type="datetime-local" name="data_hora" required><select name="tipo"><option value="entrada">Entrada</option><option value="saida">Saída</option></select><button class="btn btn-primary">Salvar</button></form></section><section class="card"><h4 style="color:var(--accent-1)">Ações em Lote</h4><form method="POST" action="/admin/registros/deletar_todos" onsubmit="return confirmarDelecao()"><div style="margin-bottom:10px"><label style="display:block;margin-bottom:6px;color:rgba(255,255,255,0.8)">Digite sua senha para confirmar:</label><input type="password" name="senha_confirmacao" id="senhaConfirmacao" placeholder="Senha do administrador" required style="max-width:300px"></div><button type="submit" class="btn btn-danger">🗑️ Apagar Todas as Batidas</button></form></section></main></div><script>function confirmarDelecao(){var senha=document.getElementById("senhaConfirmacao").value;if(!senha){alert("Por favor, digite sua senha para confirmar.");return false}if(!confirm("⚠️ ATENÇÃO: Esta ação irá apagar TODAS as batidas de TODOS os funcionários.\\n\\nEsta ação é IRREVERSÍVEL!\\n\\nDeseja realmente continuar?")){return false}return true}function loadTheme(){const cookies=document.cookie.split(";");let theme="green";let encontrou=false;for(let cookie of cookies){const[name,value]=cookie.trim().split("=");if(name==="theme"){theme=value;encontrou=true;break}}if(!encontrou){document.cookie="theme=green; path=/; max-age=31536000"}document.documentElement.setAttribute("data-theme",theme)}loadTheme()</script></body></html>'

ADMIN_EMPRESA_HTML = '''<!DOCTYPE html>
<html>
<head>
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <style>
        @import url("https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700;800&display=swap");
        
        :root {
            --bg-1: #0f2027;
            --bg-2: #203a43;
            --bg-3: #2c5364;
            --accent-1: #00c6ff;
            --accent-2: #0072ff;
            --particle-1: rgba(0, 198, 255, 0.15);
            --particle-2: rgba(44, 83, 100, 0.2);
            --particle-3: rgba(32, 58, 67, 0.15);
        }
        
        :root[data-theme="green"] {
            --bg-1: #0f2027;
            --bg-2: #203a43;
            --bg-3: #2c5364;
            --accent-1: #00c6ff;
            --accent-2: #0072ff;
            --particle-1: rgba(0, 198, 255, 0.15);
            --particle-2: rgba(44, 83, 100, 0.2);
            --particle-3: rgba(32, 58, 67, 0.15);
        }
        
        :root[data-theme="red"] {
            --bg-1: #1a0a0f;
            --bg-2: #3d1a25;
            --bg-3: #5a2a38;
            --accent-1: #ff6b6b;
            --accent-2: #ff3838;
            --particle-1: rgba(255, 107, 107, 0.15);
            --particle-2: rgba(90, 42, 56, 0.2);
            --particle-3: rgba(61, 26, 37, 0.15);
        }
        
        :root[data-theme="blue"] {
            --bg-1: #0a0f27;
            --bg-2: #1a2a5e;
            --bg-3: #2a3f7a;
            --accent-1: #4d9fff;
            --accent-2: #1e5eff;
            --particle-1: rgba(77, 159, 255, 0.15);
            --particle-2: rgba(42, 63, 122, 0.2);
            --particle-3: rgba(26, 42, 94, 0.15);
        }
        
        * { box-sizing: border-box; }
        
        body {
            margin: 0;
            font-family: Inter, sans-serif;
            background: linear-gradient(135deg, var(--bg-1), var(--bg-2), var(--bg-3));
            min-height: 100vh;
            color: #fff;
            position: relative;
            overflow-x: hidden;
            transition: background 0.5s ease;
        }
        
        body::before {
            content: "";
            position: fixed;
            width: 100%;
            height: 100%;
            background-image: 
                radial-gradient(circle at 20% 80%, var(--particle-1) 0%, transparent 50%),
                radial-gradient(circle at 80% 20%, var(--particle-2) 0%, transparent 50%),
                radial-gradient(circle at 40% 40%, var(--particle-3) 0%, transparent 50%);
            animation: particleFloat 20s ease-in-out infinite;
            z-index: 0;
        }
        
        @keyframes particleFloat {
            0%, 100% { opacity: 1; transform: translateY(0px); }
            50% { opacity: 0.8; transform: translateY(-20px); }
        }
        
        .layout { display: flex; gap: 24px; padding: 30px; position: relative; z-index: 1; }
        
        .sidebar {
            width: 260px;
            background: linear-gradient(145deg, rgba(255, 255, 255, 0.08), rgba(255, 255, 255, 0.03));
            backdrop-filter: blur(10px);
            padding: 24px;
            border-radius: 20px;
            height: calc(100vh - 60px);
            position: sticky;
            top: 30px;
            border: 1px solid rgba(255, 255, 255, 0.1);
            box-shadow: 0 8px 32px rgba(0, 0, 0, 0.3), inset 0 1px 0 rgba(255, 255, 255, 0.1);
        }
        
        .sidebar h2 {
            margin: 0 0 30px 0;
            font-size: 26px;
            font-weight: 800;
            background: linear-gradient(135deg, #00c6ff, #0072ff);
            -webkit-background-clip: text;
            -webkit-text-fill-color: transparent;
            letter-spacing: -0.5px;
        }
        
        .sidebar nav a {
            display: flex;
            align-items: center;
            gap: 12px;
            color: rgba(255, 255, 255, 0.8);
            text-decoration: none;
            padding: 14px 16px;
            margin: 8px 0;
            border-radius: 12px;
            transition: all 0.3s cubic-bezier(0.4, 0, 0.2, 1);
            font-weight: 500;
            position: relative;
            overflow: hidden;
        }
        
        .sidebar nav a::before {
            content: "";
            position: absolute;
            top: 0;
            left: -100%;
            width: 100%;
            height: 100%;
            background: linear-gradient(90deg, transparent, rgba(255, 255, 255, 0.1), transparent);
            transition: left 0.5s;
        }
        
        .sidebar nav a:hover::before { left: 100%; }
        
        .sidebar nav a:hover {
            color: #fff;
            background: rgba(255, 255, 255, 0.05);
            transform: translateX(5px);
            box-shadow: 0 4px 15px rgba(0, 198, 255, 0.2);
        }
        
        .sidebar nav a.active {
            background: linear-gradient(135deg, rgba(0, 198, 255, 0.3), rgba(0, 114, 255, 0.15));
            border: 1px solid rgba(0, 198, 255, 0.4);
            box-shadow: 0 8px 25px rgba(0, 198, 255, 0.25), inset 0 1px 0 rgba(255, 255, 255, 0.2);
            font-weight: 700;
            color: #fff;
        }
        
        .content { flex: 1; max-width: 1400px; }
        
        .card {
            background: linear-gradient(145deg, rgba(255, 255, 255, 0.1), rgba(255, 255, 255, 0.05));
            backdrop-filter: blur(10px);
            padding: 30px;
            border-radius: 20px;
            margin-bottom: 24px;
            border: 1px solid rgba(255, 255, 255, 0.1);
            box-shadow: 0 8px 32px rgba(0, 0, 0, 0.3), inset 0 1px 0 rgba(255, 255, 255, 0.1);
            animation: fadeInUp 0.5s ease-out;
        }
        
        @keyframes fadeInUp {
            from { opacity: 0; transform: translateY(30px); }
            to { opacity: 1; transform: translateY(0); }
        }
        
        .message-card {
            background: linear-gradient(135deg, rgba(0, 198, 255, 0.2), rgba(0, 114, 255, 0.1));
            padding: 16px 20px;
            border-radius: 12px;
            margin-bottom: 20px;
            border-left: 4px solid #00c6ff;
            color: #fff;
            animation: slideInLeft 0.5s ease-out;
        }
        
        @keyframes slideInLeft {
            from { opacity: 0; transform: translateX(-30px); }
            to { opacity: 1; transform: translateX(0); }
        }
        
        .section-head {
            margin-bottom: 24px;
            padding-bottom: 20px;
            border-bottom: 2px solid rgba(0, 198, 255, 0.3);
        }
        
        .section-head h3 {
            margin: 0;
            font-size: 28px;
            font-weight: 700;
            background: linear-gradient(135deg, #fff, #00c6ff);
            -webkit-background-clip: text;
            -webkit-text-fill-color: transparent;
        }
        
        input, select {
            padding: 12px 16px;
            border-radius: 12px;
            border: 1px solid rgba(255, 255, 255, 0.15);
            background: rgba(255, 255, 255, 0.08);
            color: #fff;
            font-family: Inter, sans-serif;
            font-size: 14px;
            transition: all 0.3s ease;
            width: 100%;
        }
        
        input::placeholder { color: rgba(255, 255, 255, 0.4); }
        
        input:focus, select:focus {
            outline: none;
            border-color: #00c6ff;
            background: rgba(255, 255, 255, 0.12);
            box-shadow: 0 0 0 3px rgba(0, 198, 255, 0.15), 0 4px 20px rgba(0, 198, 255, 0.2);
            transform: translateY(-2px);
        }
        
        .form-grid {
            display: grid;
            grid-template-columns: repeat(auto-fit, minmax(200px, 1fr));
            gap: 12px;
            margin-bottom: 20px;
        }
        
        .btn {
            padding: 12px 24px;
            border-radius: 12px;
            border: none;
            cursor: pointer;
            font-family: Inter, sans-serif;
            font-weight: 600;
            font-size: 14px;
            transition: all 0.3s cubic-bezier(0.4, 0, 0.2, 1);
            position: relative;
            overflow: hidden;
        }
        
        .btn::before {
            content: "";
            position: absolute;
            top: 50%;
            left: 50%;
            width: 0;
            height: 0;
            border-radius: 50%;
            background: rgba(255, 255, 255, 0.3);
            transform: translate(-50%, -50%);
            transition: width 0.6s, height 0.6s;
        }
        
        .btn:hover::before { width: 300px; height: 300px; }
        
        .btn span { position: relative; z-index: 1; }
        
        .btn-primary {
            background: linear-gradient(135deg, #00c6ff, #0072ff);
            color: #fff;
            box-shadow: 0 4px 15px rgba(0, 198, 255, 0.4);
        }
        
        .btn-primary:hover {
            box-shadow: 0 6px 25px rgba(0, 198, 255, 0.6);
            transform: translateY(-2px);
        }
        
        .btn-primary:active {
            transform: translateY(0);
        }
        
        .btn-danger {
            background: linear-gradient(135deg, #e74c3c, #c0392b);
            color: #fff;
            box-shadow: 0 4px 15px rgba(231, 76, 60, 0.4);
        }
        
        .btn-danger:hover {
            box-shadow: 0 6px 25px rgba(231, 76, 60, 0.6);
            transform: translateY(-2px);
        }
        
        .btn-warning {
            background: linear-gradient(135deg, #f39c12, #e67e22);
            color: #fff;
            box-shadow: 0 4px 15px rgba(243, 156, 18, 0.4);
        }
        
        .btn-warning:hover {
            box-shadow: 0 6px 25px rgba(243, 156, 18, 0.6);
            transform: translateY(-2px);
        }
        
        .table-section h4 {
            margin: 0 0 20px 0;
            font-size: 18px;
            font-weight: 600;
            color: #fff;
            display: flex;
            align-items: center;
            gap: 8px;
        }
        
        .table-section h4::before {
            content: "📋";
            font-size: 20px;
        }
        
        .table-wrap {
            overflow-x: auto;
            border-radius: 16px;
            background: rgba(0, 0, 0, 0.2);
            padding: 2px;
        }
        
        .table {
            width: 100%;
            border-collapse: separate;
            border-spacing: 0;
        }
        
        .table thead {
            background: linear-gradient(135deg, rgba(0, 198, 255, 0.3), rgba(0, 114, 255, 0.2));
        }
        
        .table th {
            padding: 16px;
            text-align: left;
            font-weight: 700;
            font-size: 13px;
            text-transform: uppercase;
            letter-spacing: 0.5px;
            color: #00c6ff;
            border-bottom: 2px solid rgba(0, 198, 255, 0.5);
        }
        
        .table th:first-child {
            border-top-left-radius: 14px;
        }
        
        .table th:last-child {
            border-top-right-radius: 14px;
        }
        
        .table tbody tr {
            background: linear-gradient(145deg, rgba(255, 255, 255, 0.05), rgba(255, 255, 255, 0.02));
            transition: all 0.3s ease;
        }
        
        .table tbody tr:hover {
            background: linear-gradient(145deg, rgba(0, 198, 255, 0.15), rgba(0, 114, 255, 0.1));
            transform: scale(1.01);
            box-shadow: 0 4px 20px rgba(0, 198, 255, 0.2);
        }
        
        .table td {
            padding: 16px;
            border-bottom: 1px solid rgba(255, 255, 255, 0.05);
            color: rgba(255, 255, 255, 0.9);
        }
        
        .table tbody tr:last-child td {
            border-bottom: none;
        }
        
        .action-buttons {
            display: flex;
            gap: 8px;
            flex-wrap: wrap;
            align-items: center;
        }
        
        .link-btn {
            padding: 8px 14px;
            border-radius: 8px;
            text-decoration: none;
            font-weight: 600;
            font-size: 13px;
            transition: all 0.3s ease;
            display: inline-block;
        }
        
        .link-btn-primary {
            background: linear-gradient(135deg, rgba(0, 198, 255, 0.2), rgba(0, 114, 255, 0.1));
            color: #00c6ff;
            border: 1px solid rgba(0, 198, 255, 0.3);
        }
        
        .link-btn-primary:hover {
            background: linear-gradient(135deg, rgba(0, 198, 255, 0.3), rgba(0, 114, 255, 0.2));
            box-shadow: 0 4px 15px rgba(0, 198, 255, 0.3);
            transform: translateY(-2px);
        }
        
        .link-btn-warning {
            background: linear-gradient(135deg, rgba(243, 156, 18, 0.2), rgba(230, 126, 34, 0.1));
            color: #f39c12;
            border: 1px solid rgba(243, 156, 18, 0.3);
        }
        
        .link-btn-warning:hover {
            background: linear-gradient(135deg, rgba(243, 156, 18, 0.3), rgba(230, 126, 34, 0.2));
            box-shadow: 0 4px 15px rgba(243, 156, 18, 0.3);
            transform: translateY(-2px);
        }
        
        .jornada-form {
            display: flex;
            gap: 8px;
            align-items: center;
        }
        
        .jornada-form input {
            width: 80px;
            margin: 0;
        }
        
        .jornada-form button {
            padding: 8px 16px;
            font-size: 12px;
        }
        
        .message-card {
            background: linear-gradient(135deg, rgba(0, 198, 255, 0.2), rgba(0, 114, 255, 0.1));
            padding: 16px 20px;
            border-radius: 12px;
            margin-bottom: 20px;
            border-left: 4px solid #00c6ff;
            color: #fff;
            animation: slideInLeft 0.5s ease-out;
        }
        
        @keyframes slideInLeft {
            from { opacity: 0; transform: translateX(-30px); }
            to { opacity: 1; transform: translateX(0); }
        }
        
        /* Responsive */
        @media(max-width: 1024px) {
            .layout { flex-direction: column; }
            .sidebar { width: 100%; height: auto; position: relative; top: 0; }
            .form-row { grid-template-columns: 1fr; }
        }
        
        @media(max-width: 768px) {
            .layout { padding: 20px; }
            .card { padding: 20px; }
            .form-grid { grid-template-columns: 1fr; }
        }
    </style>
</head>
<body>
    <div class="layout">
        <aside class="sidebar">
            <h2>⏱ PontoPro</h2>
            <nav>
                <a href="/admin">📊 Visão Geral</a>
                <a href="/dashboard">📈 Dashboard</a>
                <a href="/admin/empresa" class="active">🏢 Empresa</a>
                <a href="/admin/funcionarios">👥 Funcionários</a>
                <a href="/admin/gerar">⚙️ Gerar Batidas</a>
                <a href="/admin/inserir_batidas">📝 Inserir Batidas Manuais</a>
                <a href="/logout">🚪 Sair</a>
            </nav>
        </aside>
        
        <main class="content">
            {% if mensagem %}
            <div class="message-card">
                {{ mensagem }}
            </div>
            {% endif %}
            
            <section class="card">
                <div class="section-head">
                    <h3>🏢 Cadastro da Empresa</h3>
                </div>
                
                <form method="POST">
                    <input type="hidden" name="form_tipo" value="empresa">
                    
                    <div class="form-section">
                        <h4>📋 Identificação</h4>
                        <div class="form-row">
                            <input name="nome_empresa" placeholder="Nome da Empresa" value="{{empresa.nome if empresa}}" required>
                        </div>
                        <div class="form-row">
                            <input name="cnpj" placeholder="CNPJ (xx.xxx.xxx/xxxx-xx)" value="{{empresa.cnpj if empresa}}" required>
                        </div>
                    </div>
                    
                    <div class="form-section">
                        <h4>🔒 Registros (Portaria 1510)</h4>
                        <div class="form-row">
                            <input name="cei" placeholder="CEI - Código de Estabelecimento INSS" value="{{empresa.cei if empresa and empresa.cei}}">
                        </div>
                        <div class="info-box">
                            💡 CEI é obrigatório para gerar relatórios AFD conforme Portaria 1510
                        </div>
                    </div>
                    
                    <div class="form-section">
                        <h4>📍 Localização</h4>
                        <div class="form-row full">
                            <input name="endereco" placeholder="Endereço completo" value="{{empresa.endereco if empresa}}">
                        </div>
                    </div>
                    
                    <div class="form-section">
                        <h4>⚙️ Configurações</h4>
                        <div class="form-row">
                            <input name="tolerancia" placeholder="Tolerância (minutos)" type="number" min="0" max="60" value="{{empresa.tolerancia_min if empresa and empresa.tolerancia_min is not none else 5}}" required>
                        </div>
                        <div class="form-row">
                            <input name="logo" placeholder="URL do Logo" value="{{empresa.logo if empresa}}">
                        </div>
                    </div>
                    
                    <div style="margin-top: 24px;">
                        <button type="submit" class="btn btn-primary"><span>💾 Salvar Empresa</span></button>
                    </div>
                </form>
            </section>
        </main>
    </div>
    <script>
        function loadTheme() {
            const cookies = document.cookie.split(';');
            let theme = 'green';
            let encontrou = false;
            for (let cookie of cookies) {
                const [name, value] = cookie.trim().split('=');
                if (name === 'theme') {
                    theme = value;
                    encontrou = true;
                    break;
                }
            }
            if (!encontrou) {
                document.cookie = "theme=green; path=/; max-age=31536000";
            }
            document.documentElement.setAttribute('data-theme', theme);
        }
        loadTheme();
        
        // Marcar link ativo no menu
        const links = document.querySelectorAll('.sidebar nav a');
        const currentPath = window.location.pathname;
        
        links.forEach(link => {
            const href = link.getAttribute('href');
            if (href === currentPath) {
                link.classList.add('active');
            } else {
                link.classList.remove('active');
            }
        });
    </script>
</body>
</html>'''

ADMIN_FUNCIONARIOS_HTML = '''<!DOCTYPE html>
<html>
<head>
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <style>
        @import url("https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700;800&display=swap");
        
        :root {
            --bg-1: #0f2027;
            --bg-2: #203a43;
            --bg-3: #2c5364;
            --accent-1: #00c6ff;
            --accent-2: #0072ff;
            --particle-1: rgba(0, 198, 255, 0.15);
            --particle-2: rgba(44, 83, 100, 0.2);
            --particle-3: rgba(32, 58, 67, 0.15);
        }
        
        :root[data-theme="green"] {
            --bg-1: #0f2027;
            --bg-2: #203a43;
            --bg-3: #2c5364;
            --accent-1: #00c6ff;
            --accent-2: #0072ff;
            --particle-1: rgba(0, 198, 255, 0.15);
            --particle-2: rgba(44, 83, 100, 0.2);
            --particle-3: rgba(32, 58, 67, 0.15);
        }
        
        :root[data-theme="red"] {
            --bg-1: #1a0a0f;
            --bg-2: #3d1a25;
            --bg-3: #5a2a38;
            --accent-1: #ff6b6b;
            --accent-2: #ff3838;
            --particle-1: rgba(255, 107, 107, 0.15);
            --particle-2: rgba(90, 42, 56, 0.2);
            --particle-3: rgba(61, 26, 37, 0.15);
        }
        
        :root[data-theme="blue"] {
            --bg-1: #0a0f27;
            --bg-2: #1a2a5e;
            --bg-3: #2a3f7a;
            --accent-1: #4d9fff;
            --accent-2: #1e5eff;
            --particle-1: rgba(77, 159, 255, 0.15);
            --particle-2: rgba(42, 63, 122, 0.2);
            --particle-3: rgba(26, 42, 94, 0.15);
        }
        
        * {
            box-sizing: border-box;
        }
        
        body {
            margin: 0;
            font-family: Inter, sans-serif;
            background: linear-gradient(135deg, var(--bg-1), var(--bg-2), var(--bg-3));
            min-height: 100vh;
            color: #fff;
            position: relative;
            overflow-x: hidden;
            transition: background 0.5s ease;
        }
        
        /* Animated background particles */
        body::before {
            content: "";
            position: fixed;
            width: 100%;
            height: 100%;
            background-image: 
                radial-gradient(circle at 20% 80%, var(--particle-1) 0%, transparent 50%),
                radial-gradient(circle at 80% 20%, var(--particle-2) 0%, transparent 50%),
                radial-gradient(circle at 40% 40%, var(--particle-3) 0%, transparent 50%);
            animation: particleFloat 20s ease-in-out infinite;
            z-index: 0;
        }
        
        @keyframes particleFloat {
            0%, 100% { opacity: 1; transform: translateY(0px); }
            50% { opacity: 0.8; transform: translateY(-20px); }
        }
        
        .layout {
            display: flex;
            gap: 24px;
            padding: 30px;
            position: relative;
            z-index: 1;
        }
        
        .sidebar {
            width: 260px;
            background: linear-gradient(145deg, rgba(255, 255, 255, 0.08), rgba(255, 255, 255, 0.03));
            backdrop-filter: blur(10px);
            padding: 24px;
            border-radius: 20px;
            height: calc(100vh - 60px);
            position: sticky;
            top: 30px;
            border: 1px solid rgba(255, 255, 255, 0.1);
            box-shadow: 0 8px 32px rgba(0, 0, 0, 0.3), 
                        inset 0 1px 0 rgba(255, 255, 255, 0.1);
        }
        
        .sidebar h2 {
            margin: 0 0 30px 0;
            font-size: 26px;
            font-weight: 800;
            background: linear-gradient(135deg, #00c6ff, #0072ff);
            -webkit-background-clip: text;
            -webkit-text-fill-color: transparent;
            text-shadow: 0 0 30px rgba(0, 198, 255, 0.5);
            letter-spacing: -0.5px;
        }
        
        .sidebar nav a {
            display: flex;
            align-items: center;
            gap: 12px;
            color: rgba(255, 255, 255, 0.8);
            text-decoration: none;
            padding: 14px 16px;
            margin: 8px 0;
            border-radius: 12px;
            transition: all 0.3s cubic-bezier(0.4, 0, 0.2, 1);
            font-weight: 500;
            position: relative;
            overflow: hidden;
        }
        
        .sidebar nav a::before {
            content: "";
            position: absolute;
            top: 0;
            left: -100%;
            width: 100%;
            height: 100%;
            background: linear-gradient(90deg, transparent, rgba(255, 255, 255, 0.1), transparent);
            transition: left 0.5s;
        }
        
        .sidebar nav a:hover::before {
            left: 100%;
        }
        
        .sidebar nav a:hover {
            color: #fff;
            background: rgba(255, 255, 255, 0.05);
            transform: translateX(5px);
            box-shadow: 0 4px 15px rgba(0, 198, 255, 0.2);
        }
        
        .sidebar nav a.active {
            background: linear-gradient(135deg, rgba(0, 198, 255, 0.3), rgba(0, 114, 255, 0.15));
            border: 1px solid rgba(0, 198, 255, 0.4);
            box-shadow: 0 8px 25px rgba(0, 198, 255, 0.25),
                        inset 0 1px 0 rgba(255, 255, 255, 0.2);
            font-weight: 700;
            color: #fff;
        }
        
        .content {
            flex: 1;
            max-width: 1400px;
        }
        
        .card {
            background: linear-gradient(145deg, rgba(255, 255, 255, 0.1), rgba(255, 255, 255, 0.05));
            backdrop-filter: blur(10px);
            padding: 30px;
            border-radius: 20px;
            margin-bottom: 24px;
            border: 1px solid rgba(255, 255, 255, 0.1);
            box-shadow: 0 8px 32px rgba(0, 0, 0, 0.3),
                        inset 0 1px 0 rgba(255, 255, 255, 0.1);
            animation: fadeInUp 0.5s ease-out;
        }
        
        @keyframes fadeInUp {
            from { opacity: 0; transform: translateY(30px); }
            to { opacity: 1; transform: translateY(0); }
        }
        
        .section-head {
            margin-bottom: 24px;
            padding-bottom: 20px;
            border-bottom: 2px solid rgba(0, 198, 255, 0.3);
        }
        
        .section-head h3 {
            margin: 0 0 8px 0;
            font-size: 28px;
            font-weight: 700;
            background: linear-gradient(135deg, #fff, #00c6ff);
            -webkit-background-clip: text;
            -webkit-text-fill-color: transparent;
        }
        
        .section-head .subtitle {
            color: rgba(255, 255, 255, 0.6);
            font-size: 14px;
            font-weight: 400;
        }
        
        .form-card {
            background: linear-gradient(145deg, rgba(0, 198, 255, 0.1), rgba(0, 114, 255, 0.05));
            padding: 24px;
            border-radius: 16px;
            margin-bottom: 30px;
            border: 1px solid rgba(0, 198, 255, 0.2);
            box-shadow: 0 4px 20px rgba(0, 198, 255, 0.15);
        }
        
        .form-card h4 {
            margin: 0 0 20px 0;
            font-size: 18px;
            font-weight: 600;
            color: #00c6ff;
            display: flex;
            align-items: center;
            gap: 8px;
        }
        
        .form-card h4::before {
            content: "➕";
            font-size: 20px;
        }
        
        input, select {
            width: 100%;
            padding: 14px 16px;
            border-radius: 12px;
            border: 1px solid rgba(255, 255, 255, 0.15);
            background: rgba(255, 255, 255, 0.08);
            color: #fff;
            font-family: Inter, sans-serif;
            font-size: 14px;
            transition: all 0.3s ease;
            margin-bottom: 12px;
        }
        
        input::placeholder {
            color: rgba(255, 255, 255, 0.4);
        }
        
        input:focus, select:focus {
            outline: none;
            border-color: #00c6ff;
            background: rgba(255, 255, 255, 0.12);
            box-shadow: 0 0 0 3px rgba(0, 198, 255, 0.15),
                        0 4px 20px rgba(0, 198, 255, 0.2);
            transform: translateY(-2px);
        }
        
        .form-row {
            display: grid;
            grid-template-columns: 1fr 1fr;
            gap: 12px;
        }
        
        .btn {
            padding: 12px 24px;
            border-radius: 12px;
            border: none;
            cursor: pointer;
            font-family: Inter, sans-serif;
            font-weight: 600;
            font-size: 14px;
            transition: all 0.3s cubic-bezier(0.4, 0, 0.2, 1);
            position: relative;
            overflow: hidden;
        }
        
        .btn::before {
            content: "";
            position: absolute;
            top: 50%;
            left: 50%;
            width: 0;
            height: 0;
            border-radius: 50%;
            background: rgba(255, 255, 255, 0.3);
            transform: translate(-50%, -50%);
            transition: width 0.6s, height 0.6s;
        }
        
        .btn:hover::before {
            width: 300px;
            height: 300px;
        }
        
        .btn span {
            position: relative;
            z-index: 1;
        }
        
        .btn-primary {
            background: linear-gradient(135deg, #00c6ff, #0072ff);
            color: #fff;
            box-shadow: 0 4px 15px rgba(0, 198, 255, 0.4);
        }
        
        .btn-primary:hover {
            box-shadow: 0 6px 25px rgba(0, 198, 255, 0.6);
            transform: translateY(-2px);
        }
        
        .btn-primary:active {
            transform: translateY(0);
        }
        
        .btn-danger {
            background: linear-gradient(135deg, #e74c3c, #c0392b);
            color: #fff;
            box-shadow: 0 4px 15px rgba(231, 76, 60, 0.4);
        }
        
        .btn-danger:hover {
            box-shadow: 0 6px 25px rgba(231, 76, 60, 0.6);
            transform: translateY(-2px);
        }
        
        .btn-warning {
            background: linear-gradient(135deg, #f39c12, #e67e22);
            color: #fff;
            box-shadow: 0 4px 15px rgba(243, 156, 18, 0.4);
        }
        
        .btn-warning:hover {
            box-shadow: 0 6px 25px rgba(243, 156, 18, 0.6);
            transform: translateY(-2px);
        }
        
        .table-section h4 {
            margin: 0 0 20px 0;
            font-size: 18px;
            font-weight: 600;
            color: #fff;
            display: flex;
            align-items: center;
            gap: 8px;
        }
        
        .table-section h4::before {
            content: "📋";
            font-size: 20px;
        }
        
        .table-wrap {
            overflow-x: auto;
            border-radius: 16px;
            background: rgba(0, 0, 0, 0.2);
            padding: 2px;
        }
        
        .table {
            width: 100%;
            border-collapse: separate;
            border-spacing: 0;
        }
        
        .table thead {
            background: linear-gradient(135deg, rgba(0, 198, 255, 0.3), rgba(0, 114, 255, 0.2));
        }
        
        .table th {
            padding: 16px;
            text-align: left;
            font-weight: 700;
            font-size: 13px;
            text-transform: uppercase;
            letter-spacing: 0.5px;
            color: #00c6ff;
            border-bottom: 2px solid rgba(0, 198, 255, 0.5);
        }
        
        .table th:first-child {
            border-top-left-radius: 14px;
        }
        
        .table th:last-child {
            border-top-right-radius: 14px;
        }
        
        .table tbody tr {
            background: linear-gradient(145deg, rgba(255, 255, 255, 0.05), rgba(255, 255, 255, 0.02));
            transition: all 0.3s ease;
        }
        
        .table tbody tr:hover {
            background: linear-gradient(145deg, rgba(0, 198, 255, 0.15), rgba(0, 114, 255, 0.1));
            transform: scale(1.01);
            box-shadow: 0 4px 20px rgba(0, 198, 255, 0.2);
        }
        
        .table td {
            padding: 16px;
            border-bottom: 1px solid rgba(255, 255, 255, 0.05);
            color: rgba(255, 255, 255, 0.9);
        }
        
        .table tbody tr:last-child td {
            border-bottom: none;
        }
        
        .action-buttons {
            display: flex;
            gap: 8px;
            flex-wrap: wrap;
            align-items: center;
        }
        
        .link-btn {
            padding: 8px 14px;
            border-radius: 8px;
            text-decoration: none;
            font-weight: 600;
            font-size: 13px;
            transition: all 0.3s ease;
            display: inline-block;
        }
        
        .link-btn-primary {
            background: linear-gradient(135deg, rgba(0, 198, 255, 0.2), rgba(0, 114, 255, 0.1));
            color: #00c6ff;
            border: 1px solid rgba(0, 198, 255, 0.3);
        }
        
        .link-btn-primary:hover {
            background: linear-gradient(135deg, rgba(0, 198, 255, 0.3), rgba(0, 114, 255, 0.2));
            box-shadow: 0 4px 15px rgba(0, 198, 255, 0.3);
            transform: translateY(-2px);
        }
        
        .link-btn-warning {
            background: linear-gradient(135deg, rgba(243, 156, 18, 0.2), rgba(230, 126, 34, 0.1));
            color: #f39c12;
            border: 1px solid rgba(243, 156, 18, 0.3);
        }
        
        .link-btn-warning:hover {
            background: linear-gradient(135deg, rgba(243, 156, 18, 0.3), rgba(230, 126, 34, 0.2));
            box-shadow: 0 4px 15px rgba(243, 156, 18, 0.3);
            transform: translateY(-2px);
        }
        
        .jornada-form {
            display: flex;
            gap: 8px;
            align-items: center;
        }
        
        .jornada-form input {
            width: 80px;
            margin: 0;
        }
        
        .jornada-form button {
            padding: 8px 16px;
            font-size: 12px;
        }
        
        .message-card {
            background: linear-gradient(135deg, rgba(0, 198, 255, 0.2), rgba(0, 114, 255, 0.1));
            padding: 16px 20px;
            border-radius: 12px;
            margin-bottom: 20px;
            border-left: 4px solid #00c6ff;
            color: #fff;
            animation: slideInLeft 0.5s ease-out;
        }
        
        @keyframes slideInLeft {
            from { opacity: 0; transform: translateX(-30px); }
            to { opacity: 1; transform: translateX(0); }
        }
        
        /* Responsive */
        @media(max-width: 1024px) {
            .layout { flex-direction: column; }
            .sidebar { width: 100%; height: auto; position: relative; top: 0; }
            .form-row { grid-template-columns: 1fr; }
        }
        
        @media(max-width: 768px) {
            .layout { padding: 20px; }
            .card { padding: 20px; }
            
            .table-wrap {
                border-radius: 12px;
            }
            
            .table {
                font-size: 13px;
            }
            
            .table th, .table td {
                padding: 12px 8px;
            }
            
            .action-buttons {
                flex-direction: column;
                width: 100%;
            }
            
            .action-buttons .link-btn,
            .action-buttons form {
                width: 100%;
            }
            
            .action-buttons button {
                width: 100%;
            }
            
            .jornada-form {
                flex-direction: column;
                width: 100%;
            }
            
            .jornada-form input {
                width: 100%;
            }
            
            .jornada-form button {
                width: 100%;
            }
        }
    </style>
</head>
<body>
    <div class="layout">
        <aside class="sidebar">
            <h2>⏱ PontoPro</h2>
            <nav>
                <a href="/admin">📊 Visão Geral</a>
                <a href="/dashboard">📈 Dashboard</a>
                <a href="/admin/empresa">🏢 Empresa</a>
                <a href="/admin/funcionarios">👥 Funcionários</a>
                <a href="/admin/gerar">⚙️ Gerar Batidas</a>
                <a href="/admin/inserir_batidas">📝 Inserir Batidas Manuais</a>
                <a href="/logout">🚪 Sair</a>
            </nav>
        </aside>
        
        <main class="content">
            {% if mensagem %}
            <div class="message-card">
                {{ mensagem }}
            </div>
            {% endif %}
            
            <section class="card">
                <div class="section-head">
                    <h3>👥 Gestão de Funcionários</h3>
                    <div class="subtitle">Cadastre e gerencie funcionários e suas jornadas de trabalho</div>
                </div>
                
                <div class="form-card">
                    <h4>Cadastrar Novo Funcionário</h4>
                    <form method="POST">
                        <input type="hidden" name="form_tipo" value="funcionario">
                        
                        <input name="nome" placeholder="Nome Completo" required>
                        <input name="cpf" placeholder="CPF" maxlength="14" oninput="formatarCpf(this)" required>
                        <input name="pis" placeholder="PIS (Portaria 1510)">
                        <input name="senha" placeholder="Senha de Acesso" required>
                        
                        <div class="form-row">
                            <input name="jornada" placeholder="Jornada Diária (ex: 8)" required>
                            <input name="matricula" placeholder="Matrícula">
                        </div>
                        
                        <div class="form-row">
                            <input name="cargo" placeholder="Cargo">
                            <input name="departamento" placeholder="Departamento">
                        </div>
                        
                        <div style="margin-top: 8px;">
                            <button class="btn btn-primary"><span>✓ Cadastrar Funcionário</span></button>
                        </div>
                    </form>
                </div>
                
                <div class="table-section">
                    <h4>Funcionários Cadastrados</h4>
                    <div class="table-wrap">
                        <table class="table">
                            <thead>
                                <tr>
                                    <th>Nome</th>
                                    <th>CPF</th>
                                    <th>Jornada Diária</th>
                                    <th>Ações</th>
                                </tr>
                            </thead>
                            <tbody>
                                {% for f in funcionarios %}
                                <tr>
                                    <td data-label="Nome">{{ f.nome }}</td>
                                    <td data-label="CPF">{{ f.cpf }}</td>
                                    <td data-label="Jornada">
                                        <form method="POST" action="/admin/usuario/editar" class="jornada-form">
                                            <input type="hidden" name="usuario_id" value="{{ f.id }}">
                                            <input name="jornada" type="number" min="1" max="12" step="0.5" value="{{ f.jornada_diaria }}">
                                            <button class="btn btn-primary"><span>Atualizar</span></button>
                                        </form>
                                    </td>
                                    <td data-label="Ações">
                                        <div class="action-buttons">
                                            <a href="/admin/espelho?usuario_id={{ f.id }}" class="link-btn link-btn-primary">📄 Ver Espelho</a>
                                            <a href="/admin/usuario/editar_form?usuario_id={{ f.id }}" class="link-btn link-btn-warning">✏️ Editar</a>
                                            <form method="POST" action="/admin/registros/deletar_usuario" onsubmit="return confirm(\'Confirma apagar todas as batidas deste funcionário?\');" style="display:inline-block;">
                                                <input type="hidden" name="usuario_id" value="{{ f.id }}">
                                                <button class="btn btn-danger" type="submit"><span>🗑️ Apagar Batidas</span></button>
                                            </form>
                                            <form method="POST" action="/admin/usuario/deletar" onsubmit="return confirm(\'Confirma excluir este usuário? Esta ação removerá o usuário e todas as batidas.\');" style="display:inline-block;">
                                                <input type="hidden" name="usuario_id" value="{{ f.id }}">
                                                <button class="btn btn-danger" type="submit"><span>❌ Excluir Usuário</span></button>
                                            </form>
                                        </div>
                                    </td>
                                </tr>
                                {% endfor %}
                            </tbody>
                        </table>
                    </div>
                </div>
            </section>
        </main>
    </div>
    
    <script src="/static/cpf.js"></script>
    <script>
        function loadTheme() {
            const cookies = document.cookie.split(';');
            let theme = 'green';
            let encontrou = false;
            for (let cookie of cookies) {
                const [name, value] = cookie.trim().split('=');
                if (name === 'theme') {
                    theme = value;
                    encontrou = true;
                    break;
                }
            }
            if (!encontrou) {
                document.cookie = "theme=green; path=/; max-age=31536000";
            }
            document.documentElement.setAttribute('data-theme', theme);
        }
        loadTheme();
    </script>
</body>
</html>'''

ADMIN_GERAR_HTML = '''<!DOCTYPE html>
<html>
<head>
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <link rel="stylesheet" href="/static/themes.css">
    <style>
        @import url("https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700;800&display=swap");
        
        :root {
            --bg-1: #0f2027;
            --bg-2: #203a43;
            --bg-3: #2c5364;
            --accent-1: #00c6ff;
            --accent-2: #0072ff;
            --particle-1: rgba(0, 198, 255, 0.15);
            --particle-2: rgba(44, 83, 100, 0.2);
            --particle-3: rgba(32, 58, 67, 0.15);
        }
        
        :root[data-theme="green"] {
            --bg-1: #0f2027;
            --bg-2: #203a43;
            --bg-3: #2c5364;
            --accent-1: #00c6ff;
            --accent-2: #0072ff;
            --particle-1: rgba(0, 198, 255, 0.15);
            --particle-2: rgba(44, 83, 100, 0.2);
            --particle-3: rgba(32, 58, 67, 0.15);
        }
        
        :root[data-theme="red"] {
            --bg-1: #1a0a0f;
            --bg-2: #3d1a25;
            --bg-3: #5a2a38;
            --accent-1: #ff6b6b;
            --accent-2: #ff3838;
            --particle-1: rgba(255, 107, 107, 0.15);
            --particle-2: rgba(90, 42, 56, 0.2);
            --particle-3: rgba(61, 26, 37, 0.15);
        }
        
        :root[data-theme="blue"] {
            --bg-1: #0a0f27;
            --bg-2: #1a2a5e;
            --bg-3: #2a3f7a;
            --accent-1: #4d9fff;
            --accent-2: #1e5eff;
            --particle-1: rgba(77, 159, 255, 0.15);
            --particle-2: rgba(42, 63, 122, 0.2);
            --particle-3: rgba(26, 42, 94, 0.15);
        }
        
        * { box-sizing: border-box; }
        
        body {
            margin: 0;
            font-family: Inter, sans-serif;
            background: linear-gradient(135deg, var(--bg-1), var(--bg-2), var(--bg-3));
            min-height: 100vh;
            color: #fff;
            position: relative;
            overflow-x: hidden;
            transition: background 0.5s ease;
        }
        
        body::before {
            content: "";
            position: fixed;
            width: 100%;
            height: 100%;
            background-image: 
                radial-gradient(circle at 20% 80%, var(--particle-1) 0%, transparent 50%),
                radial-gradient(circle at 80% 20%, var(--particle-2) 0%, transparent 50%),
                radial-gradient(circle at 40% 40%, var(--particle-3) 0%, transparent 50%);
            animation: particleFloat 20s ease-in-out infinite;
            z-index: 0;
        }
        
        @keyframes particleFloat {
            0%, 100% { opacity: 1; transform: translateY(0px); }
            50% { opacity: 0.8; transform: translateY(-20px); }
        }
        
        .layout { display: flex; gap: 24px; padding: 30px; position: relative; z-index: 1; }
        
        .sidebar {
            width: 260px;
            background: linear-gradient(145deg, rgba(255, 255, 255, 0.08), rgba(255, 255, 255, 0.03));
            backdrop-filter: blur(10px);
            padding: 24px;
            border-radius: 20px;
            height: calc(100vh - 60px);
            position: sticky;
            top: 30px;
            border: 1px solid rgba(255, 255, 255, 0.1);
            box-shadow: 0 8px 32px rgba(0, 0, 0, 0.3), inset 0 1px 0 rgba(255, 255, 255, 0.1);
        }
        
        .sidebar h2 {
            margin: 0 0 30px 0;
            font-size: 26px;
            font-weight: 800;
            background: linear-gradient(135deg, #00c6ff, #0072ff);
            -webkit-background-clip: text;
            -webkit-text-fill-color: transparent;
            letter-spacing: -0.5px;
        }
        
        .sidebar nav a {
            display: flex;
            align-items: center;
            gap: 12px;
            color: rgba(255, 255, 255, 0.8);
            text-decoration: none;
            padding: 14px 16px;
            margin: 8px 0;
            border-radius: 12px;
            transition: all 0.3s cubic-bezier(0.4, 0, 0.2, 1);
            font-weight: 500;
            position: relative;
            overflow: hidden;
        }
        
        .sidebar nav a::before {
            content: "";
            position: absolute;
            top: 0;
            left: -100%;
            width: 100%;
            height: 100%;
            background: linear-gradient(90deg, transparent, rgba(255, 255, 255, 0.1), transparent);
            transition: left 0.5s;
        }
        
        .sidebar nav a:hover::before { left: 100%; }
        
        .sidebar nav a:hover {
            color: #fff;
            background: rgba(255, 255, 255, 0.05);
            transform: translateX(5px);
            box-shadow: 0 4px 15px rgba(0, 198, 255, 0.2);
        }
        
        .sidebar nav a.active {
            background: linear-gradient(135deg, rgba(0, 198, 255, 0.3), rgba(0, 114, 255, 0.15));
            border: 1px solid rgba(0, 198, 255, 0.4);
            box-shadow: 0 8px 25px rgba(0, 198, 255, 0.25), inset 0 1px 0 rgba(255, 255, 255, 0.2);
            font-weight: 700;
            color: #fff;
        }
        
        .content { flex: 1; max-width: 1400px; }
        
        .card {
            background: linear-gradient(145deg, rgba(255, 255, 255, 0.1), rgba(255, 255, 255, 0.05));
            backdrop-filter: blur(10px);
            padding: 30px;
            border-radius: 20px;
            margin-bottom: 24px;
            border: 1px solid rgba(255, 255, 255, 0.1);
            box-shadow: 0 8px 32px rgba(0, 0, 0, 0.3), inset 0 1px 0 rgba(255, 255, 255, 0.1);
            animation: fadeInUp 0.5s ease-out;
        }
        
        @keyframes fadeInUp {
            from { opacity: 0; transform: translateY(30px); }
            to { opacity: 1; transform: translateY(0); }
        }
        
        .header-section {
            display: flex;
            justify-content: space-between;
            align-items: flex-start;
            flex-wrap: wrap;
            gap: 20px;
            margin-bottom: 24px;
            padding-bottom: 20px;
            border-bottom: 2px solid rgba(0, 198, 255, 0.3);
        }
        
        .header-section h2 {
            margin: 0 0 8px 0;
            font-size: 28px;
            font-weight: 700;
            background: linear-gradient(135deg, #fff, #00c6ff);
            -webkit-background-clip: text;
            -webkit-text-fill-color: transparent;
        }
        
        .subtitle {
            color: rgba(255, 255, 255, 0.6);
            font-size: 14px;
            font-weight: 400;
        }
        
        .top-actions {
            display: flex;
            gap: 10px;
            align-items: center;
            flex-wrap: wrap;
        }
        
        .btn-export {
            color: #fff;
            padding: 10px 18px;
            border-radius: 12px;
            text-decoration: none;
            font-size: 14px;
            font-weight: 600;
            transition: all 0.3s cubic-bezier(0.4, 0, 0.2, 1);
            position: relative;
            overflow: hidden;
            display: inline-flex;
            align-items: center;
            gap: 6px;
        }
        
        .btn-export::before {
            content: "";
            position: absolute;
            top: 50%;
            left: 50%;
            width: 0;
            height: 0;
            border-radius: 50%;
            background: rgba(255, 255, 255, 0.3);
            transform: translate(-50%, -50%);
            transition: width 0.6s, height 0.6s;
        }
        
        .btn-export:hover::before {
            width: 300px;
            height: 300px;
        }
        
        .btn-export span {
            position: relative;
            z-index: 1;
        }
        
        .btn-excel {
            background: linear-gradient(135deg, #00cc44, #00aa33);
            box-shadow: 0 4px 15px rgba(0, 204, 68, 0.4);
        }
        
        .btn-excel:hover {
            transform: translateY(-2px);
            box-shadow: 0 6px 25px rgba(0, 204, 68, 0.6);
        }
        
        .btn-afd {
            background: linear-gradient(135deg, #00c6ff, #0072ff);
            box-shadow: 0 4px 15px rgba(0, 198, 255, 0.4);
        }
        
        .btn-afd:hover {
            transform: translateY(-2px);
            box-shadow: 0 6px 25px rgba(0, 198, 255, 0.6);
        }
        
        .btn-txt {
            background: linear-gradient(135deg, #9b59b6, #8e44ad);
            box-shadow: 0 4px 15px rgba(155, 89, 182, 0.4);
        }
        
        .btn-txt:hover {
            transform: translateY(-2px);
            box-shadow: 0 6px 25px rgba(155, 89, 182, 0.6);
        }
        
        .btn-csv {
            background: linear-gradient(135deg, #f39c12, #e67e22);
            box-shadow: 0 4px 15px rgba(243, 156, 18, 0.4);
        }
        
        .btn-csv:hover {
            transform: translateY(-2px);
            box-shadow: 0 6px 25px rgba(243, 156, 18, 0.6);
        }
        
        .filter-form {
            background: linear-gradient(145deg, rgba(0, 198, 255, 0.1), rgba(0, 114, 255, 0.05));
            padding: 20px;
            border-radius: 16px;
            margin-bottom: 24px;
            border: 1px solid rgba(0, 198, 255, 0.2);
            box-shadow: 0 4px 20px rgba(0, 198, 255, 0.15);
            display: flex;
            gap: 12px;
            align-items: center;
            flex-wrap: wrap;
        }
        
        input, select {
            padding: 12px 16px;
            border-radius: 12px;
            border: 1px solid rgba(255, 255, 255, 0.15);
            background: rgba(255, 255, 255, 0.08);
            color: #fff;
            font-family: Inter, sans-serif;
            font-size: 14px;
            transition: all 0.3s ease;
            margin-bottom: 12px;
        }
        
        input::placeholder {
            color: rgba(255, 255, 255, 0.4);
        }
        
        input:focus, select:focus {
            outline: none;
            border-color: #00c6ff;
            background: rgba(255, 255, 255, 0.12);
            box-shadow: 0 0 0 3px rgba(0, 198, 255, 0.15),
                        0 4px 20px rgba(0, 198, 255, 0.2);
            transform: translateY(-2px);
        }
        
        .btn-primary {
            background: linear-gradient(135deg, #00c6ff, #0072ff);
            color: #fff;
            padding: 12px 24px;
            border-radius: 12px;
            border: none;
            cursor: pointer;
            font-family: Inter, sans-serif;
            font-weight: 600;
            font-size: 14px;
            transition: all 0.3s cubic-bezier(0.4, 0, 0.2, 1);
            box-shadow: 0 4px 15px rgba(0, 198, 255, 0.4);
        }
        
        .btn-primary:hover {
            box-shadow: 0 6px 25px rgba(0, 198, 255, 0.6);
            transform: translateY(-2px);
        }
        
        .table-wrap {
            overflow-x: auto;
            border-radius: 16px;
            background: rgba(0, 0, 0, 0.2);
            padding: 2px;
        }
        
        table {
            width: 100%;
            border-collapse: separate;
            border-spacing: 0;
        }
        
        table thead {
            background: linear-gradient(135deg, rgba(0, 198, 255, 0.3), rgba(0, 114, 255, 0.2));
        }
        
        table th {
            padding: 16px;
            text-align: left;
            font-weight: 700;
            font-size: 13px;
            text-transform: uppercase;
            letter-spacing: 0.5px;
            color: #00c6ff;
            border-bottom: 2px solid rgba(0, 198, 255, 0.5);
        }
        
        table th:first-child {
            border-top-left-radius: 14px;
        }
        
        table th:last-child {
            border-top-right-radius: 14px;
        }
        
        table tbody tr {
            background: linear-gradient(145deg, rgba(255, 255, 255, 0.05), rgba(255, 255, 255, 0.02));
            transition: all 0.3s ease;
        }
        
        table tbody tr:hover {
            background: linear-gradient(145deg, rgba(0, 198, 255, 0.15), rgba(0, 114, 255, 0.1));
            transform: scale(1.01);
            box-shadow: 0 4px 20px rgba(0, 198, 255, 0.2);
        }
        
        table td {
            padding: 16px;
            border-bottom: 1px solid rgba(255, 255, 255, 0.05);
            color: rgba(255, 255, 255, 0.9);
        }
        
        table tbody tr:last-child td {
            border-bottom: none;
        }
        
        .edit-form {
            display: flex;
            gap: 8px;
            align-items: center;
            flex-wrap: wrap;
        }
        
        .edit-form input,
        .edit-form select {
            padding: 8px 12px;
            margin: 0;
            font-size: 13px;
        }
        
        .btn-edit {
            background: linear-gradient(135deg, #00c6ff, #0072ff);
            color: #fff;
            padding: 8px 14px;
            border-radius: 8px;
            border: none;
            cursor: pointer;
            font-weight: 600;
            font-size: 12px;
            transition: all 0.3s ease;
            box-shadow: 0 2px 10px rgba(0, 198, 255, 0.3);
        }
        
        .btn-edit:hover {
            box-shadow: 0 4px 15px rgba(0, 198, 255, 0.5);
            transform: translateY(-2px);
        }
        
        .btn-del {
            background: linear-gradient(135deg, #e74c3c, #c0392b);
            color: #fff;
            padding: 8px 14px;
            border-radius: 8px;
            border: none;
            cursor: pointer;
            font-weight: 600;
            font-size: 12px;
            transition: all 0.3s ease;
            box-shadow: 0 2px 10px rgba(231, 76, 60, 0.3);
        }
        
        .btn-del:hover {
            box-shadow: 0 4px 15px rgba(231, 76, 60, 0.5);
            transform: translateY(-2px);
        }
        
        /* Responsive */
        @media(max-width: 1024px) {
            .layout { flex-direction: column; }
            .sidebar { width: 100%; height: auto; position: relative; top: 0; }
            .filter-form { gap: 12px; }
        }
        
        @media(max-width: 768px) {
            .layout { padding: 20px; }
            .card { padding: 20px; }
            .form-grid { grid-template-columns: 1fr; }
            .table-wrap { height: 300px; }
        }
    </style>
</head>
<body>
    <div class="layout">
        <aside class="sidebar">
            <h2>⏱ PontoPro</h2>
            <nav>
                <a href="/admin">📊 Visão Geral</a>
                <a href="/dashboard">📈 Dashboard</a>
                <a href="/admin/empresa">🏢 Empresa</a>
                <a href="/admin/funcionarios">👥 Funcionários</a>
                <a href="/admin/gerar">⚙️ Gerar Batidas</a>
                <a href="/admin/inserir_batidas">📝 Inserir Batidas Manuais</a>
                <a href="/logout">🚪 Sair</a>
            </nav>
        </aside>
        
        <main class="content">
            <div class="card">
                <div class="header-section">
                    <div>
                        <h2>⚙️ Gerar Batidas Aleatórias</h2>
                        <p class="subtitle">Generate time entries with random variations (±5 min)</p>
                    </div>
                </div>

                {% if mensagem %}
                <div style="background: rgba(0, 198, 255, 0.2); border: 1px solid rgba(0, 198, 255, 0.5); padding: 16px; border-radius: 12px; margin-bottom: 20px; color: #00c6ff;">
                    ✅ {{ mensagem }}
                </div>
                {% endif %}

                <form method="POST" style="background: linear-gradient(145deg, rgba(0, 198, 255, 0.1), rgba(0, 114, 255, 0.05)); padding: 24px; border-radius: 16px; border: 1px solid rgba(0, 198, 255, 0.2);">
                    <input type="hidden" name="form_tipo" value="gerar_batidas">
                    
                    <div style="display: grid; grid-template-columns: 1fr 1fr; gap: 16px; margin-bottom: 20px;">
                        <!-- Funcionário -->
                        <div style="display: flex; flex-direction: column;">
                            <label style="font-weight: 600; margin-bottom: 8px; color: rgba(255, 255, 255, 0.9);">👤 Funcionário</label>
                            <select name="usuario_id" required style="padding: 12px; border-radius: 10px; border: 1px solid rgba(0, 198, 255, 0.3); background: rgba(255, 255, 255, 0.08); color: #fff; font-weight: 500;">
                                <option value="all" selected>📋 Todos os Funcionários</option>
                                {% for f in funcionarios %}
                                <option value="{{f.id}}">{{f.nome}}</option>
                                {% endfor %}
                            </select>
                        </div>

                        <!-- Data Inicial -->
                        <div style="display: flex; flex-direction: column;">
                            <label style="font-weight: 600; margin-bottom: 8px; color: rgba(255, 255, 255, 0.9);">📅 Data Inicial</label>
                            <input type="date" name="data_inicial" required style="padding: 12px; border-radius: 10px; border: 1px solid rgba(0, 198, 255, 0.3); background: rgba(255, 255, 255, 0.08); color: #fff;">
                        </div>
                    </div>

                    <div style="display: grid; grid-template-columns: 1fr 1fr; gap: 16px; margin-bottom: 24px;">
                        <!-- Data Final -->
                        <div style="display: flex; flex-direction: column;">
                            <label style="font-weight: 600; margin-bottom: 8px; color: rgba(255, 255, 255, 0.9);">📅 Data Final</label>
                            <input type="date" name="data_final" required style="padding: 12px; border-radius: 10px; border: 1px solid rgba(0, 198, 255, 0.3); background: rgba(255, 255, 255, 0.08); color: #fff;">
                        </div>

                        <!-- Info -->
                        <div style="display: flex; align-items: flex-end;">
                            <div style="background: rgba(0, 198, 255, 0.1); padding: 12px; border-radius: 10px; border-left: 3px solid #00c6ff; font-size: 13px; color: rgba(255, 255, 255, 0.8); width: 100%;">
                                💡 <strong>4 batidas/dia:</strong> 08:00, 12:00, 13:00, 18:00
                            </div>
                        </div>
                    </div>

                    <button type="submit" class="btn-primary" style="width: 100%; padding: 14px; font-size: 16px; font-weight: 700; cursor: pointer; border: none;">
                        🚀 Gerar Batidas Aleatórias
                    </button>
                </form>

                <div style="margin-top: 30px; padding: 20px; background: linear-gradient(145deg, rgba(255, 255, 255, 0.08), rgba(255, 255, 255, 0.03)); border-radius: 16px; border: 1px solid rgba(255, 255, 255, 0.1);">
                    <h3 style="margin-top: 0; color: #00c6ff;">📋 Regras de Geração</h3>
                    <ul style="line-height: 1.8; color: rgba(255, 255, 255, 0.8);">
                        <li>✓ Respeita registros existentes (não sobrescreve)</li>
                        <li>✓ Pula finais de semana e feriados</li>
                        <li>✓ 4 batidas diárias: entrada (08:00), saída (12:00), entrada (13:00), saída (18:00)</li>
                        <li>✓ Variação aleatória de ±5 minutos em cada batida</li>
                        <li>✓ Geração apenas até a data de hoje</li>
                    </ul>
                </div>
            </div>
        </main>
    </div>
    
    <script>
        const ctx = document.getElementById("chartHoras").getContext("2d");
        const gradient = ctx.createLinearGradient(0, 0, 0, 400);
        gradient.addColorStop(0, "rgba(0, 198, 255, 0.6)");
        gradient.addColorStop(1, "rgba(0, 198, 255, 0.05)");
        
        new Chart(ctx, {
            type: "bar",
            data: {
                labels: {{labels|tojson}},
                datasets: [{
                    label: "Horas",
                    data: {{horas|tojson}},
                    backgroundColor: gradient,
                    borderColor: "#00c6ff",
                    borderWidth: 2,
                    borderRadius: 10,
                    borderSkipped: false
                }]
            },
            options: {
                responsive: true,
                maintainAspectRatio: false,
                plugins: {
                    legend: {
                        labels: {
                            color: "#fff",
                            font: {
                                size: 14,
                                weight: 600
                            }
                        }
                    },
                    tooltip: {
                        backgroundColor: "rgba(0, 198, 255, 0.9)",
                        titleColor: "#fff",
                        bodyColor: "#fff",
                        borderColor: "#00c6ff",
                        borderWidth: 1,
                        padding: 12,
                        displayColors: false,
                        callbacks: {
                            label: function(ctx) {
                                return ctx.parsed.y.toFixed(2) + " horas";
                            }
                        }
                    }
                },
                scales: {
                    x: {
                        ticks: {
                            color: "#fff",
                            font: { size: 12 }
                        },
                        grid: {
                            color: "rgba(255, 255, 255, 0.05)",
                            drawBorder: false
                        }
                    },
                    y: {
                        ticks: {
                            color: "#fff",
                            font: { size: 12 },
                            callback: function(value) {
                                return value.toFixed(0) + "h";
                            }
                        },
                        grid: {
                            color: "rgba(255, 255, 255, 0.05)",
                            drawBorder: false
                        }
                    }
                }
            }
        });
    </script>
    <script>
        function loadTheme() {
            const cookies = document.cookie.split(';');
            let theme = 'green';
            let encontrou = false;
            for (let cookie of cookies) {
                const [name, value] = cookie.trim().split('=');
                if (name === 'theme') {
                    theme = value;
                    encontrou = true;
                    break;
                }
            }
            if (!encontrou) {
                document.cookie = "theme=green; path=/; max-age=31536000";
            }
            document.documentElement.setAttribute('data-theme', theme);
        }
        loadTheme();
        
        // Marcar link ativo no menu
        const links = document.querySelectorAll('.sidebar nav a');
        const currentPath = window.location.pathname;
        
        links.forEach(link => {
            const href = link.getAttribute('href');
            if (href === currentPath || (currentPath === '/admin/gerar' && href === '/admin/gerar')) {
                link.classList.add('active');
            } else {
                link.classList.remove('active');
            }
        });
    </script>
</body>
</html>'''

ESPelho_HTML = '''<!DOCTYPE html>
<html>
<head>
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <style>
        @import url("https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700;800&display=swap");
        
        * {
            box-sizing: border-box;
        }
        
        body {
            margin: 0;
            font-family: Inter, sans-serif;
            background: linear-gradient(135deg, #0f2027, #203a43, #2c5364);
            min-height: 100vh;
            color: #fff;
            position: relative;
            overflow-x: hidden;
            transition: background 0.5s ease;
        }
        
        /* Animated background particles */
        body::before {
            content: "";
            position: fixed;
            width: 100%;
            height: 100%;
            background-image: 
                radial-gradient(circle at 20% 80%, rgba(0, 198, 255, 0.15) 0%, transparent 50%),
                radial-gradient(circle at 80% 20%, rgba(44, 83, 100, 0.2) 0%, transparent 50%),
                radial-gradient(circle at 40% 40%, rgba(32, 58, 67, 0.15) 0%, transparent 50%);
            animation: particleFloat 20s ease-in-out infinite;
            z-index: 0;
        }
        
        @keyframes particleFloat {
            0%, 100% { opacity: 1; transform: translateY(0px); }
            50% { opacity: 0.8; transform: translateY(-20px); }
        }
        
        .layout { display: flex; gap: 24px; padding: 30px; position: relative; z-index: 1; }
        
        .sidebar {
            width: 260px;
            background: linear-gradient(145deg, rgba(255, 255, 255, 0.08), rgba(255, 255, 255, 0.03));
            backdrop-filter: blur(10px);
            padding: 24px;
            border-radius: 20px;
            height: calc(100vh - 60px);
            position: sticky;
            top: 30px;
            border: 1px solid rgba(255, 255, 255, 0.1);
            box-shadow: 0 8px 32px rgba(0, 0, 0, 0.3), 
                        inset 0 1px 0 rgba(255, 255, 255, 0.1);
        }
        
        .sidebar h2 {
            margin: 0 0 30px 0;
            font-size: 26px;
            font-weight: 800;
            background: linear-gradient(135deg, #00c6ff, #0072ff);
            -webkit-background-clip: text;
            -webkit-text-fill-color: transparent;
            text-shadow: 0 0 30px rgba(0, 198, 255, 0.5);
            letter-spacing: -0.5px;
        }
        
        .sidebar nav a {
            display: flex;
            align-items: center;
            gap: 12px;
            color: rgba(255, 255, 255, 0.8);
            text-decoration: none;
            padding: 14px 16px;
            margin: 8px 0;
            border-radius: 12px;
            transition: all 0.3s cubic-bezier(0.4, 0, 0.2, 1);
            font-weight: 500;
            position: relative;
            overflow: hidden;
        }
        
        .sidebar nav a::before {
            content: "";
            position: absolute;
            top: 0;
            left: -100%;
            width: 100%;
            height: 100%;
            background: linear-gradient(90deg, transparent, rgba(255, 255, 255, 0.1), transparent);
            transition: left 0.5s;
        }
        
        .sidebar nav a:hover::before { left: 100%; }
        
        .sidebar nav a:hover {
            color: #fff;
            background: rgba(255, 255, 255, 0.05);
            transform: translateX(5px);
            box-shadow: 0 4px 15px rgba(0, 198, 255, 0.2);
        }
        
        .sidebar nav a.active {
            background: linear-gradient(135deg, rgba(0, 198, 255, 0.3), rgba(0, 114, 255, 0.15));
            border: 1px solid rgba(0, 198, 255, 0.4);
            box-shadow: 0 8px 25px rgba(0, 198, 255, 0.25), inset 0 1px 0 rgba(255, 255, 255, 0.2);
            font-weight: 700;
            color: #fff;
        }
        
        .content { flex: 1; max-width: 1400px; }
        
        .card {
            background: linear-gradient(145deg, rgba(255, 255, 255, 0.1), rgba(255, 255, 255, 0.05));
            backdrop-filter: blur(10px);
            padding: 30px;
            border-radius: 20px;
            margin-bottom: 24px;
            border: 1px solid rgba(255, 255, 255, 0.1);
            box-shadow: 0 8px 32px rgba(0, 0, 0, 0.3), inset 0 1px 0 rgba(255, 255, 255, 0.1);
            animation: fadeInUp 0.5s ease-out;
        }
        
        @keyframes fadeInUp {
            from { opacity: 0; transform: translateY(30px); }
            to { opacity: 1; transform: translateY(0); }
        }
        
        .header-section {
            display: flex;
            justify-content: space-between;
            align-items: flex-start;
            flex-wrap: wrap;
            gap: 20px;
            margin-bottom: 24px;
            padding-bottom: 20px;
            border-bottom: 2px solid rgba(0, 198, 255, 0.3);
        }
        
        .header-section h2 {
            margin: 0 0 8px 0;
            font-size: 28px;
            font-weight: 700;
            background: linear-gradient(135deg, #fff, #00c6ff);
            -webkit-background-clip: text;
        }
        
        .btn-del {
            background: linear-gradient(135deg, #e74c3c, #c0392b);
            color: #fff;
            padding: 8px 14px;
            border-radius: 8px;
            border: none;
            cursor: pointer;
            font-weight: 600;
            font-size: 12px;
            transition: all 0.3s ease;
            box-shadow: 0 2px 10px rgba(231, 76, 60, 0.3);
        }
        
        .btn-del:hover {
            box-shadow: 0 4px 15px rgba(231, 76, 60, 0.5);
            transform: translateY(-2px);
        }
        
        /* Responsive */
        @media(max-width: 1024px) {
            .layout {
                flex-direction: column;
            }
            
            .sidebar {
                width: 100%;
                height: auto;
                position: relative;
                top: 0;
            }
            
            .header-section {
                flex-direction: column;
                align-items: flex-start;
            }
            
            .top-actions {
                width: 100%;
            }
        }
        
        @media(max-width: 768px) {
            .layout {
                padding: 20px;
            }
            
            .card {
                padding: 20px;
            }
            
            table {
                font-size: 13px;
            }
            
            table th, table td {
                padding: 12px 8px;
            }
            
            .edit-form {
                flex-direction: column;
                width: 100%;
            }
            
            .edit-form input,
            .edit-form select,
            .edit-form button {
                width: 100%;
            }
        }
    </style>
</head>
<body>
    <div class="layout">
        <aside class="sidebar">
            <h2>⏱ PontoPro</h2>
            <nav>
                <a href="/admin">📊 Visão Geral</a>
                <a href="/dashboard">📈 Dashboard</a>
                <a href="/admin/empresa">🏢 Empresa</a>
                <a href="/admin/funcionarios">👥 Funcionários</a>
                <a href="/admin/gerar">⚙️ Gerar Batidas</a>
                <a href="/logout">🚪 Sair</a>
            </nav>
        </aside>
        
        <main class="content">
            <div class="card">
                <div class="header-section">
                    <div>
                        <h2>📄 Espelho de Ponto — {{usuario.nome}}</h2>
                        <div class="subtitle">Período: {{mes}}/{{ano}}</div>
                    </div>
                    <div class="top-actions">
                        <a href="/admin/espelho/export?usuario_id={{usuario.id}}&tipo=excel&mes={{mes}}&ano={{ano}}" class="btn-export btn-excel">
                            <span>📊 Excel</span>
                        </a>
                        <a href="/admin/espelho/export?usuario_id={{usuario.id}}&tipo=afd&mes={{mes}}&ano={{ano}}" class="btn-export btn-afd">
                            <span>📋 AFD</span>
                        </a>
                        <a href="/admin/espelho/export?usuario_id={{usuario.id}}&tipo=txt&mes={{mes}}&ano={{ano}}" class="btn-export btn-txt">
                            <span>📝 TXT</span>
                        </a>
                        <a href="/admin/espelho/export?usuario_id={{usuario.id}}&tipo=csv&mes={{mes}}&ano={{ano}}" class="btn-export btn-csv">
                            <span>📑 CSV</span>
                        </a>
                    </div>
                </div>
                
                <form method="GET" action="/admin/espelho" class="filter-form">
                    <input type="hidden" name="usuario_id" value="{{usuario.id}}">
                    <input name="start" type="date" placeholder="Data Início">
                    <input name="end" type="date" placeholder="Data Fim">
                    <button class="btn-primary">🔍 Filtrar Período</button>
                </form>
                
                <div class="table-wrap">
                    <table>
                        <thead>
                            <tr>
                                <th>Data e Hora</th>
                                <th>Tipo</th>
                                <th>Justificativa</th>
                                <th>Ações</th>
                            </tr>
                        </thead>
                        <form method="POST" action="/admin/espelho/editar_lote">
                        <input type="hidden" name="usuario_id" value="{{usuario.id}}">
                        <tbody>
                            for r in registros:
                                dia = r['data_hora'][:10]
                                registros_por_dia[dia].append(r)
                                if r['justificativa']:
                                    abonos_por_dia.add(dia)
                            for d in range(1, dias_mes+1):
                                data = datetime(int(ano), int(mes_pad), d)
                                if is_feriado(data):
                                    feriados_por_dia.add(data.strftime('%Y-%m-%d'))

                            # Montar matriz de semanas (4-6 semanas)
                            semanas = []
                            semana = []
                            dia_atual = 1
                            # Preencher dias vazios antes do 1º dia
                            for _ in range(primeiro_dia_semana):
                                semana.append(None)
                            while dia_atual <= dias_mes:
                                data_str = f"{ano}-{mes_pad}-{str(dia_atual).zfill(2)}"
                                semana.append(data_str)
                                    </form>
                                </td>
                            </tr>
                            {% endfor %}
                        </tbody>
                        </table>
                        <div style="margin-top:16px;text-align:right;">
                            <button type="submit" class="btn-edit" style="font-size:1.1em;padding:10px 24px;">💾 Salvar todas alterações</button>
                        </div>
                        </form>
                    </table>
                </div>
            </div>
        </main>
    </div>
    <script>
        function loadTheme() {
            const cookies = document.cookie.split(';');
            let theme = 'green';
            let encontrou = false;
            for (let cookie of cookies) {
                const [name, value] = cookie.trim().split('=');
                if (name === 'theme') {
                    theme = value;
                    encontrou = true;
                    break;
                }
            }
            if (!encontrou) {
                document.cookie = "theme=green; path=/; max-age=31536000";
            }
            document.documentElement.setAttribute('data-theme', theme);
        }
        loadTheme();
        
        // Marcar link ativo no menu
        const links = document.querySelectorAll('.sidebar nav a');
        const currentPath = window.location.pathname;
        
        links.forEach(link => {
            const href = link.getAttribute('href');
            if (href === currentPath || (currentPath === '/admin/espelho' && href === '/admin/espelho')) {
                link.classList.add('active');
            } else {
                link.classList.remove('active');
            }
        });
    </script>
</body>
</html>'''

FUNCIONARIO_PONTO_HTML = '<!DOCTYPE html><html><head><meta name="viewport" content="width=device-width, initial-scale=1.0"><style>@import url("https://fonts.googleapis.com/css2?family=Inter:wght@300;400;600;700&display=swap");:root{--bg-1:#0f2027;--bg-2:#203a43;--bg-3:#2c5364;--card-bg:rgba(255,255,255,0.06);--accent-1:#00c6ff;--accent-2:#0072ff}:root[data-theme="green"]{--bg-1:#0f2027;--bg-2:#203a43;--bg-3:#2c5364;--accent-1:#00c6ff;--accent-2:#0072ff}:root[data-theme="red"]{--bg-1:#1a0a0f;--bg-2:#3d1a25;--bg-3:#5a2a38;--accent-1:#ff6b6b;--accent-2:#ff3838}:root[data-theme="blue"]{--bg-1:#0a0f27;--bg-2:#1a2a5e;--bg-3:#2a3f7a;--accent-1:#4d9fff;--accent-2:#1e5eff}*{box-sizing:border-box}body{margin:0;font-family:"Inter",sans-serif;background:linear-gradient(135deg,var(--bg-1),var(--bg-2),var(--bg-3));color:#fff;transition:background 0.5s ease}.layout{display:flex;gap:24px;padding:24px}.sidebar{width:240px;background:rgba(255,255,255,0.08);backdrop-filter:blur(10px);padding:20px;border-radius:14px;height:calc(100vh - 48px);position:sticky;top:24px;border-left:3px solid var(--accent-1)}.content{flex:1}.card{background:var(--card-bg);backdrop-filter:blur(10px);padding:18px;border-radius:14px;margin-bottom:18px;border:1px solid rgba(255,255,255,0.1)}.header{display:flex;justify-content:space-between;align-items:center;margin-bottom:20px;flex-wrap:wrap;gap:10px}.clock-large{font-size:48px;font-weight:700;color:var(--accent-1);text-align:center;padding:20px;background:var(--card-bg);border-radius:12px;border:2px solid rgba(255,255,255,0.2)}.btn-ponto{flex:1;min-width:150px;padding:16px;border-radius:10px;border:none;font-weight:600;font-size:16px;cursor:pointer;transition:0.3s}.btn-entrada{background:linear-gradient(90deg,var(--accent-2),var(--accent-1));color:white}.btn-entrada:hover{transform:scale(1.05);box-shadow:0 0 20px var(--accent-1)}.btn-saida{background:linear-gradient(90deg,var(--bg-2),var(--accent-2));color:white}.btn-saida:hover{transform:scale(1.05);box-shadow:0 0 20px var(--accent-2)}.ponto-buttons{display:flex;gap:12px;margin-bottom:20px;flex-wrap:wrap}table{width:100%;border-collapse:collapse;margin-top:10px;font-size:14px}td,th{padding:10px;border-bottom:1px solid rgba(255,255,255,0.1);text-align:left}th{background:var(--card-bg);color:var(--accent-1);font-weight:600}tr:hover{background:rgba(255,255,255,0.03)}td{color:rgba(255,255,255,0.9)}.badge-entrada{background:rgba(0,200,100,0.3);color:#00ff88;padding:4px 8px;border-radius:4px}.badge-saida{background:rgba(200,100,0,0.3);color:#ffbb33;padding:4px 8px;border-radius:4px}.info-box{display:grid;grid-template-columns:1fr 1fr;gap:12px;margin-bottom:20px}.info-card{background:var(--card-bg);padding:12px;border-radius:8px;border-left:3px solid var(--accent-1)}.info-label{font-size:12px;color:rgba(255,255,255,0.7)}.info-value{font-size:18px;font-weight:700;color:var(--accent-1)}.toast{position:fixed;bottom:24px;right:24px;background:rgba(0,0,0,0.75);color:#fff;padding:12px 16px;border-radius:10px;border:1px solid rgba(255,255,255,0.15);box-shadow:0 10px 25px rgba(0,0,0,0.4);opacity:0;transform:translateY(8px);transition:opacity 0.25s ease, transform 0.25s ease;z-index:9999;pointer-events:none}.toast.show{opacity:1;transform:translateY(0)}.toast.error{background:rgba(192,57,43,0.9)}@media(max-width:900px){.layout{flex-direction:column}.sidebar{width:100%;height:auto;position:relative}.ponto-buttons{flex-direction:column}.clock-large{font-size:32px}}</style></head><body><div class="layout"><aside class="sidebar"><h2 style="background:linear-gradient(135deg,var(--accent-1),var(--accent-2));-webkit-background-clip:text;-webkit-text-fill-color:transparent">⏱ Ponto</h2><nav style="display:flex;flex-direction:column;gap:8px"><a href="/funcionario" style="display:block;color:#fff;padding:10px;border-radius:8px;background:var(--card-bg);border-left:3px solid var(--accent-1);text-decoration:none">Meu Ponto</a><a href="/logout" style="display:block;color:#fff;padding:10px;border-radius:8px;text-decoration:none">Sair</a></nav></aside><main class="content"><div class="header"><div><h1 style="margin:0;color:var(--accent-1)">Bem-vindo, {{nome}}</h1><div style="color:rgba(255,255,255,0.7)">Relógio de Ponto</div></div></div><div class="clock-large" id="relogio">00:00:00</div><div id="toast" class="toast" role="status" aria-live="polite"></div><div class="ponto-buttons"><button class="btn-ponto btn-entrada" onclick="registrarPonto(\'entrada\')">📥 Registrar Entrada</button><button class="btn-ponto btn-saida" onclick="registrarPonto(\'saida\')">📤 Registrar Saída</button></div><div class="card"><h3 style="margin-top:0;color:var(--accent-1)">Informações Hoje</h3><div class="info-box"><div class="info-card"><div class="info-label">Jornada Diária</div><div class="info-value">{{jornada}}h</div></div><div class="info-card"><div class="info-label">Horas Hoje</div><div class="info-value" id="horasHoje">{{horas_hoje}}</div></div></div></div><div class="card"><h3 style="margin-top:0;color:var(--accent-1)">Batidas de Hoje</h3><table id="tabelaBatidas"><tr><th>Horário</th><th>Tipo</th><th>Hora</th></tr>{% if batidas_hoje %}{% for b in batidas_hoje %}<tr><td>{{b.data}}</td><td>{% if b.tipo == "entrada" %}<span class="badge-entrada">Entrada</span>{% else %}<span class="badge-saida">Saida</span>{% endif %}</td><td>{{b.hora}}</td></tr>{% endfor %}{% else %}<tr><td colspan=3 style="text-align:center;color:rgba(255,255,255,0.7)">Nenhuma batida hoje</td></tr>{% endif %}</table></div><div class="card"><h3 style="margin-top:0;color:var(--accent-1)">Histórico da Semana</h3><table><tr><th>Data</th><th>Entrada 1</th><th>Saída 1</th><th>Entrada 2</th><th>Saída 2</th></tr>{% for dia in dias_semana %}<tr><td>{{dia.data}}</td><td>{{dia.entrada1 or "--"}}</td><td>{{dia.saida1 or "--"}}</td><td>{{dia.entrada2 or "--"}}</td><td>{{dia.saida2 or "--"}}</td></tr>{% endfor %}</table></div></main></div><script>function loadTheme(){const cookies=document.cookie.split(";");let theme="green";let encontrou=false;for(let cookie of cookies){const[name,value]=cookie.trim().split("=");if(name==="theme"){theme=value;encontrou=true;break}}if(!encontrou){document.cookie="theme=green; path=/; max-age=31536000"}document.documentElement.setAttribute("data-theme",theme)}loadTheme();var toastTimer=null;function atualizarRelogio(){const agora=new Date();const h=String(agora.getHours()).padStart(2,"0");const m=String(agora.getMinutes()).padStart(2,"0");const s=String(agora.getSeconds()).padStart(2,"0");document.getElementById("relogio").innerHTML=h+":"+m+":"+s}function showToast(msg,isError){var t=document.getElementById("toast");if(!t){return}t.textContent=msg;t.className="toast show"+(isError?" error":"");if(toastTimer){clearTimeout(toastTimer)}toastTimer=setTimeout(function(){t.className="toast"},2000)}setInterval(atualizarRelogio,1000);atualizarRelogio();function registrarPonto(tipo){fetch("/bater",{method:"POST",headers:{"Content-Type":"application/x-www-form-urlencoded"},body:"tipo="+tipo}).then(r=>r.text()).then(d=>{showToast("Ponto registrado com sucesso!",false);setTimeout(function(){location.reload()},800)}).catch(e=>{showToast("Erro ao registrar ponto",true)})} </script></body></html>'

DASHBOARD_HTML = '''<!DOCTYPE html>
<html>
<head>
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <script src="https://cdn.jsdelivr.net/npm/chart.js@4.3.0/dist/chart.umd.min.js"></script>
    <style>
        @import url("https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700;800&display=swap");
        
        :root {
            --bg-1: #0f2027;
            --bg-2: #203a43;
            --bg-3: #2c5364;
            --accent-1: #00c6ff;
            --accent-2: #0072ff;
            --particle-1: rgba(0, 198, 255, 0.15);
            --particle-2: rgba(44, 83, 100, 0.2);
            --particle-3: rgba(32, 58, 67, 0.15);
        }
        
        :root[data-theme="green"] {
            --bg-1: #0f2027;
            --bg-2: #203a43;
            --bg-3: #2c5364;
            --accent-1: #00c6ff;
            --accent-2: #0072ff;
            --particle-1: rgba(0, 198, 255, 0.15);
            --particle-2: rgba(44, 83, 100, 0.2);
            --particle-3: rgba(32, 58, 67, 0.15);
        }
        
        :root[data-theme="red"] {
            --bg-1: #1a0a0f;
            --bg-2: #3d1a25;
            --bg-3: #5a2a38;
            --accent-1: #ff6b6b;
            --accent-2: #ff3838;
            --particle-1: rgba(255, 107, 107, 0.15);
            --particle-2: rgba(90, 42, 56, 0.2);
            --particle-3: rgba(61, 26, 37, 0.15);
        }
        
        :root[data-theme="blue"] {
            --bg-1: #0a0f27;
            --bg-2: #1a2a5e;
            --bg-3: #2a3f7a;
            --accent-1: #4d9fff;
            --accent-2: #1e5eff;
            --particle-1: rgba(77, 159, 255, 0.15);
            --particle-2: rgba(42, 63, 122, 0.2);
            --particle-3: rgba(26, 42, 94, 0.15);
        }
        
        * { box-sizing: border-box; }
        
        body {
            margin: 0;
            font-family: Inter, sans-serif;
            background: linear-gradient(135deg, var(--bg-1), var(--bg-2), var(--bg-3));
            min-height: 100vh;
            color: #fff;
            position: relative;
            overflow-x: hidden;
            transition: background 0.5s ease;
        }
        
        body::before {
            content: "";
            position: fixed;
            width: 100%;
            height: 100%;
            background-image: 
                radial-gradient(circle at 20% 80%, var(--particle-1) 0%, transparent 50%),
                radial-gradient(circle at 80% 20%, var(--particle-2) 0%, transparent 50%),
                radial-gradient(circle at 40% 40%, var(--particle-3) 0%, transparent 50%);
            animation: particleFloat 20s ease-in-out infinite;
            z-index: 0;
        }
        
        @keyframes particleFloat {
            0%, 100% { opacity: 1; transform: translateY(0px); }
            50% { opacity: 0.8; transform: translateY(-20px); }
        }
        
        .layout { display: flex; gap: 24px; padding: 30px; position: relative; z-index: 1; }
        
        .sidebar {
            width: 260px;
            background: linear-gradient(145deg, rgba(255, 255, 255, 0.08), rgba(255, 255, 255, 0.03));
            backdrop-filter: blur(10px);
            padding: 24px;
            border-radius: 20px;
            height: calc(100vh - 60px);
            position: sticky;
            top: 30px;
            border: 1px solid rgba(255, 255, 255, 0.1);
            box-shadow: 0 8px 32px rgba(0, 0, 0, 0.3), inset 0 1px 0 rgba(255, 255, 255, 0.1);
        }
        
        .sidebar h2 {
            margin: 0 0 30px 0;
            font-size: 26px;
            font-weight: 800;
            background: linear-gradient(135deg, #00c6ff, #0072ff);
            -webkit-background-clip: text;
            -webkit-text-fill-color: transparent;
            letter-spacing: -0.5px;
        }
        
        .sidebar nav a {
            display: flex;
            align-items: center;
            gap: 12px;
            color: rgba(255, 255, 255, 0.8);
            text-decoration: none;
            padding: 14px 16px;
            margin: 8px 0;
            border-radius: 12px;
            transition: all 0.3s cubic-bezier(0.4, 0, 0.2, 1);
            font-weight: 500;
            position: relative;
            overflow: hidden;
        }
        
        .sidebar nav a::before {
            content: "";
            position: absolute;
            top: 0;
            left: -100%;
            width: 100%;
            height: 100%;
            background: linear-gradient(90deg, transparent, rgba(255, 255, 255, 0.1), transparent);
            transition: left 0.5s;
        }
        
        .sidebar nav a:hover::before { left: 100%; }
        
        .sidebar nav a:hover {
            color: #fff;
            background: rgba(255, 255, 255, 0.05);
            transform: translateX(5px);
            box-shadow: 0 4px 15px rgba(0, 198, 255, 0.2);
        }
        
        .sidebar nav a.active {
            background: linear-gradient(135deg, rgba(0, 198, 255, 0.3), rgba(0, 114, 255, 0.15));
            border: 1px solid rgba(0, 198, 255, 0.4);
            box-shadow: 0 8px 25px rgba(0, 198, 255, 0.25), inset 0 1px 0 rgba(255, 255, 255, 0.2);
            font-weight: 700;
            color: #fff;
        }
        
        .content { flex: 1; max-width: 1400px; }
        
        .card {
            background: linear-gradient(145deg, rgba(255, 255, 255, 0.1), rgba(255, 255, 255, 0.05));
            backdrop-filter: blur(10px);
            padding: 30px;
            border-radius: 20px;
            margin-bottom: 24px;
            border: 1px solid rgba(255, 255, 255, 0.1);
            box-shadow: 0 8px 32px rgba(0, 0, 0, 0.3), inset 0 1px 0 rgba(255, 255, 255, 0.1);
            animation: fadeInUp 0.5s ease-out;
        }
        
        @keyframes fadeInUp {
            from { opacity: 0; transform: translateY(30px); }
            to { opacity: 1; transform: translateY(0); }
        }
        
        .card h2 {
            margin: 0 0 20px 0;
            font-size: 28px;
            font-weight: 700;
            background: linear-gradient(135deg, #fff, #00c6ff);
            -webkit-background-clip: text;
            -webkit-text-fill-color: transparent;
        }
        
        .chart-card {
            background: linear-gradient(145deg, rgba(255, 255, 255, 0.1), rgba(255, 255, 255, 0.05));
            backdrop-filter: blur(10px);
            padding: 30px;
            border-radius: 20px;
            border: 1px solid rgba(255, 255, 255, 0.1);
            box-shadow: 0 8px 32px rgba(0, 0, 0, 0.3), inset 0 1px 0 rgba(255, 255, 255, 0.1);
        }
        
        .chart-card h3 {
            margin: 0 0 24px 0;
            font-size: 22px;
            font-weight: 600;
            color: #00c6ff;
        }
        
        .filters {
            display: flex;
            gap: 12px;
            flex-wrap: wrap;
            margin-bottom: 20px;
        }
        
        .filters form {
            display: flex;
            gap: 12px;
            flex-wrap: wrap;
            width: 100%;
        }
        
        select {
            padding: 12px 16px;
            border-radius: 12px;
            border: 1px solid rgba(255, 255, 255, 0.25);
            background: rgba(20, 30, 45, 0.85);
            color: #f8fafc;
            font-family: Inter, sans-serif;
            font-size: 14px;
            transition: all 0.3s ease;
            cursor: pointer;
        }
        
        select option {
            background: #1a2332;
            color: #f8fafc;
        }
        
        select:focus {
            outline: none;
            border-color: #00c6ff;
            background: rgba(20, 30, 45, 0.95);
            box-shadow: 0 0 0 3px rgba(0, 198, 255, 0.15), 0 4px 20px rgba(0, 198, 255, 0.2);
        }
        
        .chart-wrap {
            width: 100%;
            height: 450px;
            position: relative;
        }
        
        canvas {
            width: 100% !important;
            height: 100% !important;
        }
        
        @media(max-width: 1024px) {
            .layout { flex-direction: column; }
            .sidebar { width: 100%; height: auto; position: relative; top: 0; }
            .chart-wrap { height: 350px; }
        }
        
        @media(max-width: 768px) {
            .layout { padding: 20px; }
            .card, .chart-card { padding: 20px; }
            .chart-wrap { height: 300px; }
        }
    </style>
</head>
<body>
    <div class="layout">
        <aside class="sidebar">
            <h2>⏱ PontoPro</h2>
            <nav>
                <a href="/admin">📊 Visão Geral</a>
                <a href="/dashboard" class="active">📈 Dashboard</a>
                <a href="/admin/empresa">🏢 Empresa</a>
                <a href="/admin/funcionarios">👥 Funcionários</a>
                <a href="/admin/gerar">⚙️ Gerar Batidas</a>
                <a href="/logout">🚪 Sair</a>
            </nav>
        </aside>
        
        <main class="content">
            <div class="card">
                <h2>📊 Dashboard de Horas Trabalhadas</h2>
                <div class="filters">
                    <form method="GET">
                        <select name="filtro" onchange="this.form.submit()">
                            <option value="dia" {% if filtro_atual=="dia" %}selected{% endif %}>Por Dia (7 dias)</option>
                            <option value="semana" {% if filtro_atual=="semana" %}selected{% endif %}>Por Semana (4 semanas)</option>
                            <option value="mes" {% if filtro_atual=="mes" %}selected{% endif %}>Por Mês (6 meses)</option>
                            <option value="ano" {% if filtro_atual=="ano" %}selected{% endif %}>Por Ano (3 anos)</option>
                        </select>
                        <select name="usuario_id" onchange="this.form.submit()">
                            <option value="all" {% if usuario_id=="all" %}selected{% endif %}>Todos Funcionários</option>
                            {% for f in funcionarios %}
                            <option value="{{f.id}}" {% if usuario_id==f.id|string %}selected{% endif %}>{{f.nome}}</option>
                            {% endfor %}
                        </select>
                    </form>
                </div>
            </div>
            
            <div class="chart-card">
                <h3>📈 Horas Trabalhadas</h3>
                <div class="chart-wrap">
                    <canvas id="chartHoras"></canvas>
                </div>
            </div>
        </main>
    </div>
    
    <script>
        const ctx = document.getElementById("chartHoras").getContext("2d");
        const gradient = ctx.createLinearGradient(0, 0, 0, 400);
        gradient.addColorStop(0, "rgba(0, 198, 255, 0.6)");
        gradient.addColorStop(1, "rgba(0, 198, 255, 0.05)");
        
        new Chart(ctx, {
            type: "bar",
            data: {
                labels: {{labels|tojson}},
                datasets: [{
                    label: "Horas",
                    data: {{horas|tojson}},
                    backgroundColor: gradient,
                    borderColor: "#00c6ff",
                    borderWidth: 2,
                    borderRadius: 10,
                    borderSkipped: false
                }]
            },
            options: {
                responsive: true,
                maintainAspectRatio: false,
                plugins: {
                    legend: {
                        labels: {
                            color: "#fff",
                            font: {
                                size: 14,
                                weight: 600
                            }
                        }
                    },
                    tooltip: {
                        backgroundColor: "rgba(0, 198, 255, 0.9)",
                        titleColor: "#fff",
                        bodyColor: "#fff",
                        borderColor: "#00c6ff",
                        borderWidth: 1,
                        padding: 12,
                        displayColors: false,
                        callbacks: {
                            label: function(ctx) {
                                return ctx.parsed.y.toFixed(2) + " horas";
                            }
                        }
                    }
                },
                scales: {
                    x: {
                        ticks: {
                            color: "#fff",
                            font: { size: 12 }
                        },
                        grid: {
                            color: "rgba(255, 255, 255, 0.05)",
                            drawBorder: false
                        }
                    },
                    y: {
                        ticks: {
                            color: "#fff",
                            font: { size: 12 },
                            callback: function(value) {
                                return value.toFixed(0) + "h";
                            }
                        },
                        grid: {
                            color: "rgba(255, 255, 255, 0.05)",
                            drawBorder: false
                        }
                    }
                }
            }
        });
    </script>
    <script>
        function loadTheme() {
            const cookies = document.cookie.split(';');
            let theme = 'green';
            let encontrou = false;
            for (let cookie of cookies) {
                const [name, value] = cookie.trim().split('=');
                if (name === 'theme') {
                    theme = value;
                    encontrou = true;
                    break;
                }
            }
            if (!encontrou) {
                document.cookie = "theme=green; path=/; max-age=31536000";
            }
            document.documentElement.setAttribute('data-theme', theme);
        }
        loadTheme();
    </script>
</body>
</html>'''

# ==========================================================
# ROTAS
# ==========================================================

SQL_INSERT_REGISTRO = "INSERT INTO registros (usuario_id, tipo, data_hora) VALUES (?, ?, ?)"
SQL_AND_USUARIO_ID = " AND usuario_id=?"

ADMIN_ROUTE = "/admin"
# Use this constant everywhere instead of hardcoding "/admin"


@app.route("/", methods=["GET", "POST"])
@app.route("/login", methods=["GET", "POST"])
def login():
    erro = ""
    if request.method == "POST":
        cpf = request.form.get("cpf", "").strip()
        senha = request.form.get("senha", "").strip()

        entrada_login = cpf
        cpf_normalizado = normalizar_cpf(entrada_login)
        usuario_txt = entrada_login.strip().lower()
        append_debug_log(f"Tentativa: entrada='{entrada_login}' cpf_normalizado='{cpf_normalizado}'")

        conn = get_db()
        user = conn.execute(
            """
            SELECT * FROM usuarios
            WHERE replace(replace(cpf, '.', ''), '-', '') = ?
               OR lower(trim(nome)) = ?
               OR lower(trim(COALESCE(matricula, ''))) = ?
            LIMIT 1
            """,
            (cpf_normalizado, usuario_txt, usuario_txt)
        ).fetchone()
        conn.close()

        if user:
            append_debug_log(f"Usuário encontrado: id={user['id']} tipo={user['tipo']} cpf={user['cpf']}")
        else:
            append_debug_log("Usuário não encontrado.")

        senha_ok = False
        if user:
            senha_salva = user["senha"] or ""
            try:
                senha_ok = check_password_hash(senha_salva, senha)
            except ValueError:
                senha_ok = False
            if not senha_ok and senha_salva == senha:
                senha_ok = True

        if user and senha_ok:
            porta = request.environ.get('SERVER_PORT')
            append_debug_log(f"Login OK: id={user['id']} tipo={user['tipo']} porta={porta}")
            if porta == '5050' and user["tipo"] != "admin":
                erro = "Acesso negado! Somente administradores podem fazer login pela porta 5050."
                append_debug_log("Acesso negado na porta 5050 para não-admin.")
                return render_with_theme(LOGIN_HTML, erro=erro)

            session["user_id"] = user["id"]
            session["tipo"] = user["tipo"]

            if user["tipo"] == "admin":
                session["admin_redirect_espelho"] = True
                append_debug_log("Redirecionando admin para ADMIN_ROUTE.")
                return redirect(ADMIN_ROUTE)
            else:
                append_debug_log("Redirecionando funcionário para /funcionario.")
                return redirect("/funcionario")
        else:
            erro = "CPF ou senha inválidos"
            append_debug_log(f"Falha login: entrada='{entrada_login}'")

    return render_with_theme(LOGIN_HTML, erro=erro)


# ==========================================================
# CADASTRO
# ==========================================================

@app.route("/cadastro", methods=["GET", "POST"])
def cadastro():
    mensagem = None
    tipo_mensagem = None

    # Permitir cadastro apenas para usuário admin logado
    if session.get("tipo") != "admin":
        mensagem = "Apenas o usuário admin pode adicionar cadastros."
        tipo_mensagem = "erro"
    elif request.method == "POST":
        try:
            conn = get_db()
            c = conn.cursor()
            
            cpf = request.form.get("cpf", "").strip()
            nome = request.form.get("nome", "").strip()
            senha = request.form.get("senha", "").strip()
            confirmar_senha = request.form.get("confirmar_senha", "").strip()
            jornada = request.form.get("jornada", "8").strip()
            cpf_norm = normalizar_cpf(cpf)
            
            # Validações
            if not all([cpf, nome, senha, confirmar_senha]):
                mensagem = "Todos os campos são obrigatórios!"
                tipo_mensagem = "erro"
            elif not cpf_valido(cpf):
                mensagem = "CPF inválido!"
                tipo_mensagem = "erro"
            elif senha != confirmar_senha:
                mensagem = "As senhas não conferem!"
                tipo_mensagem = "erro"
            elif len(senha) < 6:
                mensagem = "Senha deve ter no mínimo 6 caracteres!"
                tipo_mensagem = "erro"
            else:
                # Verificar se CPF já existe
                existe = c.execute(
                    "SELECT id FROM usuarios WHERE cpf=?",
                    (cpf_norm,)
                ).fetchone()
                
                if existe:
                    mensagem = "CPF já cadastrado no sistema!"
                    tipo_mensagem = "erro"
                else:
                    # Cadastrar novo usuário
                    c.execute(
                        "INSERT INTO usuarios (nome, cpf, senha, tipo, jornada_diaria) VALUES (?, ?, ?, 'funcionario', ?)",
                        (
                            nome,
                            cpf_norm,
                            generate_password_hash(senha),
                            float(jornada)
                        )
                    )
                    conn.commit()
                    mensagem = "Cadastro realizado com sucesso! Você será redirecionado para login..."
                    tipo_mensagem = "sucesso"
            
            conn.close()
        except Exception as e:
            mensagem = f"Erro ao cadastrar: {str(e)}"
            tipo_mensagem = "erro"

    return render_with_theme(CADASTRO_HTML, 
        mensagem=mensagem,
        tipo_mensagem=tipo_mensagem
    )


# ==========================================================
# ==========================================================
# FUNCIONÁRIO - RELÓGIO DE PONTO
# ==========================================================

@app.route('/funcionario')
def funcionario():
    if "user_id" not in session:
        return redirect("/")

    conn = get_db()
    user = conn.execute(
        "SELECT * FROM usuarios WHERE id=?",
        (session["user_id"],)
    ).fetchone()

    if not user:
        conn.close()
        return redirect("/")

    hoje = datetime.now().date()
    batidas_hoje = conn.execute(
        f"SELECT * FROM registros WHERE 1=1{SQL_AND_USUARIO_ID} AND date(data_hora)=? ORDER BY data_hora",
        (user["id"], hoje)
    ).fetchall()

    batidas_hoje_data = _processar_batidas_hoje(batidas_hoje, hoje)
    total_horas = calcular_total_dia([r["data_hora"] for r in batidas_hoje])
    horas_hoje = formatar_horas(total_horas)

    dias_semana_data = _processar_dias_semana(conn, user["id"])

    conn.close()

    return render_with_theme(FUNCIONARIO_PONTO_HTML,
        nome=user["nome"],
        jornada=user["jornada_diaria"],
        dias_semana=dias_semana_data,
        batidas_hoje=batidas_hoje_data,
        horas_hoje=horas_hoje
    )

def _processar_batidas_hoje(batidas_hoje, hoje):
    batidas_hoje_data = []
    for reg in batidas_hoje:
        data_hora = reg["data_hora"] if "data_hora" in reg.keys() else None
        hora = "--:--"
        if data_hora and isinstance(data_hora, str) and " " in data_hora:
            partes = data_hora.split(" ")
            if len(partes) > 1 and len(partes[1]) >= 5:
                hora = partes[1][:5]
            else:
                hora = "--:--"
        batidas_hoje_data.append({
            "data": hoje.strftime("%d/%m/%Y"),
            "tipo": reg["tipo"],
            "hora": hora
        })
    return batidas_hoje_data

def _processar_dias_semana(conn, user_id):
    inicio_semana = datetime.now() - timedelta(days=datetime.now().weekday())
    dias_semana_data = []
    dias_pt = ["Seg", "Ter", "Qua", "Qui", "Sex", "Sáb", "Dom"]
    for i in range(7):
        dia = (inicio_semana + timedelta(days=i)).date()
        registros = conn.execute(
            f"SELECT data_hora, tipo FROM registros WHERE 1=1{SQL_AND_USUARIO_ID} AND date(data_hora)=? ORDER BY data_hora",
            (user_id, dia)
        ).fetchall()
        batidas = [{
            "tipo": reg["tipo"],
            "hora": reg["data_hora"].split(" ")[1][:5] if " " in reg["data_hora"] else None
        } for reg in registros]
        entradas = [b["hora"] for b in batidas if b["tipo"] == "entrada"]
        saidas = [b["hora"] for b in batidas if b["tipo"] == "saida"]
        entrada1 = entradas[0] if len(entradas) > 0 else None
        saida1 = saidas[0] if len(saidas) > 0 else None
        entrada2 = entradas[1] if len(entradas) > 1 else None
        saida2 = saidas[1] if len(saidas) > 1 else None
        # Traduzir o dia da semana para português
        dia_semana_idx = dia.weekday()  # 0=Segunda, 6=Domingo
        dia_semana_pt = dias_pt[dia_semana_idx]
        dias_semana_data.append({
            "data": f"{dia_semana_pt}, {dia.strftime('%d/%m')}",
            "entrada1": entrada1,
            "saida1": saida1,
            "entrada2": entrada2,
            "saida2": saida2
        })
    return dias_semana_data


# ==========================================================
# BATER PONTO
# ==========================================================

@app.route("/bater", methods=["POST"])
def bater():
    # permitir funcionário e admin acessarem este endpoint (admin tem acesso a todas as páginas)
    if session.get("tipo") not in ("funcionario", "admin"):
        return redirect("/funcionario")

    tipo = request.form["tipo"]


    conn = get_db()
    agora = datetime.now()
    # Verifica se já existe batida do mesmo tipo no mesmo minuto
    ultimo = conn.execute(
        f"SELECT data_hora, tipo FROM registros WHERE 1=1{SQL_AND_USUARIO_ID} ORDER BY data_hora DESC LIMIT 1",
        (session["user_id"],)
    ).fetchone()
    if ultimo:
        try:
            from datetime import datetime as dt
            dt_ultimo = dt.strptime(ultimo["data_hora"], "%Y-%m-%d %H:%M:%S")
        except Exception:
            dt_ultimo = agora
        # Impede batida do mesmo tipo no mesmo minuto
        if ultimo["tipo"] == tipo and dt_ultimo.strftime("%Y-%m-%d %H:%M") == agora.strftime("%Y-%m-%d %H:%M"):
            msg = f"Já existe uma batida '{tipo}' registrada neste minuto."
            conn.close()
            return redirect(f"/funcionario?msg={urllib.parse.quote(msg)}")
        # Impede batidas consecutivas do mesmo tipo no mesmo dia
        if ultimo["tipo"] == tipo and dt_ultimo.date() == agora.date():
            msg = f"Não é permitido registrar duas batidas consecutivas do tipo '{tipo}' no mesmo dia."
            conn.close()
            return redirect(f"/funcionario?msg={urllib.parse.quote(msg)}")
        # Impede batidas do mesmo tipo nos últimos 10 segundos (proteção extra)
        diff = (agora - dt_ultimo).total_seconds()
        if ultimo["tipo"] == tipo and diff < 10:
            msg = f"Aguarde {int(10-diff)} segundos para registrar outra batida '{tipo}'."
            conn.close()
            return redirect(f"/funcionario?msg={urllib.parse.quote(msg)}")

    conn.execute(
        SQL_INSERT_REGISTRO,
        (
            session["user_id"],
            tipo,
            agora.strftime("%Y-%m-%d %H:%M:%S")
        )
    )
    conn.commit()
    conn.close()

    msg = f"Batida '{tipo}' registrada com sucesso!"
    return redirect(f"/funcionario?msg={urllib.parse.quote(msg)}")


# ==========================================================

# ==========================================================
# ADMIN - GERENCIAR BATIDAS
# ==========================================================

@app.route('/admin/gerar', methods=['GET', 'POST'])
def admin_gerar_route_fix():
    return admin_gerar()

# Rota para edição em lote do espelho de ponto
@app.route("/admin/espelho/editar_lote", methods=["POST"])
def admin_espelho_editar_lote():
    if session.get("tipo") != "admin":
        return redirect("/")

    usuario_id = request.form.get("usuario_id")
    # Exemplo: espera campos como registro_id[], data_hora[], tipo[], justificativa[]
    registro_ids = request.form.getlist("registro_id[]")
    datas_hora = request.form.getlist("data_hora[]")
    tipos = request.form.getlist("tipo[]")
    justificativas = request.form.getlist("justificativa[]")

    conn = get_db()
    c = conn.cursor()
    for i, reg_id in enumerate(registro_ids):
        data_hora = datas_hora[i]
        tipo = tipos[i]
        justificativa = justificativas[i] if i < len(justificativas) else ""
        # Normalizar formato vindo de <input type=datetime-local> (YYYY-MM-DDTHH:MM) para 'YYYY-MM-DD HH:MM:SS'
        if "T" in data_hora:
            data_hora = data_hora.replace("T", " ")
        if len(data_hora.split(" ")[1]) == 5:
            data_hora = data_hora + ":00"
        c.execute(
            "UPDATE registros SET data_hora=?, tipo=?, justificativa=? WHERE id=? AND usuario_id=?",
            (data_hora, tipo, justificativa, reg_id, usuario_id)
        )
    conn.commit()
    conn.close()
    return redirect(f"/admin/espelho?usuario_id={usuario_id}")

@app.route("/admin/espelho/deletar_registro", methods=["POST"])
def admin_espelho_deletar_registro():
    if session.get("tipo") != "admin":
        return redirect("/")

    usuario_id = request.form.get("usuario_id")
    registro_id = request.form.get("registro_id")

    if not usuario_id or not registro_id:
        return redirect(f"/admin/espelho?usuario_id={usuario_id or ''}&erro=Dados+inv%C3%A1lidos+para+exclus%C3%A3o")

    conn = get_db()
    c = conn.cursor()
    c.execute(
        "DELETE FROM registros WHERE id=? AND usuario_id=?",
        (registro_id, usuario_id)
    )
    removidos = c.rowcount
    conn.commit()
    conn.close()

    if removidos:
        return redirect(f"/admin/espelho?usuario_id={usuario_id}&sucesso=Batida+exclu%C3%ADda+com+sucesso")
    return redirect(f"/admin/espelho?usuario_id={usuario_id}&erro=Batida+n%C3%A3o+encontrada")

@app.route(ADMIN_ROUTE, methods=["GET", "POST"])
def admin():
    if session.get("tipo") != "admin":
        return redirect("/")

    # Redireciona para o espelho só após login
    if session.pop("admin_redirect_espelho", False):
        usuario_id = session.get('user_id')
        if usuario_id:
            return redirect(f'{ADMIN_ROUTE}/espelho?usuario_id={usuario_id}')

    conn = get_db()
    c = conn.cursor()
    # Buscar primeiro funcionário cadastrado
    funcionario = c.execute("SELECT id FROM usuarios WHERE tipo='funcionario' ORDER BY nome LIMIT 1").fetchone()
    if funcionario and not session.get('user_id'):
        conn.close()
        return redirect(f'{ADMIN_ROUTE}/espelho?usuario_id={funcionario[0]}')

    mensagem = request.args.get('mensagem') if request.args.get('mensagem') else None

    if request.method == "POST":
        form_tipo = request.form.get("form_tipo")
        if form_tipo == "funcionario":
            mensagem = cadastrar_funcionario(request, c, conn)
        elif form_tipo == "empresa":
            mensagem = cadastrar_empresa(request, c, conn)
        elif form_tipo == "gerar_batidas":
            mensagem = gerar_batidas(request, c, conn)

    funcionarios_raw = c.execute(
        "SELECT id, nome, cpf, jornada_diaria FROM usuarios WHERE tipo='funcionario'"
    ).fetchall()
    funcionarios = [dict(f) for f in funcionarios_raw]

    empresa_row = c.execute("SELECT * FROM empresa").fetchone()
    empresa = dict(empresa_row) if empresa_row else None

    conn.close()

    return render_with_theme(ADMIN_HTML,
        funcionarios=funcionarios,
        empresa=empresa,
        mensagem=mensagem
    )

def cadastrar_funcionario(request, c, conn):
    try:
        nome = request.form.get("nome")
        cpf = request.form.get("cpf")
        senha_h = generate_password_hash(request.form.get("senha"))
        jornada_v = request.form.get("jornada")
        matricula = request.form.get("matricula")
        cargo = request.form.get("cargo")
        departamento = request.form.get("departamento")
        pis = request.form.get("pis", "")
        cpf_norm = normalizar_cpf(cpf or "")

        if not cpf_valido(cpf or ""):
            return "CPF inválido."
        existe = c.execute(
            "SELECT id FROM usuarios WHERE cpf=?",
            (cpf_norm,)
        ).fetchone()
        if existe:
            return "Erro: CPF já cadastrado."
        c.execute(
            "INSERT INTO usuarios (nome, cpf, senha, tipo, jornada_diaria, matricula, cargo, departamento, pis) VALUES (?, ?, ?, 'funcionario', ?, ?, ?, ?, ?)",
            (
                nome,
                cpf_norm,
                senha_h,
                jornada_v,
                matricula,
                cargo,
                departamento,
                pis
            )
        )
        conn.commit()
        return "Funcionário cadastrado com sucesso!"
    except Exception:
        return "Erro ao cadastrar funcionário."

def cadastrar_empresa(request, c, conn):
    try:
        c.execute("DELETE FROM empresa")
        tolerancia_val = request.form.get("tolerancia", "5")
        try:
            tolerancia_val = int(tolerancia_val)
        except (TypeError, ValueError):
            tolerancia_val = 5
        c.execute(
            "INSERT INTO empresa (nome, cnpj, endereco, logo, tolerancia_min, cei) VALUES (?, ?, ?, ?, ?, ?)",
            (
                request.form["nome_empresa"],
                request.form["cnpj"],
                request.form["endereco"],
                request.form["logo"],
                tolerancia_val,
                request.form.get("cei", "")
            )
        )
        conn.commit()
        return "Empresa atualizada com sucesso!"
    except Exception:
        return "Erro ao atualizar empresa."

def gerar_batidas(request, c, conn):
    """
    Gera batidas para funcionários, refatorado para reduzir a complexidade.
    """
    usuario_id = request.form.get("usuario_id", "all")
    mes_form = request.form.get("mes", "").strip()
    ano_form = request.form.get("ano", "").strip()
    meses_form = request.form.get("meses", "1").strip()

    usuarios_gerar = _get_usuarios_para_gerar(c, usuario_id)
    if not usuarios_gerar:
        return "Nenhum funcionário selecionado para geração."

    mes, ano = _get_mes_ano(mes_form, ano_form)
    if mes < 1 or mes > 12 or ano < 2000 or ano > 3000:
        return "Mês/Ano inválido."

    try:
        meses_count = int(meses_form) if meses_form and int(meses_form) > 0 else 1
    except ValueError:
        meses_count = 1

    tolerancia = _get_tolerancia(c)
    hoje_data = datetime.now().date()
    total_registros = 0

    for u in usuarios_gerar:
        user_id = u["id"]
        jornada = float(u["jornada_diaria"]) if u["jornada_diaria"] else 8.0
        total_registros += _inserir_batidas(c, user_id, jornada, ano, mes, meses_count, tolerancia, hoje_data)

    conn.commit()
    return f"Batidas geradas para {meses_count} mês(es) terminando em {mes:02d}/{ano}. Registros inseridos: {total_registros}"

def _get_usuarios_para_gerar(c, usuario_id):
    if usuario_id == "all":
        return c.execute(
            "SELECT id, nome, jornada_diaria FROM usuarios WHERE tipo='funcionario'"
        ).fetchall()
    else:
        return c.execute(
            "SELECT id, nome, jornada_diaria FROM usuarios WHERE id=? AND tipo='funcionario'",
            (usuario_id,)
        ).fetchall()

def _get_mes_ano(mes_form, ano_form):
    if mes_form.isdigit() and ano_form.isdigit():
        return int(mes_form), int(ano_form)
    hoje = datetime.now()
    primeiro_dia_mes_atual = datetime(hoje.year, hoje.month, 1)
    ultimo_mes_anterior = primeiro_dia_mes_atual - timedelta(days=1)
    return ultimo_mes_anterior.month, ultimo_mes_anterior.year

def _get_tolerancia(c):
    tol_row = c.execute("SELECT tolerancia_min FROM empresa").fetchone()
    try:
        return int(tol_row["tolerancia_min"]) if tol_row and tol_row["tolerancia_min"] is not None else 5
    except (TypeError, ValueError, KeyError):
        return 5

def _inserir_batidas(c, user_id, jornada, ano, mes, meses_count, tolerancia, hoje_data):
    registros_inseridos = 0
    cur_year = ano
    cur_month = mes
    for _ in range(meses_count):
        registros_inseridos += _inserir_batidas_mes(c, user_id, jornada, cur_year, cur_month, tolerancia, hoje_data)
        cur_month -= 1
        if cur_month == 0:
            cur_month = 12
            cur_year -= 1
    return registros_inseridos

def _inserir_batidas_mes(c, user_id, jornada, ano, mes, tolerancia, hoje_data):
    registros_inseridos = 0
    total_dias = calendar.monthrange(ano, mes)[1]
    
    # Horários padrão de batida (em minutos desde meia-noite)
    horarios_padrao = [
        (8 * 60, 'entrada'),      # 08:00 - Entrada 1
        (12 * 60, 'saida'),        # 12:00 - Saída 1 (Almoço)
        (13 * 60, 'entrada'),      # 13:00 - Entrada 2 (Retorno)
        (18 * 60, 'saida')         # 18:00 - Saída 2 (Saída final)
    ]
    
    # Variação máxima: ±5 minutos
    variacao_maxima = 5
    
    for dia in range(1, total_dias + 1):
        data_base = datetime(ano, mes, dia)
        if data_base.date() > hoje_data:
            continue
        if data_base.weekday() >= 5 or is_feriado_completo(data_base):
            continue
        existente = c.execute(
            f"SELECT COUNT(*) as cnt FROM registros WHERE 1=1{SQL_AND_USUARIO_ID} AND date(data_hora)=?",
            (user_id, data_base.date())
        ).fetchone()["cnt"]
        if existente > 0:
            continue
        
        # Inserir as 4 batidas padrão com variação aleatória de ±5 minutos
        for hora_minutos_base, tipo_batida in horarios_padrao:
            # Variação aleatória de ±5 minutos
            variacao = random.randint(-variacao_maxima, variacao_maxima)
            hora_minutos_final = hora_minutos_base + variacao
            
            # Converter de minutos para hora e minuto
            hora = hora_minutos_final // 60
            minuto = hora_minutos_final % 60
            
            # Criar datetime com a hora/minuto variado
            data_hora = datetime(ano, mes, dia, hora, minuto)
            
            c.execute(
                SQL_INSERT_REGISTRO,
                (user_id, tipo_batida, data_hora.strftime("%Y-%m-%d %H:%M:%S"))
            )
            registros_inseridos += 1
    
    return registros_inseridos


@app.route(f'{ADMIN_ROUTE}/empresa', methods=['GET', 'POST'])
def admin_empresa():
    if session.get('tipo') != 'admin':
        return redirect('/')
    
    conn = get_db()
    c = conn.cursor()
    mensagem = None
    
    # Cadastro empresa
    if request.method == 'POST' and request.form.get('form_tipo') == 'empresa':
        c.execute('DELETE FROM empresa')
        tolerancia_val = request.form.get('tolerancia', '5')
        try:
            tolerancia_val = int(tolerancia_val)
        except (TypeError, ValueError):
            tolerancia_val = 5
        c.execute(
            'INSERT INTO empresa (nome, cnpj, endereco, logo, tolerancia_min, cei) VALUES (?, ?, ?, ?, ?, ?)',
            (
                request.form['nome_empresa'],
                request.form['cnpj'],
                request.form['endereco'],
                request.form['logo'],
                tolerancia_val,
                request.form.get('cei', '')
            )
        )
        conn.commit()
        mensagem = 'Empresa atualizada com sucesso!'
    
    empresa = c.execute('SELECT * FROM empresa').fetchone()
    conn.close()
    
    return render_with_theme(ADMIN_EMPRESA_HTML, empresa=empresa, mensagem=mensagem)


@app.route(f'{ADMIN_ROUTE}/funcionarios', methods=['GET', 'POST'])
def admin_funcionarios():
    if session.get('tipo') != 'admin':
        return redirect('/')

    conn = get_db()
    c = conn.cursor()
    mensagem = request.args.get('mensagem') if request.args.get('mensagem') else None

    def cadastrar_funcionario_form():
        nome = request.form.get('nome')
        cpf = request.form.get('cpf')
        senha_h = generate_password_hash(request.form.get('senha'))
        jornada_v = request.form.get('jornada')
        matricula = request.form.get('matricula')
        cargo = request.form.get('cargo')
        departamento = request.form.get('departamento')
        pis = request.form.get('pis', '')
        cpf_norm = normalizar_cpf(cpf or '')

        if not cpf_valido(cpf or ''):
            return 'CPF inválido.'
        existe = c.execute(
            'SELECT id FROM usuarios WHERE cpf=?',
            (cpf_norm,)
        ).fetchone()
        if existe:
            return 'Erro: CPF já cadastrado.'
        c.execute(
            'INSERT INTO usuarios (nome, cpf, senha, tipo, jornada_diaria, matricula, cargo, departamento, pis) VALUES (?, ?, ?, "funcionario", ?, ?, ?, ?, ?)',
            (nome, cpf_norm, senha_h, jornada_v, matricula, cargo, departamento, pis)
        )
        conn.commit()
        return 'Funcionário cadastrado com sucesso!'

    # Cadastro funcionário
    if request.method == 'POST' and request.form.get('form_tipo') == 'funcionario':
        try:
            mensagem = cadastrar_funcionario_form()
        except Exception:
            if not mensagem:
                mensagem = 'Erro ao cadastrar funcionário.'

    funcionarios_raw = c.execute(
        'SELECT id, nome, cpf, jornada_diaria FROM usuarios WHERE tipo="funcionario"'
    ).fetchall()
    funcionarios = [dict(f) for f in funcionarios_raw]

    conn.close()

    return render_with_theme(ADMIN_FUNCIONARIOS_HTML, funcionarios=funcionarios, mensagem=mensagem)


def get_usuarios_gerar(c, usuario_id):
    if usuario_id == 'all':
        return c.execute(
            'SELECT id, nome, jornada_diaria FROM usuarios WHERE tipo="funcionario"'
        ).fetchall()
    else:
        return c.execute(
            'SELECT id, nome, jornada_diaria FROM usuarios WHERE id=? AND tipo="funcionario"',
            (usuario_id,)
        ).fetchall()

def get_mes_ano(mes_form, ano_form):
    if mes_form.isdigit() and ano_form.isdigit():
        return int(mes_form), int(ano_form)
    hoje = datetime.now()
    primeiro_dia_mes_atual = datetime(hoje.year, hoje.month, 1)
    ultimo_mes_anterior = primeiro_dia_mes_atual - timedelta(days=1)
    return ultimo_mes_anterior.month, ultimo_mes_anterior.year

def get_tolerancia(c):
    tol_row = c.execute('SELECT tolerancia_min FROM empresa').fetchone()
    try:
        return int(tol_row['tolerancia_min']) if tol_row and tol_row['tolerancia_min'] is not None else 5
    except (TypeError, ValueError, KeyError):
        return 5

def inserir_batidas_usuario(c, user_id, jornada, ano, mes, meses_count, tolerancia, hoje_data):
    registros_inseridos = 0
    cur_year = ano
    cur_month = mes
    for _ in range(meses_count):
        registros_inseridos += inserir_batidas_mes(c, user_id, cur_year, cur_month, hoje_data)
        cur_month -= 1
        if cur_month == 0:
            cur_month = 12
            cur_year -= 1
    return registros_inseridos

def inserir_batidas_mes(c, user_id, ano, mes, hoje_data):
    registros_inseridos = 0
    total_dias = calendar.monthrange(ano, mes)[1]
    
    # Horários padrão de batida (em minutos desde meia-noite)
    horarios_padrao = [
        (8 * 60, 'entrada'),      # 08:00 - Entrada 1
        (12 * 60, 'saida'),        # 12:00 - Saída 1 (Almoço)
        (13 * 60, 'entrada'),      # 13:00 - Entrada 2 (Retorno)
        (18 * 60, 'saida')         # 18:00 - Saída 2 (Saída final)
    ]
    
    # Variação máxima: ±5 minutos
    variacao_maxima = 5
    
    for dia in range(1, total_dias + 1):
        data_base = datetime(ano, mes, dia)
        if data_base.date() > hoje_data:
            continue
        if data_base.weekday() >= 5 or is_feriado_completo(data_base):
            continue
        
        # Verificar se já existem batidas para este dia
        existente = c.execute(
            'SELECT COUNT(*) as cnt FROM registros WHERE usuario_id=? AND date(data_hora)=?',
            (user_id, data_base.date())
        ).fetchone()['cnt']
        if existente > 0:
            continue
        
        # Inserir as 4 batidas padrão com variação aleatória de ±5 minutos
        SQL_INSERT = 'INSERT INTO registros (usuario_id, tipo, data_hora) VALUES (?, ?, ?)'
        
        for hora_minutos_base, tipo_batida in horarios_padrao:
            # Variação aleatória de ±5 minutos
            variacao = random.randint(-variacao_maxima, variacao_maxima)
            hora_minutos_final = hora_minutos_base + variacao
            
            # Converter de minutos para hora e minuto
            hora = hora_minutos_final // 60
            minuto = hora_minutos_final % 60
            
            # Criar datetime com a hora/minuto variado
            data_hora = datetime(ano, mes, dia, hora, minuto)
            
            c.execute(SQL_INSERT, (user_id, tipo_batida, data_hora.strftime('%Y-%m-%d %H:%M:%S')))
            registros_inseridos += 1
    
    return registros_inseridos

def processar_geracao_batidas(request, conn):
    usuario_id = request.form.get('usuario_id', 'all')
    data_inicial_str = request.form.get('data_inicial', '').strip()
    data_final_str = request.form.get('data_final', '').strip()

    c = conn.cursor()
    usuarios_gerar = get_usuarios_gerar(c, usuario_id)
    if not usuarios_gerar:
        return 'Nenhum funcionário selecionado para geração.'
    
    # Validar e converter datas
    try:
        data_inicial = datetime.strptime(data_inicial_str, '%Y-%m-%d').date()
        data_final = datetime.strptime(data_final_str, '%Y-%m-%d').date()
    except ValueError:
        return '❌ Datas inválidas. Use o formato YYYY-MM-DD.'
    
    if data_inicial > data_final:
        return '❌ Data inicial não pode ser posterior à data final.'
    
    hoje = datetime.now().date()
    if data_final > hoje:
        data_final = hoje
    
    # Gerar batidas para o período especificado
    registros_inseridos = 0
    for u in usuarios_gerar:
        user_id = u['id']
        registros_inseridos += inserir_batidas_por_periodo(c, user_id, data_inicial, data_final)

    conn.commit()
    dias_count = (data_final - data_inicial).days + 1
    return f'✅ Batidas geradas de {data_inicial.strftime("%d/%m/%Y")} a {data_final.strftime("%d/%m/%Y")} ({dias_count} dias). Total: {registros_inseridos} registro(s).'

def inserir_batidas_por_periodo(c, user_id, data_inicial, data_final):
    """Insere batidas para um período específico"""
    registros_inseridos = 0
    
    # Horários padrão de batida (em minutos desde meia-noite)
    horarios_padrao = [
        (8 * 60, 'entrada'),      # 08:00
        (12 * 60, 'saida'),        # 12:00
        (13 * 60, 'entrada'),      # 13:00
        (18 * 60, 'saida')         # 18:00
    ]
    
    variacao_maxima = 5
    
    data_atual = data_inicial
    while data_atual <= data_final:
        # Pular finais de semana e feriados
        if data_atual.weekday() >= 5 or is_feriado_completo(datetime.combine(data_atual, datetime.min.time())):
            data_atual += timedelta(days=1)
            continue
        
        # Verificar se já existem batidas para este dia
        existente = c.execute(
            'SELECT COUNT(*) as cnt FROM registros WHERE usuario_id=? AND date(data_hora)=?',
            (user_id, data_atual)
        ).fetchone()['cnt']
        
        if existente == 0:
            # Inserir as 4 batidas padrão com variação aleatória de ±5 minutos
            SQL_INSERT = 'INSERT INTO registros (usuario_id, tipo, data_hora) VALUES (?, ?, ?)'
            
            for hora_minutos_base, tipo_batida in horarios_padrao:
                variacao = random.randint(-variacao_maxima, variacao_maxima)
                hora_minutos_final = hora_minutos_base + variacao
                
                hora = hora_minutos_final // 60
                minuto = hora_minutos_final % 60
                
                data_hora = datetime.combine(data_atual, datetime.min.time()).replace(hour=hora, minute=minuto)
                
                c.execute(SQL_INSERT, (user_id, tipo_batida, data_hora.strftime('%Y-%m-%d %H:%M:%S')))
                registros_inseridos += 1
        
        data_atual += timedelta(days=1)
    
    return registros_inseridos

def admin_gerar():
    if session.get('tipo') != 'admin':
        return redirect('/')

    conn = get_db()
    c = conn.cursor()
    mensagem = None

    # Gerar batidas aleatórias
    if request.method == 'POST' and request.form.get('form_tipo') == 'gerar_batidas':
        mensagem = processar_geracao_batidas(request, conn)

    funcionarios_raw = c.execute(
        'SELECT id, nome, cpf, jornada_diaria FROM usuarios WHERE tipo="funcionario"'
    ).fetchall()
    funcionarios = []
    for f in funcionarios_raw:
        funcionarios.append({
            'id': f['id'] if f['id'] is not None else '',
            'nome': f['nome'] if f['nome'] is not None else '',
            'cpf': f['cpf'] if f['cpf'] is not None else '',
            'jornada_diaria': f['jornada_diaria'] if f['jornada_diaria'] is not None else 8.0
        })
    conn.close()
    return render_with_theme(
        ADMIN_GERAR_HTML,
        funcionarios=funcionarios,
        mensagem=mensagem,
        labels=[],
        horas=[]
    )


@app.route('/admin/registros/deletar_todos', methods=['POST'])
def admin_registros_deletar_todos():
    if session.get('tipo') != 'admin':
        return redirect('/')
    
    # Validar senha do admin
    senha_confirmacao = request.form.get('senha_confirmacao', '')
    
    if not senha_confirmacao:
        return redirect('/admin?erro=Senha de confirmação obrigatória')
    
    # Buscar admin logado
    conn = get_db()
    admin = conn.execute(
        'SELECT senha FROM usuarios WHERE id=?',
        (session['user_id'],)
    ).fetchone()
    
    if not admin or not check_password_hash(admin['senha'], senha_confirmacao):
        conn.close()
        return redirect('/admin?erro=Senha incorreta')
    
    # Senha correta, deletar todos os registros
    conn.execute('DELETE FROM registros')
    conn.commit()
    conn.close()
    return redirect(f'{ADMIN_ROUTE}?sucesso=Todas as batidas foram apagadas')


@app.route('/admin/registros/deletar_usuario', methods=['POST'])
def admin_registros_deletar_usuario():
    if session.get('tipo') != 'admin':
        return redirect('/')
    usuario_id = request.args.get("usuario_id")
    if not usuario_id:
        return redirect("/admin")

    # Removed unused variables 'mes' and 'ano'

    conn = get_db()
    c = conn.cursor()
    usuario_row = c.execute("SELECT id, nome, jornada_diaria FROM usuarios WHERE id=?", (usuario_id,)).fetchone()
    if usuario_row:
        usuario = dict(usuario_row)
        if 'jornada_diaria' not in usuario or usuario['jornada_diaria'] is None:
            pass  # Optionally handle missing jornada_diaria
    # Removed unused assignment to usuario and checks for undefined 'u'

    conn.execute('DELETE FROM usuarios WHERE id=?', (usuario_id,))
    conn.commit()
    conn.close()
    msg = urllib.parse.quote('Usuário excluído com sucesso.')
    return redirect(f'{ADMIN_ROUTE}?mensagem={msg}')


@app.route('/admin/usuario/editar_form')
def admin_usuario_editar_form():
    if session.get('tipo') != 'admin':
        return redirect('/')
    usuario_id = request.args.get('usuario_id')
    if not usuario_id:
        return redirect(ADMIN_ROUTE)
    conn = get_db()
    u_row = conn.execute('SELECT id, nome, cpf, jornada_diaria, matricula, cargo, departamento, pis FROM usuarios WHERE id=?', (usuario_id,)).fetchone()
    u = dict(u_row) if u_row else None
    conn.close()
    if not u:
        return redirect('/admin')
    
    # Formulário completo para editar usuário
    html = '''<!DOCTYPE html>
<html>
<head>
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <style>
        @import url("https://fonts.googleapis.com/css2?family=Inter:wght@300;400;600;700&display=swap");
        
        :root {
            --bg-1: #0f2027;
            --bg-2: #203a43;
            --bg-3: #2c5364;
            --accent-1: #00c6ff;
            --accent-2: #0072ff;
        }
        
        :root[data-theme="green"] {
            --bg-1: #0f2027;
            --bg-2: #203a43;
            --bg-3: #2c5364;
            --accent-1: #00c6ff;
            --accent-2: #0072ff;
        }
        
        :root[data-theme="red"] {
            --bg-1: #1a0a0f;
            --bg-2: #3d1a25;
            --bg-3: #5a2a38;
            --accent-1: #ff6b6b;
            --accent-2: #ff3838;
        }
        
        :root[data-theme="blue"] {
            --bg-1: #0a0f27;
            --bg-2: #1a2a5e;
            --bg-3: #2a3f7a;
            --accent-1: #4d9fff;
            --accent-2: #1e5eff;
        }
        
        body {
            margin: 0;
            font-family: Inter, sans-serif;
            background: linear-gradient(135deg, var(--bg-1), var(--bg-2), var(--bg-3));
            color: #fff;
            padding: 24px;
            transition: background 0.5s ease;
        }
        .card {
            background: rgba(255,255,255,0.06);
            backdrop-filter: blur(10px);
            padding: 24px;
            border-radius: 12px;
            max-width: 600px;
            margin: 0 auto;
            border: 1px solid rgba(255,255,255,0.1);
        }
        h2 {
            margin-top: 0;
            color: var(--accent-1);
        }
        .form-group {
            margin-bottom: 16px;
        }
        label {
            display: block;
            margin-bottom: 6px;
            color: rgba(255,255,255,0.8);
            font-size: 14px;
        }
        input, select {
            width: 100%;
            padding: 10px;
            border-radius: 8px;
            border: 1px solid rgba(255,255,255,0.1);
            font-size: 14px;
            box-sizing: border-box;
            background: rgba(255,255,255,0.05);
            color: #fff;
        }
        input:focus, select:focus {
            outline: none;
            border-color: var(--accent-1);
            box-shadow: 0 0 0 2px rgba(0, 198, 255, 0.2);
        }
        .btn {
            padding: 12px 20px;
            border-radius: 8px;
            border: none;
            cursor: pointer;
            font-size: 14px;
            font-weight: 600;
            margin-right: 8px;
            transition: all 0.3s;
        }
        .btn-primary {
            background: linear-gradient(135deg, var(--accent-1), var(--accent-2));
            color: #fff;
        }
        .btn-primary:hover {
            transform: scale(1.02);
            box-shadow: 0 4px 15px rgba(0, 198, 255, 0.3);
        }
        .btn-secondary {
            background: rgba(255,255,255,0.1);
            color: #fff;
        }
        .btn-secondary:hover {
            background: rgba(255,255,255,0.15);
        }
        .actions {
            margin-top: 24px;
            display: flex;
            gap: 8px;
        }
    </style>
</head>
<body>
    <div class="card">
        <h2>✏️ Editar Funcionário</h2>
        <form method="POST" action="/admin/usuario/editar_cadastro">
            <input type="hidden" name="usuario_id" value="{{u.id}}">
            
            <div class="form-group">
                <label>Nome Completo</label>
                <input name="nome" type="text" value="{{u.nome}}" required>
            </div>
            
            <div class="form-group">
                <label>CPF</label>
                <input name="cpf" type="text" maxlength="14" oninput="formatarCpf(this)" value="{{u.cpf}}" required>
            </div>
            
            <div class="form-group">
                <label>PIS (Portaria 1510)</label>
                <input name="pis" type="text" value="{{u.pis if u.pis}}" placeholder="Opcional">
            </div>
            
            <div class="form-group">
                <label>Jornada Diária (horas)</label>
                <input name="jornada" type="number" step="0.5" min="1" max="12" value="{{u.jornada_diaria if u.jornada_diaria is not none else 8.0}}" required>
            </div>
            
            <div class="form-group">
                <label>Matrícula</label>
                <input name="matricula" type="text" value="{{u.matricula if u.matricula is not none else ''}}">
            </div>
            
            <div class="form-group">
                <label>Cargo</label>
                <input name="cargo" type="text" value="{{u.cargo if u.cargo is not none else ''}}">
            </div>
            
            <div class="form-group">
                <label>Departamento</label>
                <input name="departamento" type="text" value="{{u.departamento if u.departamento is not none else ''}}">
            </div>
            
            <div class="actions">
                <button type="submit" class="btn btn-primary">💾 Salvar Alterações</button>
                <a href="/admin/funcionarios" class="btn btn-secondary" style="text-decoration:none;display:inline-block;text-align:center">🚫 Cancelar</a>
            </div>
        </form>
    </div>
    <script src="/static/cpf.js"></script>
    <script>
        function loadTheme() {
            const cookies = document.cookie.split(';');
            let theme = 'green';
            let encontrou = false;
            for (let cookie of cookies) {
                const [name, value] = cookie.trim().split('=');
                if (name === 'theme') {
                    theme = value;
                    encontrou = true;
                    break;
                }
            }
            if (!encontrou) {
                document.cookie = "theme=green; path=/; max-age=31536000";
            }
            document.documentElement.setAttribute('data-theme', theme);
        }
        loadTheme();
    </script>
</body>
</html>'''
    return render_with_theme(html, u=u)


@app.route('/admin/usuario/editar_cadastro', methods=['POST'])
def admin_usuario_editar_cadastro():
    if session.get('tipo') != 'admin':
        return redirect('/')
    usuario_id = request.form.get('usuario_id')
    nome = request.form.get('nome')
    cpf = request.form.get('cpf')
    jornada = request.form.get('jornada')
    matricula = request.form.get('matricula')
    cargo = request.form.get('cargo')
    departamento = request.form.get('departamento')
    pis = request.form.get('pis', '')

    try:
        jornada_val = float(jornada) if jornada else None
    except (TypeError, ValueError):
        jornada_val = None

    conn = get_db()
    cpf_norm = normalizar_cpf(cpf or '')
    if not cpf_valido(cpf or ''):
        conn.close()
        return redirect('/admin/funcionarios?mensagem=CPF%20inv%C3%A1lido.')

    existe = conn.execute(
        'SELECT id FROM usuarios WHERE cpf=? AND id<>?',
        (cpf_norm, usuario_id)
    ).fetchone()
    if existe:
        conn.close()
        return redirect('/admin/funcionarios?mensagem=CPF%20j%C3%A1%20cadastrado.')

    # atualizar campos fornecidos
    conn.execute('UPDATE usuarios SET nome=?, cpf=?, jornada_diaria=?, matricula=?, cargo=?, departamento=?, pis=? WHERE id=?',
                 (nome, cpf_norm, jornada_val, matricula, cargo, departamento, pis, usuario_id))
    conn.commit()
    conn.close()
    return redirect('/admin/funcionarios')


@app.route('/admin/usuario/editar', methods=['POST'])
def admin_usuario_editar():
    if session.get('tipo') != 'admin':
        return redirect('/')

    usuario_id = request.form.get('usuario_id')
    jornada = request.form.get('jornada')
    try:
        jornada_val = float(jornada)
    except ValueError:
        jornada_val = None

    if jornada_val is not None:
        conn = get_db()
        conn.execute('UPDATE usuarios SET jornada_diaria=? WHERE id=?', (jornada_val, usuario_id))
        conn.commit()
        conn.close()

    return redirect('/admin')


@app.route("/admin/restart", methods=["POST"])
def admin_restart():
    if session.get("tipo") != "admin":
        return redirect("/")

    # Path to helper PowerShell script that will start a new python process
    helper = r"C:\RelogioPonto\restart_helper.ps1"

    if not os.path.exists(helper):
        return "Script de reinício não encontrado. Crie 'restart_helper.ps1' em " + helper

    try:
        subprocess.Popen([
            "powershell",
            "-NoProfile",
            "-ExecutionPolicy",
            "Bypass",
            "-File",
            helper
        ], stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
    except Exception as e:
        return f"Erro ao iniciar script de reinício: {str(e)}"

    # Agendar término do processo atual breve para que helper possa iniciar nova instância
    threading.Timer(1.0, lambda: os._exit(0)).start()
    return "Reiniciando aplicação..."


@app.route("/admin/editar", methods=["POST"])
def editar_batida():
    if session.get("tipo") != "admin":
        return redirect("/")

    usuario_id = request.form["usuario_id"]
    data_hora = request.form["data_hora"]
    tipo = request.form["tipo"]
    # Normalizar formato vindo de <input type=datetime-local> (YYYY-MM-DDTHH:MM) para
    # 'YYYY-MM-DD HH:MM:SS' para compatibilidade com sqlite date().
    if "T" in data_hora:
        data_hora = data_hora.replace("T", " ")
    if len(data_hora.split(" ")[1]) == 5:
        data_hora = data_hora + ":00"

    conn = get_db()
    conn.execute(
        "INSERT INTO registros (usuario_id, tipo, data_hora) VALUES (?, ?, ?)",
        (usuario_id, tipo, data_hora)
    )
    conn.commit()
    conn.close()

    return redirect("/admin")

# Endpoint para inserir dia abonado (abono)
@app.route("/admin/inserir_abono", methods=["POST"])
def admin_inserir_abono():
    if session.get("tipo") != "admin":
        return redirect("/")

    usuario_id = request.form.get("usuario_id")
    data_abono = request.form.get("data_abono")
    justificativa = request.form.get("justificativa", "")
    if not usuario_id or not data_abono:
        return redirect("/admin?mensagem=Preencha todos os campos do abono.")

    # Inserir registro de abono como tipo 'abono' no início do dia
    data_hora = f"{data_abono} 00:00:00"
    conn = get_db()
    conn.execute(
        "INSERT INTO registros (usuario_id, tipo, data_hora, justificativa) VALUES (?, ?, ?, ?)",
        (usuario_id, "abono", data_hora, justificativa)
    )
    conn.commit()
    conn.close()
    return redirect("/admin?mensagem=Dia abonado inserido com sucesso!")


def build_registros_por_dia(registros):
    from collections import defaultdict
    registros_por_dia = defaultdict(list)
    abonos_por_dia = set()
    for r in registros:
        dia = r['data_hora'][:10]
        registros_por_dia[dia].append(r)
        if r['tipo'] == 'abono':
            abonos_por_dia.add(dia)
    return registros_por_dia, abonos_por_dia

def build_feriados_por_dia(ano, mes_pad):
    import calendar
    feriados_por_dia = set()
    dias_mes = calendar.monthrange(int(ano), int(mes_pad))[1]
    for d in range(1, dias_mes+1):
        data_obj = datetime(int(ano), int(mes_pad), d)
        if is_feriado_completo(data_obj):
            feriados_por_dia.add(data_obj.strftime('%Y-%m-%d'))
    return feriados_por_dia

def build_semanas(ano, mes_pad):
    import calendar
    dias_mes = calendar.monthrange(int(ano), int(mes_pad))[1]
    primeiro_dia_semana = datetime(int(ano), int(mes_pad), 1).weekday()
    semanas = []
    semana = [None] * primeiro_dia_semana
    for d in range(1, dias_mes + 1):
        data_str = f"{ano}-{mes_pad}-{str(d).zfill(2)}"
        semana.append(data_str)
        if len(semana) == 7:
            semanas.append(semana)
            semana = []
    if semana:
        while len(semana) < 7:
            semana.append(None)
        semanas.append(semana)
    return semanas

@app.route("/admin/espelho")
def admin_espelho():
    import calendar

    if session.get("tipo") != "admin":
        return redirect("/")

    usuario_id = request.args.get("usuario_id")
    ano = request.args.get('ano') or str(datetime.now().year)
    mes_pad = request.args.get('mes') or str(datetime.now().month).zfill(2)
    sucesso = request.args.get('sucesso')
    erro = request.args.get('erro')

    conn = get_db()
    c = conn.cursor()
    usuario_row = c.execute("SELECT id, nome, jornada_diaria FROM usuarios WHERE id=?", (usuario_id,)).fetchone()
    usuario = dict(usuario_row) if usuario_row else {'id': '', 'nome': '', 'jornada_diaria': 8.0}

    funcionarios = c.execute("SELECT id, nome FROM usuarios WHERE tipo='funcionario' ORDER BY nome").fetchall()

    registros = c.execute(
        "SELECT r.id, r.tipo, r.data_hora, r.justificativa FROM registros r WHERE r.usuario_id=? AND strftime('%Y', r.data_hora)=? AND strftime('%m', r.data_hora)=? ORDER BY r.data_hora",
        (usuario_id, ano, mes_pad)
    ).fetchall() if usuario_id else []

    registros_por_dia, abonos_por_dia = build_registros_por_dia(registros)
    feriados_por_dia = build_feriados_por_dia(ano, mes_pad)
    semanas = build_semanas(ano, mes_pad)
    conn.close()

    from flask import render_template_string
    template = '''<!DOCTYPE html>
<html>
<head>
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <style>
        @import url("https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700;800&display=swap");
        :root {
            --bg-1: #0f2027;
            --bg-2: #203a43;
            --bg-3: #2c5364;
            --accent-1: #00c6ff;
            --accent-2: #0072ff;
            --particle-1: rgba(0, 198, 255, 0.15);
            --particle-2: rgba(44, 83, 100, 0.2);
            --particle-3: rgba(32, 58, 67, 0.15);
        }
        * { box-sizing: border-box; }
        body {
            margin: 0;
            font-family: Inter, sans-serif;
            background: linear-gradient(135deg, var(--bg-1), var(--bg-2), var(--bg-3));
            min-height: 100vh;
            color: #fff;
            position: relative;
            overflow-x: hidden;
            transition: background 0.5s ease;
        }
        body::before {
            content: "";
            position: fixed;
            width: 100%;
            height: 100%;
            background-image: 
                radial-gradient(circle at 20% 80%, var(--particle-1) 0%, transparent 50%),
                radial-gradient(circle at 80% 20%, var(--particle-2) 0%, transparent 50%),
                radial-gradient(circle at 40% 40%, var(--particle-3) 0%, transparent 50%);
            animation: particleFloat 20s ease-in-out infinite;
            z-index: 0;
        }
        @keyframes particleFloat {
            0%, 100% { opacity: 1; transform: translateY(0px); }
            50% { opacity: 0.8; transform: translateY(-20px); }
        }
        .layout { display: flex; gap: 24px; padding: 30px; position: relative; z-index: 1; }
        .sidebar {
            width: 260px;
            background: linear-gradient(145deg, rgba(255, 255, 255, 0.08), rgba(255, 255, 255, 0.03));
            backdrop-filter: blur(10px);
            padding: 24px;
            border-radius: 20px;
            height: calc(100vh - 60px);
            position: sticky;
            top: 30px;
            border: 1px solid rgba(255, 255, 255, 0.1);
            box-shadow: 0 8px 32px rgba(0, 0, 0, 0.3), inset 0 1px 0 rgba(255, 255, 255, 0.1);
            display: flex;
            flex-direction: column;
            align-items: flex-start;
        }
        .sidebar h2 {
            margin: 0 0 30px 0;
            font-size: 26px;
            font-weight: 800;
            background: linear-gradient(135deg, #00c6ff, #0072ff);
            -webkit-background-clip: text;
            -webkit-text-fill-color: transparent;
            text-shadow: 0 0 30px rgba(0, 198, 255, 0.5);
            letter-spacing: -0.5px;
        }
        .sidebar nav {
            width: 100%;
            display: flex;
            flex-direction: column;
            gap: 8px;
        }
        .sidebar nav a {
            display: flex;
            align-items: center;
            gap: 12px;
            color: rgba(255, 255, 255, 0.8);
            text-decoration: none;
            padding: 14px 16px;
            margin: 4px 0;
            border-radius: 12px;
            transition: all 0.3s cubic-bezier(0.4, 0, 0.2, 1);
            font-weight: 500;
            position: relative;
            overflow: hidden;
        }
        .sidebar nav a::before {
            content: "";
            position: absolute;
            top: 0;
            left: -100%;
            width: 100%;
            height: 100%;
            background: linear-gradient(90deg, transparent, rgba(255, 255, 255, 0.1), transparent);
            transition: left 0.5s;
        }
        .sidebar nav a:hover::before { left: 100%; }
        .sidebar nav a:hover {
            color: #fff;
            background: rgba(255, 255, 255, 0.05);
            transform: translateX(5px);
            box-shadow: 0 4px 15px rgba(0, 198, 255, 0.2);
        }
        .sidebar nav a.active {
            background: linear-gradient(135deg, rgba(0, 198, 255, 0.3), rgba(0, 114, 255, 0.15));
            border: 1px solid rgba(0, 198, 255, 0.4);
            box-shadow: 0 8px 25px rgba(0, 198, 255, 0.25), inset 0 1px 0 rgba(255, 255, 255, 0.2);
            font-weight: 700;
            color: #fff;
        }
        .content { flex: 1; max-width: 1400px; }
        .card {
            background: linear-gradient(145deg, rgba(255, 255, 255, 0.1), rgba(255, 255, 255, 0.05));
            backdrop-filter: blur(10px);
            padding: 30px;
            border-radius: 20px;
            margin-bottom: 24px;
            border: 1px solid rgba(255, 255, 255, 0.1);
            box-shadow: 0 8px 32px rgba(0, 0, 0, 0.3), inset 0 1px 0 rgba(255, 255, 255, 0.1);
            animation: fadeInUp 0.5s ease-out;
        }
        @keyframes fadeInUp {
            from { opacity: 0; transform: translateY(30px); }
            to { opacity: 1; transform: translateY(0); }
        }
        .header-section {
            display: flex;
            justify-content: space-between;
            align-items: flex-start;
            flex-wrap: wrap;
            gap: 20px;
            margin-bottom: 24px;
            padding-bottom: 20px;
            border-bottom: 2px solid rgba(0, 198, 255, 0.3);
        }
        .header-section h2 {
            margin: 0 0 8px 0;
            font-size: 28px;
            font-weight: 700;
            background: linear-gradient(135deg, #fff, #00c6ff);
            -webkit-background-clip: text;
        }
        select option {
            color: #111 !important;
            background: #fff !important;
        }
        .flash-msg {
            transition: opacity 0.35s ease, transform 0.35s ease;
        }
        .flash-msg.hide {
            opacity: 0;
            transform: translateY(-4px);
        }
        .btn-del {
            background: linear-gradient(135deg, #e74c3c, #c0392b);
            color: #fff;
            padding: 8px 14px;
            border-radius: 8px;
            border: none;
            cursor: pointer;
            font-weight: 600;
            font-size: 12px;
            transition: all 0.3s ease;
            box-shadow: 0 2px 10px rgba(231, 76, 60, 0.3);
        }
        .btn-del:hover {
            box-shadow: 0 4px 15px rgba(231, 76, 60, 0.5);
            transform: translateY(-2px);
        }
        /* Responsive */
        @media(max-width: 1024px) {
            .layout {
                flex-direction: column;
            }
            .sidebar {
                width: 100%;
                height: auto;
                position: relative;
                top: 0;
            }
            .header-section {
                flex-direction: column;
                align-items: flex-start;
            }
            .top-actions {
                width: 100%;
            }
        }
        @media(max-width: 768px) {
            .layout {
                padding: 20px;
            }
            .card {
                padding: 20px;
            }
            table {
                font-size: 13px;
            }
            table th, table td {
                padding: 12px 8px;
            }
            .edit-form {
                flex-direction: column;
                width: 100%;
            }
            .edit-form input,
            .edit-form select,
            .edit-form button {
                width: 100%;
            }
        }
    </style>
</head>
<body>
    <div class="layout">
        <aside class="sidebar">
            <h2>⏱ PontoPro</h2>
            <nav>
                <a href="/admin">📊 Visão Geral</a>
                <a href="/dashboard">📈 Dashboard</a>
                <a href="/admin/empresa">🏢 Empresa</a>
                <a href="/admin/funcionarios">👥 Funcionários</a>
                <a href="/admin/gerar">⚙️ Gerar Batidas</a>
                <a href="/logout">🚪 Sair</a>
            </nav>
        </aside>
        <main class="content">
            <div class="card">
                {% if sucesso %}
                <div class="flash-msg" style="margin-bottom:14px;padding:10px 12px;border-radius:10px;background:rgba(46, 204, 113, 0.15);border:1px solid rgba(46, 204, 113, 0.45);color:#b8ffd6;font-weight:600;">✅ {{ sucesso }}</div>
                {% endif %}
                {% if erro %}
                <div class="flash-msg" style="margin-bottom:14px;padding:10px 12px;border-radius:10px;background:rgba(231, 76, 60, 0.15);border:1px solid rgba(231, 76, 60, 0.45);color:#ffd2cd;font-weight:600;">❌ {{ erro }}</div>
                {% endif %}
                <div class="header-section">
                    <div>
                        <h2>📄 Espelho de Ponto — {{ usuario.nome }}</h2>
                        <div class="subtitle">Período: {{ mes_pad }}/{{ ano }}</div>
                    </div>
                    <div class="top-actions">
                        <a href="/admin/espelho/export?usuario_id={{ usuario.id }}&tipo=excel&mes={{ mes_pad }}&ano={{ ano }}">📊 Excel</a>
                        <a href="/admin/espelho/export?usuario_id={{ usuario.id }}&tipo=afd&mes={{ mes_pad }}&ano={{ ano }}">📋 AFD</a>
                        <a href="/admin/espelho/export?usuario_id={{ usuario.id }}&tipo=txt&mes={{ mes_pad }}&ano={{ ano }}">📝 TXT</a>
                        <a href="/admin/espelho/export?usuario_id={{ usuario.id }}&tipo=csv&mes={{ mes_pad }}&ano={{ ano }}">📑 CSV</a>
                    </div>
                </div>
                <form method="GET" action="/admin/espelho" class="filter-form">
                    <div style="display: flex; gap: 12px; align-items: center; flex-wrap: wrap; width: 100%;">
                        <div style="display: flex; flex-direction: column; gap: 4px; flex: 1; min-width: 200px;">
                            <label style="font-size: 12px; color: rgba(255, 255, 255, 0.7); font-weight: 600;">👤 Funcionário</label>
                            <select name="usuario_id" onchange="this.form.submit()" style="padding: 10px 14px; border-radius: 10px; border: 1px solid rgba(0, 198, 255, 0.3); background: rgba(255, 255, 255, 0.08); color: #fff; font-weight: 500; width: 100%;">
                                <option value="">Selecione um funcionário</option>
                                {% for f in funcionarios %}
                                <option value="{{ f.id }}" {% if usuario.id|string == f.id|string %}selected{% endif %}>{{ f.nome }}</option>
                                {% endfor %}
                            </select>
                        </div>
                        
                        <div style="display: flex; flex-direction: column; gap: 4px;">
                            <label style="font-size: 12px; color: rgba(255, 255, 255, 0.7); font-weight: 600;">📅 Mês</label>
                            <select name="mes" onchange="this.form.submit()" style="padding: 10px 14px; border-radius: 10px; border: 1px solid rgba(0, 198, 255, 0.3); background: rgba(255, 255, 255, 0.08); color: #fff; font-weight: 500; min-width: 150px;">
                                <option value="01" {% if mes_pad == '01' %}selected{% endif %}>Janeiro</option>
                                <option value="02" {% if mes_pad == '02' %}selected{% endif %}>Fevereiro</option>
                                <option value="03" {% if mes_pad == '03' %}selected{% endif %}>Março</option>
                                <option value="04" {% if mes_pad == '04' %}selected{% endif %}>Abril</option>
                                <option value="05" {% if mes_pad == '05' %}selected{% endif %}>Maio</option>
                                <option value="06" {% if mes_pad == '06' %}selected{% endif %}>Junho</option>
                                <option value="07" {% if mes_pad == '07' %}selected{% endif %}>Julho</option>
                                <option value="08" {% if mes_pad == '08' %}selected{% endif %}>Agosto</option>
                                <option value="09" {% if mes_pad == '09' %}selected{% endif %}>Setembro</option>
                                <option value="10" {% if mes_pad == '10' %}selected{% endif %}>Outubro</option>
                                <option value="11" {% if mes_pad == '11' %}selected{% endif %}>Novembro</option>
                                <option value="12" {% if mes_pad == '12' %}selected{% endif %}>Dezembro</option>
                            </select>
                        </div>
                        
                        <div style="display: flex; flex-direction: column; gap: 4px;">
                            <label style="font-size: 12px; color: rgba(255, 255, 255, 0.7); font-weight: 600;">📆 Ano</label>
                            <select name="ano" onchange="this.form.submit()" style="padding: 10px 14px; border-radius: 10px; border: 1px solid rgba(0, 198, 255, 0.3); background: rgba(255, 255, 255, 0.08); color: #fff; font-weight: 500; min-width: 120px;">
                                {% for y in range(2024, 2031) %}
                                <option value="{{ y }}" {% if ano|string == y|string %}selected{% endif %}>{{ y }}</option>
                                {% endfor %}
                            </select>
                        </div>
                    </div>
                </form>
                <form method="POST" action="/admin/espelho/editar_lote">
                    <input type="hidden" name="usuario_id" value="{{ usuario.id }}">
                    <table>
                        <thead>
                            <tr>
                                <th>Data e Hora</th>
                                <th>Tipo</th>
                                <th>Justificativa</th>
                                <th>Feriado</th>
                                <th>Abono</th>
                                <th>Ações</th>
                            </tr>
                        </thead>
                        <tbody>
                            {% for dia, regs in registros_por_dia.items() %}
                                {% for r in regs %}
                                <tr>
                                    <td>
                                        <input type="hidden" name="registro_id[]" value="{{ r.id }}">
                                        <input type="datetime-local" name="data_hora[]" value="{{ r.data_hora[:10] }}T{{ r.data_hora[11:16] }}">
                                    </td>
                                    <td>
                                        <select name="tipo[]">
                                            <option value="entrada" {% if r.tipo=='entrada' %}selected{% endif %}>Entrada</option>
                                            <option value="saida" {% if r.tipo=='saida' %}selected{% endif %}>Saída</option>
                                            <option value="abono" {% if r.tipo=='abono' %}selected{% endif %}>Abono</option>
                                        </select>
                                    </td>
                                    <td>
                                        <input type="text" name="justificativa[]" value="{{ r.justificativa }}">
                                    </td>
                                    <td>
                                        <button type="button" class="btn-feriado" data-dia="{{ dia }}" style="background:#ffeaea;color:#c00;border:none;padding:6px 12px;border-radius:6px;{% if dia in feriados_por_dia %}opacity:1;{% else %}opacity:0.5;{% endif %}">
                                            {% if dia in feriados_por_dia %}Desmarcar Feriado{% else %}Marcar Feriado{% endif %}
                                        </button>
                                    </td>
                                    <td>
                                        <button type="button" class="btn-abono" data-dia="{{ dia }}" data-usuario="{{ usuario.id }}" style="background:#eaffea;color:#0a0;border:none;padding:6px 12px;border-radius:6px;{% if dia in abonos_por_dia %}opacity:1;{% else %}opacity:0.5;{% endif %}">
                                            {% if dia in abonos_por_dia %}Desmarcar Abono{% else %}Marcar Abono{% endif %}
                                        </button>
                                    </td>
                                    <td>
                                        <button
                                            type="submit"
                                            class="btn-del"
                                            formaction="/admin/espelho/deletar_registro"
                                            formmethod="post"
                                            name="registro_id"
                                            value="{{ r.id }}"
                                            onclick="return confirm('Confirma excluir esta batida?');"
                                        >🗑 Excluir</button>
                                    </td>
                                </tr>
                                {% endfor %}
                            {% endfor %}
                        </tbody>
                    </table>
                    <div style="margin-top:16px;text-align:right;">
                        <button type="submit" class="btn-edit">💾 Salvar todas alterações</button>
                    </div>
                </form>
                <div class="legenda">
                    <span style="background:#ffeaea;padding:4px 8px;border-radius:6px;color:#c00;">Feriado</span>
                    <span style="background:#eaffea;padding:4px 8px;border-radius:6px;color:#0a0;">Abono</span>
                    <span style="border:2px solid #0072ff;padding:4px 8px;border-radius:6px;">Hoje</span>
                </div>
            </div>
        </main>
    </div>
    <script>
    document.querySelectorAll('.btn-feriado').forEach(btn => {
        btn.addEventListener('click', function() {
            const dia = this.getAttribute('data-dia');
            fetch('/admin/toggle_feriado', {
                method: 'POST',
                headers: {'Content-Type': 'application/json'},
                body: JSON.stringify({data: dia})
            }).then(r => r.json()).then(d => { location.reload(); });
        });
    });
    document.querySelectorAll('.btn-abono').forEach(btn => {
        btn.addEventListener('click', function() {
            const dia = this.getAttribute('data-dia');
            const usuario_id = this.getAttribute('data-usuario');
            fetch('/admin/toggle_abono', {
                method: 'POST',
                headers: {'Content-Type': 'application/json'},
                body: JSON.stringify({data: dia, usuario_id: usuario_id})
            }).then(r => r.json()).then(d => { location.reload(); });
        });
    });

    setTimeout(() => {
        document.querySelectorAll('.flash-msg').forEach(el => {
            el.classList.add('hide');
            setTimeout(() => {
                if (el && el.parentNode) {
                    el.parentNode.removeChild(el);
                }
            }, 380);
        });
    }, 3500);
    </script>
</body>
</html>'''
    return render_template_string(template, usuario=usuario, mes_pad=mes_pad, ano=ano, registros_por_dia=registros_por_dia, feriados_por_dia=feriados_por_dia, abonos_por_dia=abonos_por_dia, funcionarios=funcionarios, sucesso=sucesso, erro=erro)

@app.route('/admin/espelho/export')
def admin_espelho_export():
    if session.get("tipo") != "admin":
        return redirect("/")

    usuario_id = request.args.get("usuario_id")
    tipo = request.args.get("tipo", "csv").lower()
    ano = request.args.get("ano") or str(datetime.now().year)
    mes = request.args.get("mes") or str(datetime.now().month).zfill(2)

    conn = get_db()
    c = conn.cursor()
    # Buscar dados completos do funcionário
    usuario_row = c.execute("SELECT id, nome, cpf, jornada_diaria, departamento, pis FROM usuarios WHERE id=?", (usuario_id,)).fetchone()
    usuario = dict(usuario_row) if usuario_row else {'id': '', 'nome': '', 'cpf': '', 'jornada_diaria': '', 'departamento': '', 'pis': ''}
    # Buscar dados da empresa
    empresa_row = c.execute("SELECT nome FROM empresa LIMIT 1").fetchone()
    empresa_nome = empresa_row[0] if empresa_row else 'EMPRESA'

    registros = c.execute(
        "SELECT r.id, r.tipo, r.data_hora, r.justificativa FROM registros r WHERE r.usuario_id=? AND strftime('%Y', r.data_hora)=? AND strftime('%m', r.data_hora)=? ORDER BY r.data_hora",
        (usuario_id, ano, mes)
    ).fetchall() if usuario_id else []
    
    # Carregar apenas feriados nacionais fixos
    # Feriados nacionais brasileiros: 01/01, 21/04, 01/05, 07/09, 12/10, 02/11, 15/11, 25/12
    # Mais 02/11 (Finados) se ainda não incluído
    feriados_set = set()
    feriados_nacionais = [
        (1, 1),    # Confraternização Universal
        (4, 21),   # Tiradentes
        (5, 1),    # Dia do Trabalho
        (9, 7),    # Independência do Brasil
        (10, 12),  # Nossa Senhora Aparecida
        (11, 2),   # Finados
        (11, 15),  # Proclamação da República
        (12, 25),  # Natal
    ]
    # Converter para formato YYYY-MM-DD
    for mes_feriado, dia_feriado in feriados_nacionais:
        data_feriado = f"{ano}-{mes_feriado:02d}-{dia_feriado:02d}"
        feriados_set.add(data_feriado)
    
    # Mapear abonos por data
    abonos_map = {}
    for r in registros:
        if r[1] == 'abono':  # tipo = abono
            data = r[2].split(' ')[0]  # YYYY-MM-DD
            abonos_map[data] = True
    conn.close()

    # Geração dos dados para exportação
    output = io.StringIO()
    filename = f"espelho_{usuario.get('nome','usuario')}_{mes}_{ano}.{tipo}"
    output = io.StringIO()
    filename = f"espelho_{usuario.get('nome','usuario')}_{mes}_{ano}.{tipo}"
    # Bloco de informações para todos os formatos
    info_empresa = f"EMPRESA: {empresa_nome}"
    info_departamento = f"DEPARTAMENTO: {usuario.get('departamento', '') or '--'}"
    info_funcionario = f"FUNCIONÁRIO: {usuario.get('nome', '')}"
    info_pis = f"PIS: {usuario.get('pis', '') or '--'}"
    meses_pt = {
        "01": "Janeiro", "02": "Fevereiro", "03": "Março", "04": "Abril",
        "05": "Maio", "06": "Junho", "07": "Julho", "08": "Agosto",
        "09": "Setembro", "10": "Outubro", "11": "Novembro", "12": "Dezembro"
    }
    dias_semana_pt = [
        "Segunda-feira", "Terça-feira", "Quarta-feira", "Quinta-feira",
        "Sexta-feira", "Sábado", "Domingo"
    ]
    info_periodo = f"PERÍODO: Mês: {meses_pt.get(mes, mes)} / {ano}"
    info_tipo = "TIPO: NORMAL"
    info_cpf = f"CPF: {usuario.get('cpf', '') or '--'}"
    info_jornada = f"JORNADA: {usuario.get('jornada_diaria', '') or '--'}h"

    if tipo == "csv":
        writer = csv.writer(output, delimiter=';', lineterminator='\n')
        writer.writerow([info_empresa, info_periodo])
        writer.writerow([info_departamento, info_tipo])
        writer.writerow([info_funcionario, info_cpf])
        writer.writerow([info_pis, info_jornada])
        writer.writerow([])
        writer.writerow(["ID", "Tipo", "Data/Hora", "Justificativa", "Status"])
        for r in registros:
            data_dia = r[2][:10] if r[2] else ""  # YYYY-MM-DD
            status = ""
            if r[1] == 'abono':
                status = "ABONADO"
            elif data_dia in feriados_set:
                status = "FERIADO"
            writer.writerow([r[0], r[1], r[2], r[3], status])
        mimetype = "text/csv"
    elif tipo == "txt":
        output.write(f"{info_empresa}\t{info_periodo}\n")
        output.write(f"{info_departamento}\t{info_tipo}\n")
        output.write(f"{info_funcionario}\t{info_cpf}\n")
        output.write(f"{info_pis}\t{info_jornada}\n\n")
        output.write("ID\tTipo\tData/Hora\tJustificativa\tStatus\n")
        for r in registros:
            data_dia = r[2][:10]  # YYYY-MM-DD
            status = ""
            if data_dia in feriados_set:
                status = "FERIADO"
            elif r[1] == 'abono':
                status = "ABONADO"
            output.write(f"{r[0]}\t{r[1]}\t{r[2]}\t{r[3]}\t{status}\n")
        mimetype = "text/plain"
    elif tipo == "excel":
        import openpyxl
        from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Espelho de Ponto"
        
        # Configurar opções avançadas do Excel
        # Usar sistema de data 1900 (padrão)
        wb.date1904 = False
        
        # Configurar cálculo para salvar valores de vínculos externos
        wb.calculation.calcMode = 'auto'
        wb.calculation.fullCalcOnLoad = True
        
        # Título centralizado
        ws.merge_cells('A1:O1')
        title_cell = ws['A1']
        title_cell.value = 'ESPELHO DE PONTO'
        title_cell.font = Font(bold=True, size=14, color='FFFFFFFF')
        title_cell.fill = PatternFill(start_color='FF4472C4', end_color='FF4472C4', fill_type='solid')
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 25
        
        # Bloco de informações (lado esquerdo e direito)
        ws['A3'] = info_empresa
        ws['A4'] = info_departamento
        ws['A5'] = info_funcionario
        ws['A6'] = info_pis
        
        ws['H3'] = info_periodo
        ws['H4'] = info_tipo
        ws['H5'] = info_cpf
        ws['H6'] = info_jornada
        
        for row in range(3, 7):
            for col in ['A', 'H']:
                cell = ws[f'{col}{row}']
                cell.font = Font(size=10)
        
        # Cabeçalhos da tabela
        header_row = 8
        headers = [
            'Data', 'Dia', '1ª E.', '1ª S.', '2ª E.', '2ª S.', '3ª E.', '3ª S.', '4ª E.',
            'Abono', 'Jornada', 'H.N.', 'H.E.', 'Saldo', 'Observação'
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=header_row, column=col, value=header)
            cell.font = Font(bold=True, color='FFFFFFFF', size=10)
            cell.fill = PatternFill(start_color='FF002060', end_color='FF002060', fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = Border(
                left=Side(border_style='thin', color='FF000000'),
                right=Side(border_style='thin', color='FF000000'),
                top=Side(border_style='thin', color='FF000000'),
                bottom=Side(border_style='thin', color='FF000000')
            )
        
        ws.row_dimensions[header_row].height = 25
        
        # Definir larguras das colunas
        col_widths = [12, 15, 9, 9, 9, 9, 9, 9, 9, 9, 10, 9, 9, 10, 18]
        for col, width in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(col)].width = width
        
        # Preencher dados dos dias do mês
        thin_border = Border(
            left=Side(border_style='thin', color='FF000000'),
            right=Side(border_style='thin', color='FF000000'),
            top=Side(border_style='thin', color='FF000000'),
            bottom=Side(border_style='thin', color='FF000000')
        )
        
        font_normal = Font(size=10)
        font_normal_black = Font(color='FF000000', size=10)  # Preto para horário normal
        font_green = Font(color='FF00B050', size=10)  # Verde para horas extras
        
        # Função para detectar se um horário é hora extra
        def is_hora_extra(horario_str, weekday):
            """
            Verifica se um horário está fora da jornada de trabalho.
            Jornada: 08:00-12:00 e 13:00-18:00 (segunda a sexta apenas)
            Sábado e domingo: QUALQUER horário é hora extra
            
            weekday: 0=seg, 1=ter, 2=qua, 3=qui, 4=sex, 5=sab, 6=dom
            Retorna True se for hora extra (verde), False se for normal (preto)
            """
            try:
                hm = horario_str.strip()
                if not hm or hm == '00:00':
                    return False
                
                # Finais de semana: TUDO é hora extra
                if weekday >= 5:  # Sábado (5) ou Domingo (6)
                    return True
                
                # Segunda a sexta: aplicar lógica de horário
                h, m = map(int, hm.split(':'))
                hora_minuto = h * 60 + m
                
                # Jornada normal: 08:00-12:00 (480-720 min) e 13:00-18:00 (780-1080 min)
                turno1_inicio = 8 * 60      # 08:00
                turno1_fim = 12 * 60       # 12:00
                turno2_inicio = 13 * 60    # 13:00
                turno2_fim = 18 * 60       # 18:00
                
                # Se está dentro de um dos turnos, não é hora extra
                if (turno1_inicio <= hora_minuto <= turno1_fim or 
                    turno2_inicio <= hora_minuto <= turno2_fim):
                    return False
                else:
                    return True  # Fora do horário normal = hora extra
            except Exception:
                return False
        
        dias_mes = calendar.monthrange(int(ano), int(mes))[1]
        
        for dia in range(1, dias_mes + 1):
            data = datetime(int(ano), int(mes), dia)
            row_idx = header_row + dia
            data_str = f"{ano}-{mes}-{str(dia).zfill(2)}"
            
            # Data
            data_cell = ws.cell(row=row_idx, column=1, value=data.strftime('%d/%m/%Y'))
            data_cell.border = thin_border
            data_cell.alignment = Alignment(horizontal='center')
            data_cell.font = font_normal
            
            # Dia da semana
            dia_cell = ws.cell(row=row_idx, column=2, value=dias_semana_pt[data.weekday()])
            dia_cell.border = thin_border
            dia_cell.alignment = Alignment(horizontal='center')
            dia_cell.font = font_normal
            
            # Batidas (entrada/saída)
            batidas_dia = [r for r in registros if r[2].startswith(data_str) and r[1] not in ('abono', 'feriado')]
            entrada_saida = []
            for r in batidas_dia:
                if r[1] in ('entrada', 'saida'):
                    entrada_saida.append(r[2][11:16])
            
            # Preencher horários nas colunas 3-9 (1ªE, 1ªS, 2ªE, 2ªS, 3ªE, 3ªS, 4ªE)
            for col_idx, horario in enumerate(entrada_saida[:9], 3):
                hora_cell = ws.cell(row=row_idx, column=col_idx, value=horario)
                hora_cell.border = thin_border
                hora_cell.alignment = Alignment(horizontal='center')
                # Verde para hora extra, preto para hora normal
                if is_hora_extra(horario, data.weekday()):
                    hora_cell.font = font_green  # Verde = hora extra
                else:
                    hora_cell.font = font_normal_black  # Preto = hora normal
            
            # Preencher células vazias
            for col_idx in range(3 + len(entrada_saida), 10):
                vazio = ws.cell(row=row_idx, column=col_idx)
                vazio.border = thin_border
            
            # Coluna 10: Abono (calcula período abonado baseado nas batidas)
            abono_cell = ws.cell(row=row_idx, column=10)
            tem_abono = data_str in abonos_map
            if tem_abono:
                # Se tem batidas no dia, calcular meio período
                if len(entrada_saida) > 0:
                    # Calcular tempo trabalhado (em minutos)
                    tempo_trabalhado_min = 0
                    for i in range(0, len(entrada_saida), 2):
                        if i + 1 < len(entrada_saida):
                            try:
                                entrada_parts = entrada_saida[i].split(':')
                                saida_parts = entrada_saida[i+1].split(':')
                                entrada_min = int(entrada_parts[0]) * 60 + int(entrada_parts[1])
                                saida_min = int(saida_parts[0]) * 60 + int(saida_parts[1])
                                tempo_trabalhado_min += (saida_min - entrada_min)
                            except Exception:
                                pass
                    
                    # Jornada completa = 9h = 540 minutos
                    jornada_min = 9 * 60
                    # Abono = jornada - tempo trabalhado
                    abono_min = jornada_min - tempo_trabalhado_min
                    
                    if abono_min > 0:
                        # Converte para horas:minutos
                        abono_h = abono_min // 60
                        abono_m = abono_min % 60
                        abono_cell.value = f"{abono_h:02d}:{abono_m:02d}"
                    else:
                        abono_cell.value = '00:00'
                else:
                    # Sem batidas = abono integral
                    abono_cell.value = '09:00'
                
                abono_cell.font = font_green
            abono_cell.border = thin_border
            abono_cell.alignment = Alignment(horizontal='center')
            
            # Coluna 11: Jornada (fixa em 09:00:00 - vazia em sábado e domingo)
            jornada_cell = ws.cell(row=row_idx, column=11)
            jornada_cell.border = thin_border
            jornada_cell.alignment = Alignment(horizontal='center')
            jornada_cell.font = font_normal
            # Não preencher Jornada em sábado (5) ou domingo (6)
            is_weekend = data.weekday() >= 5
            if not is_weekend:  # Segunda a sexta (0-4)
                jornada_cell.value = '09:00:00'
                jornada_cell.number_format = 'hh:mm:ss'
                print(f"JORNADA PREENCHIDA: {data.strftime('%d/%m')} weekday={data.weekday()}")
            else:
                print(f"JORNADA VAZIA: {data.strftime('%d/%m')} weekday={data.weekday()}")
            
            # Coluna 12: H.N. - Fórmula para calcular horas trabalhadas
            # Fórmula: =SUM(F-E,D-C) onde D,C são 1ªS,1ªE; F,E são 2ªS,2ªE
            col_letter_C = get_column_letter(3)   # 1ªE
            col_letter_D = get_column_letter(4)   # 1ªS
            col_letter_E = get_column_letter(5)   # 2ªE
            col_letter_F = get_column_letter(6)   # 2ªS
            
            hn_formula = f"=SUM({col_letter_F}{row_idx}-{col_letter_E}{row_idx},{col_letter_D}{row_idx}-{col_letter_C}{row_idx})"
            hn_cell = ws.cell(row=row_idx, column=12, value=hn_formula)
            hn_cell.border = thin_border
            hn_cell.alignment = Alignment(horizontal='center')
            hn_cell.font = font_normal
            hn_cell.number_format = 'hh:mm:ss'  # Formato hora
            
            # Coluna 13: H.E. - Fórmula para calcular horas extras
            # Fórmula: =SUM(J+L)-K onde J=Abono, L=H.N., K=Jornada
            col_letter_J = get_column_letter(10)  # Abono
            col_letter_K = get_column_letter(11)  # Jornada
            col_letter_L = get_column_letter(12)  # H.N.
            
            he_formula = f"=SUM({col_letter_J}{row_idx}+{col_letter_L}{row_idx})-{col_letter_K}{row_idx}"
            he_cell = ws.cell(row=row_idx, column=13, value=he_formula)
            he_cell.border = thin_border
            he_cell.alignment = Alignment(horizontal='center')
            he_cell.number_format = 'hh:mm:ss'  # Formato hora
            he_cell.font = font_normal
            
            # Coluna 14: Saldo (vazio por enquanto)
            saldo_cell = ws.cell(row=row_idx, column=14)
            saldo_cell.border = thin_border
            saldo_cell.alignment = Alignment(horizontal='center')
            
            # Coluna 15: Observação
            obs_cell = ws.cell(row=row_idx, column=15)
            if data_str in feriados_set:
                obs_cell.value = 'Feriado'
                obs_cell.font = Font(color='FFC00000', size=10, bold=True)
            elif data_str in abonos_map:
                obs_cell.value = 'Abono'
                obs_cell.font = font_green
            obs_cell.border = thin_border
            obs_cell.alignment = Alignment(horizontal='left')
        
        xlsx_stream = io.BytesIO()
        wb.save(xlsx_stream)
        xlsx_stream.seek(0)
        mimetype = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        filename = filename.replace('.excel', '.xlsx')
        from flask import send_file
        return send_file(
            xlsx_stream,
            mimetype=mimetype,
            as_attachment=True,
            download_name=filename
        )
    elif tipo == "afd":
        # Exportação no layout AFD conforme o print
        # Cada registro: [número do registro][PIS][data][hora][número do cartão de ponto]
        # Buscar PIS e número do cartão do funcionário
        pis = usuario.get('pis', '').zfill(11) if usuario.get('pis', '') else '0'*11
        cartao_ponto = str(usuario.get('id', '')).zfill(12) if usuario.get('id', '') else '0'*12
        for idx, r in enumerate(registros, 1):
            # Número do registro (sequencial, 9 dígitos)
            num_registro = str(idx).zfill(9)
            # Data (formato ddMMyyyy)
            data = ''
            hora = ''
            if r[2]:
                try:
                    dt = datetime.strptime(r[2][:19], '%Y-%m-%d %H:%M:%S')
                except Exception:
                    dt = None
                if dt:
                    data = dt.strftime('%d%m%Y')
                    hora = dt.strftime('%H%M')
            # Linha AFD
            linha = f"{num_registro}{pis}{data}{hora}{cartao_ponto}\n"
            output.write(linha)
        mimetype = "text/plain"
    else:
        return "Formato não suportado", 400

    output.seek(0)
    from flask import send_file
    return send_file(
        io.BytesIO(output.getvalue().encode('utf-8')),
        mimetype=mimetype,
        as_attachment=True,
        download_name=filename
    )

# ==========================================================
# ROTAS DE TOGGLE (fora de qualquer função)
# ==========================================================
@app.route('/admin/toggle_feriado', methods=['POST'])
def toggle_feriado():
    if session.get('tipo') != 'admin':
        return {'ok': False}, 403
    # Apenas feriados nacionais são permitidos
    return {'ok': False, 'msg': 'Apenas feriados nacionais são permitidos. Para adicionar feriados especiais, edite o código.'}, 400

@app.route('/admin/toggle_abono', methods=['POST'])
def toggle_abono():
    if session.get('tipo') != 'admin':
        return {'ok': False}, 403
    data = request.get_json()
    dia = data.get('data')
    usuario_id = data.get('usuario_id')
    if not dia or not usuario_id:
        return {'ok': False, 'msg': 'Dados insuficientes'}, 400
    conn = get_db()
    # Verifica se já existe abono para o dia
    abono = conn.execute("SELECT id FROM registros WHERE usuario_id=? AND tipo='abono' AND data_hora LIKE ?", (usuario_id, dia+'%')).fetchone()
    if abono:
        conn.execute("DELETE FROM registros WHERE id=?", (abono['id'],))
    else:
        conn.execute("INSERT INTO registros (usuario_id, tipo, data_hora, justificativa) VALUES (?, 'abono', ?, '')", (usuario_id, dia+' 00:00:00'))
    conn.commit()
    conn.close()
    return {'ok': True}


# ==========================================================
# FERIADOS - Função para verificar feriado completo
# ==========================================================
def is_feriado_completo(dt):
    # Exemplo: lista de feriados nacionais fixos
    feriados = [
        (1, 1),    # Ano Novo
        (4, 21),   # Tiradentes
        (5, 1),    # Dia do Trabalho
        (9, 7),    # Independência
        (10, 12),  # Nossa Senhora Aparecida
        (11, 2),   # Finados
        (11, 15),  # Proclamação da República
        (12, 25),  # Natal
    ]
    # Adicione outros feriados ou lógica para calcular móveis (ex: Carnaval, Páscoa)
    return (dt.month, dt.day) in feriados

# ==========================================================
# DASHBOARD COM GRÁFICOS
# ==========================================================

@app.route("/dashboard")
def dashboard():
    if session.get("tipo") != "admin":
        return redirect("/")
    
    conn = get_db()
    
    # Obter parâmetros de filtro
    filtro = request.args.get('filtro', 'mes')  # dia, semana, mes, ano
    usuario_id = request.args.get('usuario_id', 'all')
    # Removido variáveis não utilizadas: mes, ano, funcionarios
    
    # Preparar dados para gráficos
    labels = []
    horas = []
    
    if filtro == 'dia':
        # Últimos 7 dias
        for i in range(6, -1, -1):
            dia = datetime.now() - timedelta(days=i)
            labels.append(dia.strftime('%d/%m'))
            query = "SELECT * FROM registros WHERE date(data_hora)=?"
            params = [dia.date()]
            if usuario_id != 'all':
                query += SQL_AND_USUARIO_ID
                params.append(usuario_id)
            registros = conn.execute(query, params).fetchall()
            total = calcular_horas_do_dia(registros)
            horas.append(total)
    elif filtro == 'semana':
        # Últimas 4 semanas
        for i in range(3, -1, -1):
            semana_inicio = (datetime.now() - timedelta(days=datetime.now().weekday())) - timedelta(weeks=i)
            semana_fim = semana_inicio + timedelta(days=6)
            labels.append(f"{semana_inicio.strftime('%d/%m')} - {semana_fim.strftime('%d/%m')}")
            query = "SELECT * FROM registros WHERE date(data_hora) BETWEEN ? AND ?"
            params = [semana_inicio.date(), semana_fim.date()]
            if usuario_id != 'all':
                query += " AND usuario_id=?"
                params.append(usuario_id)
            registros = conn.execute(query, params).fetchall()
            total = calcular_horas_periodo(registros)
            horas.append(total)
    elif filtro == 'mes':
        # Últimos 6 meses
        for i in range(5, -1, -1):
            mes = (datetime.now().month - i - 1) % 12 + 1
            ano = datetime.now().year if datetime.now().month - i > 0 else datetime.now().year - 1
            labels.append(f"{mes:02d}/{ano}")
            query = "SELECT * FROM registros WHERE strftime('%m', data_hora)=? AND strftime('%Y', data_hora)=?"
            params = [f"{mes:02d}", str(ano)]
            if usuario_id != 'all':
                query += " AND usuario_id=?"
                params.append(usuario_id)
            registros = conn.execute(query, params).fetchall()
            total = calcular_horas_periodo(registros)
            horas.append(total)
    elif filtro == 'ano':
        # Últimos 3 anos
        for i in range(2, -1, -1):
            ano = datetime.now().year - i
            labels.append(str(ano))
            query = "SELECT * FROM registros WHERE strftime('%Y', data_hora)=?"
            params = [str(ano)]
            if usuario_id != 'all':
                query += " AND usuario_id=?"
                params.append(usuario_id)
            registros = conn.execute(query, params).fetchall()
            total = calcular_horas_periodo(registros)
            horas.append(total)
    else:
        # fallback
        pass

    # Buscar lista de funcionários
    funcionarios = conn.execute("SELECT id, nome FROM usuarios").fetchall()
    conn.close()

    return render_template_string(
        DASHBOARD_HTML,
        labels=labels,
        horas=horas,
        filtro_atual=filtro,
        usuario_id=usuario_id,
        funcionarios=funcionarios
    )
def calcular_horas_periodo(registros):
    """Calcula total de horas de um período (múltiplos dias)"""
    dias = {}
    for r in registros:
        data = r['data_hora'][:10]
        if data not in dias:
            dias[data] = []
        dias[data].append(r)
    total = 0
    for dia_registros in dias.values():
        total += calcular_horas_do_dia(dia_registros)
    return round(total, 2)


# ==========================================================
@app.route("/logout")
def logout():
    session.clear()
    return redirect("/")
# ==========================================================

if __name__ == "__main__":
    # Removido init_db() para evitar erro NameError

    # Configurar SSL com certificados autossinados
    import os
    from werkzeug.middleware.proxy_fix import ProxyFix
    
    certs_dir = os.path.join(os.path.dirname(__file__), 'certs')
    cert_file = os.path.join(certs_dir, 'ponto.crt')
    key_file = os.path.join(certs_dir, 'ponto.key')
    
    # Verificar se os certificados existem
    use_ssl = os.path.exists(cert_file) and os.path.exists(key_file)
    ssl_context = None
    if use_ssl:
        ssl_context = (cert_file, key_file)
        print("[SSL] OK - Certificados encontrados - usando HTTPS")
    else:
        print("[SSL] AVISO - Certificados nao encontrados - usando HTTP")

    # Criar aplicação para HTTP que redireciona para HTTPS
    app_http_redirect = Flask(__name__)
    
    @app_http_redirect.route('/', defaults={'path': ''})
    @app_http_redirect.route('/<path:path>')
    def redirect_to_https(path):
        # Obter host e porta do request
        host = request.host.split(':')[0]
        # Determinar porta HTTPS baseado na porta HTTP
        http_port = int(request.host.split(':')[1])
        https_port = 5050 if http_port == 6050 else 5000
        
        url = 'https://{}:{}/{}'.format(host, https_port, path)
        if request.query_string:
            url = url + '?' + request.query_string.decode()
        return redirect(url, code=301)

    # Executar a mesma aplicação em duas portas:
    # - 5000: porta principal HTTPS de desenvolvimento
    # - 5050: porta dedicada HTTPS para admin/login
    # - 6000: porta HTTP que redireciona para 5000
    # - 6050: porta HTTP que redireciona para 5050
    def run_port(app, port, debug_mode=False, ssl_ctx=None):
        try:
            app.run(host='0.0.0.0', port=port, debug=debug_mode, use_reloader=False, ssl_context=ssl_ctx)
        except OSError as e:
            print("[ERROR] Porta {} em uso: {}".format(port, e))

    # Iniciar servidor de redirect HTTP em background (porta 6000 -> 5000)
    if use_ssl:
        t1 = threading.Thread(target=run_port, args=(app_http_redirect, 6000, False, None), daemon=True)
        t1.start()
        
        # Iniciar servidor de redirect HTTP para admin em background (porta 6050 -> 5050)
        t2 = threading.Thread(target=run_port, args=(app_http_redirect, 6050, False, None), daemon=True)
        t2.start()

    # Iniciar servidor HTTPS principal em background (port 5050, admin)
    t3 = threading.Thread(target=run_port, args=(app, 5050, False, ssl_context), daemon=True)
    t3.start()

    proto = "https" if use_ssl else "http"
    print("")
    print("Servidor iniciado:")
    print("  Funcionario: " + proto + "://ponto.local:5000")
    print("  Admin:       " + proto + "://ponto.admin:5050")
    if use_ssl:
        print("")
        print("Redirects HTTP automaticos:")
        print("  http://ponto.local:5000     (porta 6000) -> https://ponto.local:5000")
        print("  http://ponto.admin:5050     (porta 6050) -> https://ponto.admin:5050")
    print("")
    
    # Rodar servidor HTTPS principal (bloqueante - porta 5000)
    run_port(app, 5000, True, ssl_context)





