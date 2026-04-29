#!/usr/bin/env python
# -*- coding: utf-8 -*-
import requests
import openpyxl
import time

print("Iniciando teste de formatação...")

# Criar sessão para manter cookie
session = requests.Session()

# Login
print("1. Fazendo login...")
login_data = {'cpf': '00000000000', 'senha': 'Pont0!2024@Admin#Secure'}
r = session.post('http://localhost:5050/admin', data=login_data, verify=False, timeout=10)
print(f"   Status: {r.status_code}")

# Dar um tempo para a sessão ser processada
time.sleep(0.5)

# Export Excel
print("2. Exportando arquivo...")
export_url = 'http://localhost:5050/admin/espelho/export?tipo=excel&mes=2&ano=2026'
r = session.get(export_url, verify=False, timeout=10)
print(f"   Status: {r.status_code}")
print(f"   Tamanho: {len(r.content)} bytes")

# Verificar se é um ZIP válido (Excel é ZIP)
if r.content[:2] == b'PK':
    print("   ✅ Arquivo é um ZIP válido")
else:
    print(f"   ❌ Arquivo começa com: {r.content[:50]}")
    exit(1)

# Salvar arquivo
print("3. Salvando arquivo...")
with open('teste_formato_hora.xlsx', 'wb') as f:
    f.write(r.content)
print(f"   ✅ Arquivo salvo com sucesso")

# Verificar formatação
print("\n4. Verificando formatação...")
wb = openpyxl.load_workbook('teste_formato_hora.xlsx')
ws = wb.active

print("\n=== COLUNAS L E M (H.N. e H.E.) ===")
print("\nFormatos encontrados:")
for row in range(9, 25):
    hn_fmt = ws[f'L{row}'].number_format
    he_fmt = ws[f'M{row}'].number_format
    if hn_fmt and 'h' in hn_fmt.lower():
        print(f"  Linha {row}: H.N.={hn_fmt} ✅ | H.E.={he_fmt} ✅")
        break

print("\n=== COLUNA K (JORNADA) - Verificando sábado/domingo ===")
print("\nData | Dia | Jornada")
print("-----|-----|--------")

for row in range(9, 39):  # Percorrer todo o mês
    data_cell = ws[f'A{row}'].value
    day_cell = ws[f'B{row}'].value
    jornada_cell = ws[f'K{row}'].value
    
    if data_cell:
        # Verificar dia da semana
        is_weekend = day_cell and day_cell.lower() in ('sábado', 'domingo', 'sabado')
        jornada_str = str(jornada_cell) if jornada_cell else "(vazio)"
        
        marker = "✅" if (is_weekend and not jornada_cell) or (not is_weekend and jornada_cell) else "❌"
        if is_weekend or day_cell and day_cell.lower() in ('segunda', 'terça', 'terca'):
            print(f"{data_cell} | {day_cell:10} | {jornada_str:15} {marker}")

print("\n✅ Teste concluído!")
