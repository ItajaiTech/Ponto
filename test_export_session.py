#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
Teste com manutenção de sessão aprimorada
"""
import requests
import sys

# Desabilitar SSL warnings
import urllib3
urllib3.disable_warnings()

print("Iniciando teste de export...")

# Criar session com headers mantidos
session = requests.Session()

# Fazer POST login COM esperança que mantenha a sessão
print("\n1. Fazendo login...")
login_url = 'http://localhost:5050/admin'
login_data = {'cpf': '00000000000', 'senha': 'Pont0!2024@Admin#Secure'}

r_login = session.post(login_url, data=login_data, allow_redirects=True)
print(f"   POST Status: {r_login.status_code}")
print(f"   Cookies na sessão: {session.cookies}")
print(f"   URL final: {r_login.url}")

# Tentar export IMEDIATAMENTE após login  
print("\n2. Exportando Excel (mesma sessão)...")
export_url = 'http://localhost:5050/admin/espelho/export?tipo=excel&mes=2&ano=2026'

r_export = session.get(export_url, allow_redirects=False)
print(f"   GET Status: {r_export.status_code}")
print(f"   Content-Type: {r_export.headers.get('content-type', 'N/A')}")
print(f"   Tamanho: {len(r_export.content)} bytes")

# Verificar se é ZIP
is_zip = r_export.content[:2] == b'PK'
is_html = b'<html>' in r_export.content[:200]

print(f"   É ZIP (Excel)? {is_zip}")
print(f"   É HTML? {is_html}")

if is_zip:
    print("\n✅ SUCESSO: Arquivo exportado como ZIP (Excel)")
    with open('espelho_final_teste.xlsx', 'wb') as f:
        f.write(r_export.content)
    print("   Salvo como: espelho_final_teste.xlsx")
    
    # Tentar validar estrutura
    import openpyxl
    try:
        wb = openpyxl.load_workbook('espelho_final_teste.xlsx')
        ws = wb.active
        
        print(f"\n   Sheet name: {ws.title}")
        print(f"   Primeira célula L9: {ws['L9'].value}")
        print(f"   Formato L9: {ws['L9'].number_format}")
        print(f"   Primeira célula K9: {ws['K9'].value}")
        
        # Contar Sunday vs weekday jornadacell
        sunday_empty = 0
        weekday_filled = 0
        for row in range(9, 40):
            dia = ws[f'B{row}'].value
            jornada = ws[f'K{row}'].value
            if dia and 'domingo'in str(dia).lower() and not jornada:
                sunday_empty += 1
            if dia and 'segunda' in str(dia).lower() and jornada:
                weekday_filled += 1
        
        print(f"   Domingos com Jornada vazia: {sunday_empty}")
        print(f"   Segundas com Jornada preenchida: {weekday_filled}")
        
    except Exception as e:
        print(f"   Erro ao validar: {e}")
else:
    print(f"\n❌ ERRO: Não é um ZIP válido")
    print(f"   Primeiros 200 bytes: {r_export.content[:200]}")
    sys.exit(1)

print("\n✅ Teste concluído!")
