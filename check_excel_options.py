import openpyxl

wb = openpyxl.load_workbook('espelho_final_opcoes.xlsx')

print('✅ OPÇÕES DO EXCEL CONFIGURADAS')
print()
print('Configurações de Cálculo:')
print(f'  Modo de Cálculo: {wb.calcPr.calcMode}')
print(f'  Data 1904: {wb.date1904}')

print()
print('Propriedades do Workbook:')
print(f'  Title: {wb.properties.title}')
print(f'  Subject: {wb.properties.subject}')

print()
print('Verificação de Fórmulas:')
ws = wb.active
formula_count = 0
for row in ws.iter_rows():
    for cell in row:
        if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
            formula_count += 1

print(f'  Total de fórmulas: {formula_count}')
print(f'  Exemplo (L10): {ws["L10"].value}')
print(f'  Exemplo (M10): {ws["M10"].value}')

print()
print('✅ ARQUIVO PRONTO COM TODAS AS OPÇÕES!')
