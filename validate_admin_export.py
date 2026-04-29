#!/usr/bin/env python
# -*- coding: utf-8 -*-
import openpyxl

wb = openpyxl.load_workbook('espelho_admin_final.xlsx')
ws = wb.active

print("VALIDAÇÃO FINAL DO ARQUIVO")
print("=" * 60)

print("\n1. COLUNAS L E M - FORMATO HORA")
for row in range(9, 20):
    hn_val = ws[f'L{row}'].value
    if hn_val and isinstance(hn_val, str) and hn_val.startswith('='):
        hn_fmt = ws[f'L{row}'].number_format
        he_fmt = ws[f'M{row}'].number_format
        print(f"   L{row}: {hn_fmt}")
        print(f"   M{row}: {he_fmt}")
        if hn_fmt == 'hh:mm:ss' and he_fmt == 'hh:mm:ss':
            print("   ✅ FORMATOS CORRETOS!")
        else:
            print("   ❌ Formatos incorretos")
        break

print("\n2. COLUNA K - JORNADA EM FINAIS DE SEMANA")
saturday_ok = False
sunday_ok = False

for row in range(9, 40):
    dia = ws[f'B{row}'].value
    jornada = ws[f'K{row}'].value
    
    if dia:
        if 'sábado' in str(dia).lower() and not saturday_ok:
            if not jornada:
                print(f"   Sábado ({ws[f'A{row}'].value}): vazio ✅")
                saturday_ok = True
            else:
                print(f"   Sábado ({ws[f'A{row}'].value}): {jornada} ❌")
                saturday_ok = False
        elif 'domingo' in str(dia).lower() and not sunday_ok:
            if not jornada:
                print(f"   Domingo ({ws[f'A{row}'].value}): vazio ✅")
                sunday_ok = True
            else:
                print(f"   Domingo ({ws[f'A{row}'].value}): {jornada} ❌")
                sunday_ok = False

print("\n" + "=" * 60)
if hn_fmt == 'hh:mm:ss' and he_fmt == 'hh:mm:ss' and saturday_ok and sunday_ok:
    print("✅ TODAS AS MUDANÇAS FORAM APLICADAS COM SUCESSO!")
else:
    print("❌ Algumas mudanças ainda não foram aplicadas")
print("=" * 60)
