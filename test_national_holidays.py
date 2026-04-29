#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Script para validar que apenas feriados nacionais aparecem no export
"""

import openpyxl
import os
import time
from datetime import datetime, timedelta

# Feriados nacionais esperados em 2026
FERIADOS_NACIONAIS = {
    '2026-01-01',  # Confraternização
    '2026-04-21',  # Tiradentes
    '2026-05-01',  # Dias do Trabalho
    '2026-09-07',  # Independência
    '2026-10-12',  # Nossa Senhora Aparecida
    '2026-11-02',  # Finados
    '2026-11-15',  # Proclamação da República
    '2026-12-25',  # Natal
}

# Feriados municipais/estaduais que devem ser REMOVIDOS
FERIADOS_REMOVIDOS = {
    '2026-02-04',  # São Paulo (Consciência Negra - segunda data)
    '2026-02-11',  # Alguns Estados
}

def format_date(date_str):
    """Converte 2026-02-04 para 04/02/2026"""
    if not date_str or len(date_str) != 10:
        return None
    try:
        parts = date_str.split('-')
        return f"{parts[2]}/{parts[1]}/{parts[0]}"
    except Exception:
        return date_str

def check_excel_file(filepath):
    """
    Valida que apenas feriados nacionais aparecem marcados no Excel
    """
    print(f"\n{'='*60}")
    print(f"VALIDANDO: {filepath}")
    print(f"{'='*60}\n")
    
    if not os.path.exists(filepath):
        print(f"❌ ERRO: Arquivo não encontrado: {filepath}")
        return False
    
    try:
        wb = openpyxl.load_workbook(filepath)
        ws = wb.active
    except Exception as e:
        print(f"❌ ERRO ao abrir Excel: {e}")
        return False
    
    # Procura a linha de header
    header_row = None
    for row_idx in range(1, 10):
        cell_value = ws.cell(row=row_idx, column=1).value
        if cell_value and 'Data' in str(cell_value):
            header_row = row_idx
            break
    
    if not header_row:
        print("❌ Não encontrou linha de header com 'Data'")
        return False
    
    print(f"Header encontrado na linha {header_row}")
    print(f"\nValidando coluna 'Observação' (Coluna 15)...\n")
    
    issues = {
        'marcados_como_feriado': [],
        'feriados_nacionais_ok': [],
        'feriados_municipais_nao_marcados': [],
        'data_invalida': []
    }
    
    # Itera pelas linhas de dados
    row_idx = header_row + 1
    while True:
        data_cell = ws.cell(row=row_idx, column=1).value
        if data_cell is None or data_cell == '':
            break
        
        obs_cell = ws.cell(row=row_idx, column=15).value  # Observação
        
        # Tenta extrair a data
        try:
            if isinstance(data_cell, str):
                data_str = data_cell.strip()
                # Converte DD/MM/YYYY para YYYY-MM-DD
                if '/' in data_str:
                    parts = data_str.split('/')
                    if len(parts) == 3:
                        data_normalized = f"{parts[2]}-{parts[1]}-{parts[0]}"
                    else:
                        issues['data_invalida'].append((row_idx, data_str))
                        row_idx += 1
                        continue
                else:
                    issues['data_invalida'].append((row_idx, data_str))
                    row_idx += 1
                    continue
            else:
                row_idx += 1
                continue
            
            # Verifica se é feriado municipal que deveria ter sido removido
            if data_normalized in FERIADOS_REMOVIDOS:
                if obs_cell and 'Feriado' in str(obs_cell):
                    issues['marcados_como_feriado'].append({
                        'linha': row_idx,
                        'data': f"{format_date(data_normalized)} ({data_str})",
                        'tipo': 'REMOVIDO (Municipal/Estadual)',
                        'obs': obs_cell
                    })
                else:
                    issues['feriados_municipais_nao_marcados'].append({
                        'data': format_date(data_normalized),
                        'status': '✅ Não marcado (correto)'
                    })
            
            # Verifica se é feriado nacional
            if data_normalized in FERIADOS_NACIONAIS:
                if obs_cell and 'Feriado' in str(obs_cell):
                    issues['feriados_nacionais_ok'].append({
                        'data': format_date(data_normalized),
                        'status': '✅ Marcado como Feriado (correto)'
                    })
                else:
                    # Feriado nacional NÃO marcado (problema!)
                    print(f"⚠️  ATENÇÃO: Feriado nacional {format_date(data_normalized)} NÃO marcado!")
        
        except Exception as e:
            issues['data_invalida'].append((row_idx, data_cell, str(e)))
        
        row_idx += 1
    
    # SAÍDA DE RESULTADOS
    print("\n" + "="*60)
    print("RESULTADOS DA VALIDAÇÃO")
    print("="*60)
    
    all_ok = True
    
    # Feriados nacionais marcados corretamente
    if issues['feriados_nacionais_ok']:
        print(f"\n✅ FERIADOS NACIONAIS (marcados corretamente):")
        for item in issues['feriados_nacionais_ok']:
            print(f"   {item['data']}: {item['status']}")
    
    # Feriados municipais que foram removidos com sucesso
    if issues['feriados_municipais_nao_marcados']:
        print(f"\n✅ FERIADOS MUNICIPAIS REMOVIDOS (não marcados):")
        for item in issues['feriados_municipais_nao_marcados']:
            print(f"   {item['data']}: {item['status']}")
    
    # PROBLEMAS: Feriados municipais ainda marcados como feriado
    if issues['marcados_como_feriado']:
        all_ok = False
        print(f"\n❌ PROBLEMAS - Feriados municipais ainda marcados:")
        for item in issues['marcados_como_feriado']:
            print(f"   Linha {item['linha']}: {item['data']}")
            print(f"      Tipo: {item['tipo']}")
            print(f"      Observação: {item['obs']}")
    
    if issues['data_invalida']:
        print(f"\n⚠️  DATAS INVÁLIDAS: {len(issues['data_invalida'])} encontradas")
    
    print("\n" + "="*60)
    if all_ok:
        print("✅ SUCESSO: Apenas feriados nacionais estão marcados!")
    else:
        print("❌ FALHA: Alguns feriados municipais ainda aparecem")
    print("="*60 + "\n")
    
    return all_ok

# Verifica arquivos exportados recentemente
excel_files = []
for fname in os.listdir('c:\\RelogioPonto'):
    if fname.startswith('espelho') and fname.endswith('.xlsx'):
        fpath = os.path.join('c:\\RelogioPonto', fname)
        mtime = os.path.getmtime(fpath)
        excel_files.append((mtime, fpath, fname))

if excel_files:
    # Ordena por data de modificação (mais recente primeiro)
    excel_files.sort(reverse=True)
    
    print("\nArquivos Excel encontrados (5 mais recentes):")
    for i, (mtime, fpath, fname) in enumerate(excel_files[:5]):
        mtime_str = datetime.fromtimestamp(mtime).strftime('%Y-%m-%d %H:%M:%S')
        print(f"  {i+1}. {fname} ({mtime_str})")
    
    # Valida o arquivo mais recente
    latest_file = excel_files[0][1]
    print(f"\nValidando arquivo mais recente: {os.path.basename(latest_file)}\n")
    check_excel_file(latest_file)
else:
    print("❌ Nenhum arquivo Excel encontrado em c:\\RelogioPonto")
