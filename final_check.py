import openpyxl
import os

print('✅ VALIDAÇÃO DO ARQUIVO EXCEL COM OPÇÕES')
print()

# Verificar arquivo
file_path = 'espelho_final_opcoes.xlsx'
if os.path.exists(file_path):
    file_size = os.path.getsize(file_path)
    print(f'Arquivo: {file_path}')
    print(f'Tamanho: {file_size} bytes')
    print()
    
    # Carregar workbook
    try:
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active
        
        print('✅ Arquivo Excel válido!')
        print()
        print('Conteúdo:')
        print(f'  - Sheet name: {ws.title}')
        print(f'  - Título (A1): {ws["A1"].value}')
        print(f'  - Primeiro dia (A9-A10): {ws["A10"].value}')
        print()
        
        # Contar fórmulas
        formulas = []
        for row in ws.iter_rows():
            for cell in row:
                if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                    formulas.append(cell.coordinate)
        
        print(f'  - Total de fórmulas: {len(formulas)}')
        if formulas:
            print(f'  - Primeiras 3 fórmulas:')
            for i, f in enumerate(formulas[:3]):
                cell = ws[f]
                print(f'    {f}: {cell.value}')
        
        print()
        print('✅ EXPORT COM OPÇÕES CONFIGURADAS COM SUCESSO!')
        
    except Exception as e:
        print(f'❌ Erro ao carregar: {e}')
else:
    print(f'❌ Arquivo não encontrado: {file_path}')
