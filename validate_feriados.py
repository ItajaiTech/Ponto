#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Valida feriados nacionais no export Excel"""

import openpyxl
import os
from datetime import datetime

# Feriados nacionais de 2026
NACIONAIS = {
    '2026-01-01', '2026-04-21', '2026-05-01', '2026-09-07',
    '2026-10-12', '2026-11-02', '2026-11-15', '2026-12-25'
}

# Feriados que foram REMOVIDOS (municipais/estaduais)
REMOVIDOS = {'2026-02-04', '2026-02-11'}

def check_excel():
    # Encontra arquivo mais recente
    files = [(os.path.getmtime(f'c:\\RelogioPonto\\{f}'), f) 
             for f in os.listdir('c:\\RelogioPonto') 
             if f.startswith('espelho') and f.endswith('.xlsx')]
    
    if not files:
        print("Nenhum arquivo Excel encontrado")
        return
    
    latest = max(files)[1]
    print(f"Validando: {latest}\n")
    
    wb = openpyxl.load_workbook(f'c:\\RelogioPonto\\{latest}')
    ws = wb.active
    
    # Encontra header
    for row in range(1, 10):
        if ws.cell(row, 1).value and 'Data' in str(ws.cell(row, 1).value):
            header_row = row
            break
    
    print("="*60)
    print("VALIDACAO DE FERIADOS")
    print("="*60 + "\n")
    
    removidos_marcados = []
    nacionais_marcados = []
    
    # Itera pelas linhas
    for row in range(header_row + 1, 200):
        data_val = ws.cell(row, 1).value
        obs_val = ws.cell(row, 15).value
        
        if not data_val:
            break
        
        # Converte data para formato YYYY-MM-DD
        try:
            parts = str(data_val).split('/')
            data_key = f"{parts[2]}-{parts[1]}-{parts[0]}"
        except:
            continue
        
        is_feriado = obs_val and 'Feriado' in str(obs_val)
        
        # Verifica municipais que devem estar VAZIOS
        if data_key in REMOVIDOS:
            if is_feriado:
                removidos_marcados.append(f"{data_val} - ERRO: Marcado como feriado!")
            else:
                print(f"✓ {data_val} (municipal) - nao marcado")
        
        # Verifica nacionais que devem estar PREENCHIDOS
        if data_key in NACIONAIS:
            if is_feriado:
                nacionais_marcados.append(f"{data_val} - OK")
                print(f"✓ {data_val} (nacional) - marcado como feriado")
            else:
                print(f"! {data_val} (nacional) - NAO marcado!")
    
    print("\n" + "="*60)
    if removidos_marcados:
        print("ERROS ENCONTRADOS:")
        for item in removidos_marcados:
            print(f"  X {item}")
        print("\nSTATUS: FALHA - municipais ainda aparecem")
    else:
        print("STATUS: SUCESSO - Apenas nacionais aparecem!")
    print("="*60)

if __name__ == '__main__':
    check_excel()
