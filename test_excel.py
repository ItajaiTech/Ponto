import openpyxl

wb = openpyxl.load_workbook('espelho_novo.xlsx')
ws = wb.active

def get_color_info(cell):
    """Extract color info from cell"""
    fill_color = "sem cor"
    font_color = "sem cor"
    
    if cell.fill and cell.fill.start_color:
        fill_color = f"Fill: {cell.fill.start_color.rgb if hasattr(cell.fill.start_color, 'rgb') else cell.fill.start_color.index}"
    
    if cell.font and cell.font.color:
        if hasattr(cell.font.color, 'rgb'):
            font_color = f"Font RGB: {cell.font.color.rgb}"
        elif hasattr(cell.font.color, 'theme'):
            font_color = f"Font Theme: {cell.font.color.theme}"
        else:
            font_color = f"Font: {cell.font.color.index}"
    
    return f"[{fill_color}] [{font_color}]"

print('=== TÍTULO (Linha 1) ===')
print(f'Valor: {ws["A1"].value}')
print(f'Cores: {get_color_info(ws["A1"])}')
print(f'Font bold: {ws["A1"].font.bold}, size: {ws["A1"].font.size}')

print()
print('=== CABEÇALHOS (Linha 8) ===')
print(f'Col 1 (Data): cores: {get_color_info(ws.cell(8, 1))}')

print()
print('=== DIA COM FERIADO (Linha 12 = dia 4) ===')
print(f'Col 15 (Observação): valor={ws.cell(12, 15).value}, cores: {get_color_info(ws.cell(12, 15))}')


