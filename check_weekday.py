#!/usr/bin/env python
# -*- coding: utf-8 -*-
import requests
import openpyxl

session = requests.Session()
session.post('http://localhost:5050/', data={'cpf': '00000000000', 'senha': 'Pont0!2024@Admin#Secure'}, allow_redirects=True)
session.get('http://localhost:5050/admin')
r = session.get('http://localhost:5050/admin/espelho/export?tipo=excel&mes=02&ano=2026', timeout=15)
with open('espelho_correto.xlsx', 'wb') as f:
    f.write(r.content)

wb = openpyxl.load_workbook('espelho_correto.xlsx')
ws = wb.active

print("\n" + "=" * 70)
print("VERIFICAÇÃO FINAL DO EXPORT DO ADMIN")
print("=" * 70)

# Verificar formatações
print("\nCOLUNAS L E M - FORMATAÇÃO HORA")
for row in range(9, 15):
    l_fmt = ws[f'L{row}'].number_format
    m_fmt = ws[f'M{row}'].number_format
    if l_fmt and 'h' in l_fmt.lower():
        print(f"   L{row}: {l_fmt} OK")
        print(f"   M{row}: {m_fmt} OK")
        break
    elif row == 9:
        print(f"   L{row}: {l_fmt} - INCORRETO")
        print(f"   M{row}: {m_fmt} - INCORRETO")

# Verificar Jornada
print("\nCOLUNA K - JORNADA EM FINAIS DE SEMANA")
for row in range(9, 40):
    dia = ws[f'B{row}'].value
    jornada = ws[f'K{row}'].value
    
    if dia:
        if 'Domingo' in str(dia):
            status = "vazio" if not jornada else f"preenchido"
            print(f"   Domingo ({ws[f'A{row}'].value}): {status}")
            break

print("\n" + "=" * 70 + "\n")

