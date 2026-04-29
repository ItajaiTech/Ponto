#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
Teste direto da formatação de Excel sem passar pelo Flask
"""
import sqlite3
import json
from datetime import datetime
import calendar
import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

print("Simulando exportação Excel com formatação...")

# Carregar dados do banco
conn = sqlite3.connect('ponto.db')
cursor = conn.cursor()

# Pegar dados do usuário
cursor.execute("SELECT * FROM registros WHERE usuario_id = 1 ORDER BY data_hora")
registros = cursor.fetchall()
conn.close()

print(f"Total de registros encontrados: {len(registros)}")

# Carregar feriados
try:
    with open('feriados_personalizados.json', 'r', encoding='utf-8') as f:
        feriados = set(json.load(f))
except:
    feriados = set()
    print("Arquivo de feriados não encontrado")

# Criar workbook
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Espelho de Ponto"

# Configurar opções
wb.date1904 = False

# Setup inicial (título, headers, etc)
header_row = 8

# Dados do usuário
mes = 2
ano = 2026

# Dia da semana em português
dias_semana_pt = ["segunda", "terça", "quarta", "quinta", "sexta", "sábado", "domingo"]

# Mapa de abonos
abonos_map = {}
for r in registros:
    if r[2] == 'abono':  # índice 2 = tipo
        data_str = r[3][:10]  # índice 3 = data_hora
        abonos_map[data_str] = 1

# Setup borders
thin_border = Border(
    left=Side(border_style='thin'),
    right=Side(border_style='thin'),
    top=Side(border_style='thin'),
    bottom=Side(border_style='thin')
)

font_normal = Font(size=10)
font_green = Font(color='FF00B050', size=10)

print(f"\nGerando dados para {mes}/{ano}...")

# Preencher dados
dias_mes = calendar.monthrange(int(ano), int(mes))[1]

for dia in range(1, min(10, dias_mes + 1)):  # Apenas primeiros 9 dias para teste
    data = datetime(int(ano), int(mes), dia)
    row_idx = header_row + dia
    data_str = f"{ano}-{mes:02d}-{dia:02d}"
    
    # Data
    data_cell = ws.cell(row=row_idx, column=1, value=data.strftime('%d/%m/%Y'))
    data_cell.border = thin_border
    
    # Dia da semana
    dia_cell = ws.cell(row=row_idx, column=2, value=dias_semana_pt[data.weekday()])
    dia_cell.border = thin_border
    
    # Abono
    abono_cell = ws.cell(row=row_idx, column=10)
    tem_abono = data_str in abonos_map
    if tem_abono:
        abono_cell.value = '09:00'
        abono_cell.font = font_green
    abono_cell.border = thin_border
    
    # COLUNA K: Jornada (vazia em sábado/domingo)
    jornada_cell = ws.cell(row=row_idx, column=11)
    jornada_cell.border = thin_border
    jornada_cell.alignment = Alignment(horizontal='center')
    jornada_cell.font = font_normal
    
    is_weekend = data.weekday() in (5, 6)  # 5=sábado, 6=domingo
    if not is_weekend:
        jornada_cell.value = '09:00:00'
    
    print(f"  {dia:02d}/{mes:02d}: {dias_semana_pt[data.weekday()]:8} - Jornada: {'(vazia)' if is_weekend else '09:00:00'}")
    
    # Coluna L: H.N. - Com formato hora
    col_letter_C = get_column_letter(3)
    col_letter_D = get_column_letter(4)
    col_letter_E = get_column_letter(5)
    col_letter_F = get_column_letter(6)
    
    hn_formula = f"=SUM({col_letter_F}{row_idx}-{col_letter_E}{row_idx},{col_letter_D}{row_idx}-{col_letter_C}{row_idx})"
    hn_cell = ws.cell(row=row_idx, column=12, value=hn_formula)
    hn_cell.border = thin_border
    hn_cell.alignment = Alignment(horizontal='center')
    hn_cell.number_format = 'hh:mm:ss'  # Formato hora
    
    # Coluna M: H.E. - Com formato hora
    col_letter_J = get_column_letter(10)
    col_letter_K = get_column_letter(11)
    col_letter_L = get_column_letter(12)
    
    he_formula = f"=SUM({col_letter_J}{row_idx}+{col_letter_L}{row_idx})-{col_letter_K}{row_idx}"
    he_cell = ws.cell(row=row_idx, column=13, value=he_formula)
    he_cell.border = thin_border
    he_cell.alignment = Alignment(horizontal='center')
    he_cell.number_format = 'hh:mm:ss'  # Formato hora

# Salvar
print("\nSalvando arquivo...")
wb.save('teste_formatacao_direta.xlsx')
print("✅ Arquivo salvo: teste_formatacao_direta.xlsx\n")

# Validar
print("Validando...")
wb = openpyxl.load_workbook('teste_formatacao_direta.xlsx')
ws = wb.active

print("\n=== VERIFICAÇÃO ===")
print("\nColunas L e M (H.N. e H.E.) - FORMATAÇÃO HORA:")
for row in range(9, 12):
    hn_fmt = ws[f'L{row}'].number_format
    he_fmt = ws[f'M{row}'].number_format
    print(f"  Linha {row}: H.N.={hn_fmt:15} | H.E.={he_fmt}")

print("\nColuna K (Jornada) - VERIFICAR FINAIS DE SEMANA:")
for row in range(9, 18):
    data = ws[f'A{row}'].value
    dia = ws[f'B{row}'].value
    jornada = ws[f'K{row}'].value
    if data:
        is_weekend = dia.lower() in ('sábado', 'domingo', 'sabado')
        status = "✅ vazio" if (is_weekend and not jornada) else "✅ preenchido" if (not is_weekend and jornada == '09:00:00') else f"❌ {jornada}"
        print(f"  {data} ({dia:8}): {status}")

print("\n✅ Teste concluído!")
