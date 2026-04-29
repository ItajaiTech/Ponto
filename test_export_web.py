#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
Teste de export Excel através do Flask com manutenção de sessão
"""
import requests
from requests.adapters import HTTPAdapter
from urllib3.util.retry import Retry
import openpyxl
import time

def criar_session_com_retry():
    """Criar sessão com retry automático"""
    session = requests.Session()
    retry = Retry(
        total=3,
        backoff_factor=0.5,
        status_forcelist=(502, 503, 504)
    )
    adapter = HTTPAdapter(max_retries=retry)
    session.mount('http://', adapter)
    session.mount('https://', adapter)
    return session

print("=" * 60)
print("TESTE DE EXPORT EXCEL DO SISTEMA ADMIN")
print("=" * 60)

session = criar_session_com_retry()

try:
    # Step 1: Login
    print("\n1️⃣  Fazendo login...")
    login_url = 'http://localhost:5050/'
    login_data = {
        'cpf': '00000000000',
        'senha': 'Pont0!2024@Admin#Secure'
    }
    
    r = session.post(login_url, data=login_data, timeout=10, allow_redirects=True)
    print(f"   Status: {r.status_code}")
    print(f"   URL final: {r.url}")
    print(f"   Cookies: {dict(session.cookies)}")
    
    # Verificar se chegou no admin
    if '/admin' in r.url or 'espelho' in r.url:
        print("   ✅ Login redirecionou para admin")
    elif r.status_code == 200:
        print("   ✅ Login OK")
    else:
        print("   ❌ Erro no login!")
    
    # Step 2: Aguardar um pouco
    time.sleep(1)
    
    # Step 3: Acessar página admin para garantir sessão
    print("\n2️⃣  Acessando página admin...")
    admin_url = 'http://localhost:5050/admin'
    r = session.get(admin_url, timeout=10)
    print(f"   Status: {r.status_code}")
    
    time.sleep(0.5)
    
    # Step 4: Fazer export
    print("\n3️⃣  Exportando arquivo Excel...")
    export_url = 'http://localhost:5050/admin/espelho/export?tipo=excel&mes=02&ano=2026'
    
    r = session.get(export_url, timeout=10)
    print(f"   Status: {r.status_code}")
    print(f"   Content-Type: {r.headers.get('content-type', 'N/A')}")
    print(f"   Tamanho: {len(r.content)} bytes")
    
    if r.content[:2] == b'PK':
        print("   ✅ É um arquivo ZIP válido (Excel)")
        
        # Salvar arquivo
        filename = 'espelho_admin_export.xlsx'
        with open(filename, 'wb') as f:
            f.write(r.content)
        print(f"   ✅ Arquivo salvo: {filename}")
        
        # Validar estrutura
        print("\n4️⃣  Verificando estrutura do Excel...")
        wb = openpyxl.load_workbook(filename)
        ws = wb.active
        
        print(f"   Sheet: {ws.title}")
        print(f"   Título (A1): {ws['A1'].value}")
        
        # Verificar formatações
        print("\n5️⃣  Verificando formatações...")
        
        # Procurar formulas nas colunas L e M
        found_hn = False
        found_he = False
        hn_format = None
        he_format = None
        
        for row in range(9, 25):
            if ws[f'L{row}'].value and isinstance(ws[f'L{row}'].value, str) and ws[f'L{row}'].value.startswith('='):
                found_hn = True
                hn_format = ws[f'L{row}'].number_format
                print(f"   L{row}: Fórmula encontrada")
                print(f"         Valor: {ws[f'L{row}'].value}")
                print(f"         Formato: {hn_format}")
                break
        
        for row in range(9, 25):
            if ws[f'M{row}'].value and isinstance(ws[f'M{row}'].value, str) and ws[f'M{row}'].value.startswith('='):
                found_he = True
                he_format = ws[f'M{row}'].number_format
                print(f"   M{row}: Fórmula encontrada")
                print(f"         Valor: {ws[f'M{row}'].value}")
                print(f"         Formato: {he_format}")
                break
        
        # Verificar Jornada (coluna K)
        print("\n6️⃣  Verificando Jornada (coluna K)...")
        for row in range(9, 20):
            dia = ws[f'B{row}'].value
            jornada = ws[f'K{row}'].value
            if dia:
                if 'sábado' in str(dia).lower() or 'domingo' in str(dia).lower():
                    status = "✅ vazio" if not jornada else f"❌ preenchido ({jornada})"
                else:
                    status = "✅ preenchido" if jornada else "❌ vazio"
                print(f"   {dia:15} - {status}")
        
        print("\n" + "=" * 60)
        print("✅ TESTE CONCLUÍDO COM SUCESSO!")
        print("=" * 60)
        
    else:
        print("   ❌ Não é um ZIP válido")
        print(f"   Primeiros bytes: {r.content[:100]}")
        print(f"   Texto: {r.text[:200]}")

except Exception as e:
    print(f"\n❌ Erro: {e}")
    import traceback
    traceback.print_exc()
