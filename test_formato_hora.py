#!/usr/bin/env python
# -*- coding: utf-8 -*-
import requests
import openpyxl

# Login
session = requests.Session()
login_data = {'cpf': '00000000000', 'senha': 'Pont0!2024@Admin#Secure'}
r = session.post('http://localhost:5050/admin', data=login_data)
print(f"Login: {r.status_code}")

# Export Excel
export_url = 'http://localhost:5050/admin/espelho/export?tipo=excel&mes=2&ano=2026'
r = session.get(export_url)
print(f"Export: {r.status_code}")

# Salvar arquivo
with open('teste_formato_hora.xlsx', 'wb') as f:
    f.write(r.content)
print(f"Arquivo salvo: {len(r.content)} bytes")

# Verificar formatação
wb = openpyxl.load_workbook('teste_formato_hora.xlsx')
ws = wb.active

print("\n=== VERIFICAÇÃO DA FORMATAÇÃO ===")
print("\nSemana de 02/02/2026 (segunda):")
for row in range(9, 11):  # Segunda e terça
    data = ws[f'A{row}'].value
    dia = ws[f'B{row}'].value
    hn = ws[f'L{row}'].value
    he = ws[f'M{row}'].value
    jornada = ws[f'K{row}'].value
    hn_fmt = ws[f'L{row}'].number_format
    he_fmt = ws[f'M{row}'].number_format
    print(f"  {data} ({dia})")
    print(f"    H.N. (L{row}): {hn} | Formato: {hn_fmt}")
    print(f"    H.E. (M{row}): {he} | Formato: {he_fmt}")
    print(f"    Jornada (K{row}): {jornada}")

print("\nFim de semana (04/02/2026 = quarta, 05/02/2026 = quinta):")
for row in range(12, 14):
    data = ws[f'A{row}'].value
    dia = ws[f'B{row}'].value
    jornada = ws[f'K{row}'].value
    print(f"  {data} ({dia}) - Jornada: {jornada}")

print("\nSábado e domingo (07-08/02):")
for row in range(15, 17):
    data = ws[f'A{row}'].value
    dia = ws[f'B{row}'].value
    jornada = ws[f'K{row}'].value
    hn = ws[f'L{row}'].value
    hn_fmt = ws[f'L{row}'].number_format
    print(f"  {data} ({dia}) - Jornada: {jornada} (vazia?) | H.N. formato: {hn_fmt}")

print("\n✅ Verificação concluída!")
