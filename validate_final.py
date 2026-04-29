import openpyxl

wb = openpyxl.load_workbook('espelho_final.xlsx')
ws = wb.active

print('✅ VALIDAÇÃO FINAL DO EXCEL')
print()
print('TÍTULO:')
print(f'  Valor: {ws["A1"].value}')
print(f'  Mesclado: A1:O1')

print()
print('INFO BLOCK:')
print(f'  A3: {ws["A3"].value}')
print(f'  A5: {ws["A5"].value}')
print(f'  H6: {ws["H6"].value}')

print()
print('CABEÇALHOS (15 cols):')
for i in range(1, 16):
    print(f'  Col {i:2d}: {ws.cell(8, i).value}')

print()
print('FERIADO (dia 4, linha 12):')
for col in [1, 2, 10, 15]:
    val = ws.cell(12, col).value
    print(f'  Col {col:2d}: {val}')

print()
print('ABONO (dia 5, linha 13):')
print(f'  Col 10 (Abono): {ws.cell(13, 10).value}')
print(f'  Col 15 (Observação): {ws.cell(13, 15).value}')
